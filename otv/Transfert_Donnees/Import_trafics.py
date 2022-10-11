# -*- coding: utf-8 -*-
'''
Created on 27 juin 2019

@author: martin.schoreisz

module d'importation des données de trafics forunies par les gestionnaires
'''

import pandas as pd
import geopandas as gp
import numpy as np
import os, re, csv,statistics,filecmp, warnings
from pathlib import PureWindowsPath
from unidecode import unidecode
import datetime as dt
from geoalchemy2 import Geometry,WKTElement
from shapely.geometry import LineString, Point
from shapely.ops import transform
from collections import Counter
from openpyxl import load_workbook

import Connexion_Transfert as ct
from Donnees_horaires import (comparer2Sens,verifValiditeFichier, concatIndicateurFichierHoraire, SensAssymetriqueError,
                              verifNbJoursValidDispo, tmjaDepuisHoraire, periodeDepuisHoraire, attributsHoraire, 
                              mensuelDepuisHoraire)
from Donnees_sources import FIM, MHCorbin, DataSensError, PasAssezMesureError
from Integration_nouveau_comptage import (corresp_nom_id_comptag, scinderComptagExistant, creer_comptage, structureBddOld2NewForm,
                                          geomFromIdComptagCommunal, creerCompteur)
from Import_export_comptage import (compteur_existant_bdd, insererSchemaComptage)
import Outils as O
from Params.Mensuel import dico_mois, renommerMois
from Params.Bdd_OTV import (attBddCompteur, nomConnBddOtv, schemaComptage, schemaComptageAssoc, tableComptage, 
                            tableIndicAgrege, tableIndicHoraire, attrComptage, attrIndicAgrege, enumIndicateur)   
from Params.DonneesGestionnaires import *
from Params.DonneesVitesse import valeursVmaAdmises, correspVmaVlVmaPl  
              
              
class Comptage():
    """
    classe comprenant les attributs et méthode commune à tous les comptages de tous les departements
    """
    def __init__(self, fichier):
        self.fichier=fichier
        
    def ouvrir_csv(self, delimiter=';', skiprows=0):
        """
        ouvrir un fichier csv en supprimant les caracteres accentué et en remplaçant les valeurs vides par np.naN
        """
        with open(self.fichier, newline='') as csvfile: 
            reader = csv.reader(csvfile, delimiter=delimiter)
            fichier = ([[re.sub(('é|è|ê'),'e',a) for a in row] for i, row in enumerate(reader) if i>=skiprows])
            df = pd.DataFrame(data=fichier[1:], columns=fichier[0])
            df.replace('', np.NaN, inplace=True)
        return  df
    
        
    def maj_geom(self, schema, table, nom_table_ref,nom_table_pr,dep=False):
        """
        mettre à jour les lignes de geom null
        en entree : 

            schema : string nom du schema de la table
            table : string : nom de la table
            nom_table_ref : string : nom de table contenant le referntiel (schema qualified)
            nom_table_pr : string : nom de table contenant les pr (schema qualified)
            dep : string : code departement sur 2 chiffres
        """
        if dep : 
            rqt_geom=f""" update {schema}.{table}
              set geom=(select geom_out  from comptage.geoloc_pt_comptag('{nom_table_ref}','{nom_table_pr}',id_comptag))
              where dep='{dep}' and geom is null"""
            rqt_attr=f""" update {schema}.{table}
              set src_geo='pr+abs', x_l93=round(st_x(geom)::numeric,3), y_l93=round(st_y(geom)::numeric,3)
              where dep='{dep}' and src_geo is null and geom is not null"""
        else :
            rqt_geom=f""" update {schema}.{table}
              set geom=(select geom_out  from comptage.geoloc_pt_comptag('{nom_table_ref}','{nom_table_pr}',id_comptag))
              where geom is null""" 
            rqt_attr=f""" update {schema}.{table}
              set src_geo='pr+abs', x_l93=round(st_x(geom)::numeric,3), y_l93=round(st_y(geom)::numeric,3)
              where src_geo is null and geom is not null"""   
              
        with ct.ConnexionBdd(nomConnBddOtv) as c:
                c.sqlAlchemyConn.execute(rqt_geom)
                c.sqlAlchemyConn.execute(rqt_attr)


    def filtrer_periode_ponctuels(self):
        """
        filtrer des periodesde vacances
        """
        #filtrer pt comptage pendant juillet aout
        self.df_attr=self.df_attr.loc[self.df_attr.apply(lambda x : x['debut_periode'].month not in [7,8] and x['fin_periode'].month not in [7,8], 
                                                         axis=1)].copy()
    
    def donnees_existantes(self, table_linearisation_existante,table_cpt):
        """
        recuperer une linearisation existante et les points de comptages existants
        in : 
            table_linearisation_existante : string : schema-qualified table de linearisation de reference
            table_cpt : string : table des comptages
        """    
        with ct.ConnexionBdd(nomConnBddOtv) as c:
            lin_precedente=gp.GeoDataFrame.from_postgis(f'select * from {table_linearisation_existante}',c.sqlAlchemyConn, 
                                                    geom_col='geom',crs={'init': 'epsg:2154'})
            pt_cpt=pd.read_sql(f'select id_comptag,type_poste from comptage.{table_cpt}',c.sqlAlchemyConn)
            lin_precedente=lin_precedente.merge(pt_cpt, how='left', on='id_comptag')
        return lin_precedente
    
    def plus_proche_voisin_comptage_a_inserer(self,df, schema_temp,nom_table_temp,table_linearisation_existante,table_pr,table_cpt):
        """
        trouver si nouveau point de comptage est sur  un id_comptag linearise
        in:
            df : les données de comptage nouvelles, normalement c'est self.df_attr_insert (cf Comptage_Cd17 ou Compatge cd47
            schema_temp : string : nom du schema en bdd opur calcul geom, cf localiser_comptage_a_inserer
            nom_table_temp : string : nom de latable temporaire en bdd opur calcul geom, cf localiser_comptage_a_inserer
            table_linearisation_existante : string : schema-qualified table de linearisation de reference cf donnees_existantes
            table_pr : string : schema qualifyed nom de la table de reference des pr
            table_cpt : string : table des comptages
        """
        #mettre en forme les points a inserer
        points_a_inserer=self.localiser_comptage_a_inserer(df, schema_temp,nom_table_temp,table_linearisation_existante, table_pr)
        #recuperer les donnees existante
        lin_precedente=self.donnees_existantes(table_linearisation_existante, table_cpt)
        
        #ne conserver que le spoints a inserer pour lesquels il y a une geometrie
        points_a_inserer_geom=points_a_inserer.loc[~points_a_inserer.geom.isna()].copy()
        #recherche de la ligne la plus proche pour chaque point a inserer
        ppv=O.plus_proche_voisin(points_a_inserer_geom,lin_precedente[['id_ign','geom']],5,'id_comptag','id_ign')
        #verifier si il y a un id comptage sur la ligne la plus proche, ne conserver que les lignes où c'est le cas, et recuperer la geom de l'id_comptag_lin 
        #(les points issu de lin sans geom ne sont pas conserves, et on a besoin de passer les donnees en gdf)
        ppv_id_comptagLin=ppv.merge(lin_precedente[['id_ign','id_comptag','type_poste']].rename(columns={'id_comptag':'id_comptag_lin','type_poste':'type_poste_lin'}), on='id_ign')
        ppv_id_comptagLin=ppv_id_comptagLin.loc[~ppv_id_comptagLin.id_comptag_lin.isna()].copy().merge(self.existant[['geom','id_comptag']].rename(columns=
                                                {'geom':'geom_cpt_lin','id_comptag':'id_comptag_lin'}),on='id_comptag_lin')
        ppv_id_comptagLin_p=gp.GeoDataFrame(ppv_id_comptagLin.rename(columns={'id_ign':'id_ign_cpt_new'}),geometry=ppv_id_comptagLin.geom_cpt_lin)
        ppv_id_comptagLin_p.crs = {'init' :'epsg:2154'}
        #si il y a un id_comptag linearisation : trouver le troncon le plus proche de celui-ci
        ppv_total=O.plus_proche_voisin(ppv_id_comptagLin_p,lin_precedente[['id_ign','geom']],5,'id_comptag_lin','id_ign')
        ppv_final=ppv_total.merge(ppv_id_comptagLin_p,on='id_comptag_lin').merge(df[['id_comptag','type_poste']]
                    ,on='id_comptag').rename(columns={'id_ign':'id_ign_lin','type_poste':'type_poste_new'}).drop_duplicates()
        return ppv_final,points_a_inserer_geom
    
    def troncon_elemntaires(self,schema, table_graph,table_vertex,liste_lignes,id_name):    
        """
        a reprendre avec travail Vince
        """
    
    def corresp_old_new_comptag(self,schema_temp,nom_table_temp,table_linearisation_existante,
                                schema, table_graph,table_vertex,id_name,table_pr,table_cpt):
        """
         a reprendre avec travail Vince
        """
        def pt_corresp(id_ign_lin,id_ign_cpt_new,dico_corresp) : 
            if id_ign_cpt_new in dico_corresp[id_ign_lin] : 
                return True
            else : return False
           
        #verif que les colonnes necessaires sont presentes dans le fichier de bas
        ppv_final,points_a_inserer_geom=self.plus_proche_voisin_comptage_a_inserer(self.df_attr_insert, schema_temp,nom_table_temp,
                                                             table_linearisation_existante, table_pr,table_cpt)
        print('plus proche voisin fait')
        dico_corresp=self.troncon_elemntaires(schema, table_graph,table_vertex,ppv_final.id_ign_lin.tolist(),id_name)
        print('tronc elem fait')
        ppv_final['correspondance']=ppv_final.apply(lambda x : pt_corresp(x['id_ign_lin'],x['id_ign_cpt_new'],dico_corresp),axis=1)
        df_correspondance=ppv_final.loc[(ppv_final['correspondance']) & 
              (~ppv_final['id_comptag_lin'].isin(self.df_attr.id_comptag.tolist()))].copy()[['id_comptag_lin','type_poste_lin','id_comptag','type_poste_new']]
        return df_correspondance,points_a_inserer_geom
        
    
    def creer_valeur_txt_update(self, df, liste_attr):
        """
        a partir d'une df cree un tuple selon les vaelur que l'on va vouloir inserer dans la Bdd
        en entree : 
            df: df des donnees de base
            liste_attr : liste des attributs que l'on souhaite transferer dans la bdd (avec id_comptag)
        """
        valeurs_txt=str(tuple([ tuple([elem[i].replace('\'','_') if isinstance(i,str) else elem[i] for i in range(len(liste_attr)) ]) 
                               for elem in zip(*[df[a].tolist() for a in liste_attr])]))[1:-1]
        return valeurs_txt
    
    def update_bdd(self, schema, table, valeurs_txt,dico_attr_modif,identifiant='id_comptag',filtre=None):
        """
        mise à jour des id_comptag deja presents dans la base
        en entree : 
            schema : string nom du schema de la table
            table : string : nom de la table
            valeurs_txt : tuple des valeurs pour mise à jour, issu de creer_valeur_txt_update
            dico_attr_modif : dico de string avec en clé les nom d'attribut à mettre à jour, en value des noms des attributs source dans la df (ne pas mettre id_comptag,
                            garder les attributsdans l'ordre issu de creer_valeur_txt_update)
            identifiant : string : nom de la''tribut permettant la mise a jour par jointure
            filtre : string : condition de requete where a ajouter. le permeir 'and' ne doit pas etre ecrit
        """
        rqt_attr=','.join(f'{attr_b}=v.{attr_f}' for (attr_b,attr_f) in dico_attr_modif.items())
        attr_fichier=','.join(f'{attr_f}' for attr_f in dico_attr_modif.values())
        rqt_base=f'update {schema}.{table}  as c set {rqt_attr} from (values {valeurs_txt}) as v({identifiant},{attr_fichier}) where v.{identifiant}=c.{identifiant}'
        if filtre : 
            rqt_base=rqt_base+f' and {filtre}'
        with ct.ConnexionBdd(nomConnBddOtv) as c:
                c.sqlAlchemyConn.execute(rqt_base)

    def insert_bdd(self, schema, table, df, if_exists='append',geomType='POINT'):
        """
        insérer les données dans la bdd et mettre à jour la geometrie
        en entree : 
            schema : string nom du schema de la table
            table : string : nom de la table
        """
        if isinstance(df, gp.GeoDataFrame) : 
            if df.geometry.name!='geom':
                df=O.gp_changer_nom_geom(df, 'geom')
                df.geom=df.apply(lambda x : WKTElement(x['geom'].wkt, srid=2154), axis=1)
            with ct.ConnexionBdd(nomConnBddOtv) as c:
                df.to_sql(table,c.sqlAlchemyConn,schema=schema,if_exists=if_exists, index=False,
                          dtype={'geom': Geometry()} )
        elif isinstance(df, pd.DataFrame) : 
            with ct.ConnexionBdd(nomConnBddOtv) as c:
                df.to_sql(table,c.sqlAlchemyConn,schema=schema,if_exists=if_exists, index=False )
                
    def insererComptage(self, df):
        """
        spécialisation de la fonction insert_bdd pour le cas des comptages
        """
        self.insert_bdd('comptage', 'comptage',
                        df, 'append', 'POINT')
        
    def insererAgrege(self, df):
        """
        spécialisation de la fonction insert_bdd pour le cas des indicateurs agrege
        """
        self.insert_bdd('comptage', 'indic_agrege',
                        df, 'append', 'POINT')
        
    def insererMensuel(self, df):
        """
        spécialisation de la fonction insert_bdd pour le cas des indicateurs mensuel
        """
        self.insert_bdd('comptage', 'indic_mensuel',
                        df, 'append', 'POINT')
        
    def insererHoraire(self, df):
        """
        spécialisation de la fonction insert_bdd pour le cas des indicateurs horaire
        """
        self.insert_bdd('comptage', 'indic_horaire',
                        df, 'append', 'POINT')
                
class Comptage_cd64(Comptage):
    """
    premiere annee : 2020 : base fichier excel 
    ATTENTION : le fichier continet plusieurs annee, il faut donc le filtrer pour ne garder que les annees inconnues (2018, 2019, 2020)
    """
    def __init__(self,fichier, annee) : 
        """
        attributs : 
            fichier : raw string path complet
            annee : string : annee sur 4 caracteres
        """
        self.fichier=fichier
        self.annee=annee 
        
    def miseEnForme(self):
        """
        ouvrir le fichier, chnager le nom des attributs et ne conserver que les annees qui nous interesse
        in : 
            anneeMin : string : annee mini a conserver sur 4 caracteres
        """
        dfBrute=pd.read_excel(self.fichier, skiprows=1)
        dfBrute=dfBrute.rename(columns={c:unidecode(c).lower().replace(' ','_').replace('%', 'pc_') for c in dfBrute.columns})
        dfBrute['departement']=dfBrute['departement'].astype('str')
        dfBrute['type']=dfBrute['type'].apply(lambda x : 'permanent' if x=='Per' else 'tournant')
        dfBrute['route']=dfBrute['route'].apply(lambda x : O.epurationNomRoute(x.split()[1]))
        dfBrute['type_veh']='tv/pl'
        dfBrute['reseau']='RD'
        dfBrute['gestionnai']='CD64'
        dfBrute['concession']=False
        dfBrute['src_geo']='pr+abs_gestionnaire'
        dfBrute['obs_supl']=dfBrute.apply(lambda x : f'section : {x.section} ; indice : {x.indice} ; sens : {x.sens} ; localisation : {x.localisation}', axis=1)
        dfBrute['fictif']=False
        dfBrute['src_cpt']='convention gestionnaire'
        dfBrute['convention']=True
        dfBrute['sens_cpt']='double sens'
        dfBrute['fichier']=self.fichier
        dfBrute.rename(columns={'type':'type_poste','departement':'dep', 'prc':'pr', 'abc':'abs', 'mja_tv':'tmja', 'mja_pc_pl':'pc_pl' }, inplace=True)
        dfBrute['id_comptag']=dfBrute.apply(lambda x : f"{x.dep}-{x.route}-{x.pr}+{x['abs']}", axis=1)
        
        self.df_attr=dfBrute
    
    def filtrerAnnee(self,df, anneeMin):
        """
        filtrer une dfBrute selon les annees a garder
        in : 
            anneeMin : string : annee mini a conserver sur 4 caracteres
        """
        return df.loc[df.annee>=int('2018')].copy()
        
    def classer_comptage_insert_update(self,table='compteur', schema='comptage'):
        """
        repartir les comptages dans les classes de ceux qu'on connait et les nouveaux
        in : 
            table : strin : non schema-qualifyed table de la bdd contenant les compteurs
            schema : schema contenant les donnees de comptage
        """
        self.df_attr=self.filtrerAnnee(self.df_attr, '2018')
        #creation de l'existant
        existant = compteur_existant_bdd(table, schema=schema,dep='64', type_poste=False)
        #mise a jour des noms selon la table de corresp existnate
        self.corresp_nom_id_comptag(self.df_attr)
        #ajouter une colonne de geometrie
        #classement
        self.df_attr_update = self.df_attr.loc[self.df_attr.id_comptag.isin(existant.id_comptag.tolist())].copy()
        self.df_attr_insert = self.df_attr.loc[~self.df_attr.id_comptag.isin(existant.id_comptag.tolist())].copy()
        return
    
    def verifComptageInsert(self, dicoCorresp,gdfInsertCorrigee) : 
        """
        verifier que les comptages a inserer sont bine nouveau en les geolocalisant et comparant avec les donnes connues 
        in : 
            dicoCorresp = dico avec en clé l'id_comptag new et en value l'id_comptag gti
            gdfInsertCorrigee = geoDataFrame avec les geometries nulles remplies le plus possibles via SIG
        """

        #il reste des geoms nulles, ajoutées manuellement
        #après verif on passe le dico corresp dans la table bbd des coreesp_id_comptag
        self.insert_bdd('comptage', 'corresp_id_comptag', 
                        pd.DataFrame.from_dict({'id_gest': [k for k in  dicoCorresp.keys()],'id_gti':[v for v in  dicoCorresp.values()]}))
        #recalcul des correspondances
        self.classer_comptage_insert_update()
        self.df_attr_insert= gp.GeoDataFrame(self.df_attr_insert.merge(gdfInsertCorrigee[['id_comptag', 'geometry']], on='id_comptag'))
        return
        

class Comptage_cd23(Comptage):
    """
    Pour le moment dans le 23 on ne reçoit qu'un fichier des comptages tournant, format xls
    attention : pour le point de comptage D941 6+152 à Aubusson, le pR est 32 et non 6. il faut donc corriger à la main le fihcier excel
    car il y a 2 pr 6+32 sur la d941
    """
    def __init__(self,fichier, annee) : 
        self.fichier=fichier
        self.annee=annee
        
    def ouvrirMiseEnForme(self):
        """
        ouvertur, nettoyage du fichier
        """
        # ouvrir le classeur
        df_excel=pd.read_excel(r'Q:\DAIT\TI\DREAL33\2020\OTV\Doc_travail\Donnees_source\CD23\2019-CD23_trafics.xls',skiprows=11)
        # renomer les champs
        df_excel_rennome=df_excel.rename(columns={'1er trimestre  du 01 janvier au 31 mars':'trim1_TV', 'Unnamed: 9':'trim1_pcpl',
                         '2ème trimestre du 01 avril au 30 juin':'trim2_TV', 'Unnamed: 11':'trim2_pcpl',
                         '3ème trimestre du 01 juillet au 30 septembre':'trim3_TV', 'Unnamed: 13':'trim3_pcpl',
                         '4ème trimestre du 01 octobre au 31 décembre':'trim4_TV', 'Unnamed: 15':'trim4_pcpl',
                         'Unnamed: 17':'pc_pl', f'TMJA {self.annee}':'tmja'})
        #supprimer la 1ere ligne
        df_excel_filtre=df_excel_rennome.loc[1:,:].copy()
        #mise en forme attribut
        df_excel_filtre['Route']=df_excel_filtre.apply(lambda x : str(x['Route']).upper(), axis=1)
        #attribut id_comptag
        for i in ['DEP','PR','ABS'] : 
            df_excel_filtre[i]=df_excel_filtre.apply(lambda x : str(int(x[i])),axis=1)
        df_excel_filtre['id_comptag']=df_excel_filtre.apply(lambda x : '-'.join([x['DEP'],'D'+str(x['Route']),
                                                                                 x['PR']+'+'+x['ABS']]),axis=1)
        df_excel_filtre['fichier']=self.fichier
        df_excel_filtre['src']='tableur CD23'
        self.df_attr=df_excel_filtre
        
    def classer_compteur_update_insert(self, table_cpt='compteur'):
        """
        separer les donnes a mettre a jour de celles a inserer, selon les correspondances connues et les points existants
        """
        self.corresp_nom_id_comptag(self.df_attr)
        self.existant  = compteur_existant_bdd(table_cpt, dep='23')
        self.df_attr_update=self.df_attr.loc[self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        self.df_attr_insert=self.df_attr.loc[~self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        
    def update_bdd_23(self, schema, table):
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr_update,['id_comptag','tmja','pc_pl','src', 'fichier'])
        dico_attr_modif={f'tmja_{self.annee}':'tmja', f'pc_pl_{self.annee}':'pc_pl',f'src_{self.annee}':'src', 'fichier':'fichier'}
        self.update_bdd(schema, table, valeurs_txt,dico_attr_modif)
    
    def donneesMens(self):
        """
        calcul des données mensuelles uniquement sur la base du fichier excel de regroupement
        """
        list_id_comptag=[val for val in self.df_attr.id_comptag.tolist() for _ in (0, 1)]
        donnees_type=['tmja','pc_pl']*len(self.df_attr.id_comptag.tolist())
        annee_df=[str(self.annee)]*2*len(self.df_attr.id_comptag.tolist())
        janv, fev, mars,avril,mai,juin,juil,aout,sept,octo,nov,dec=[],[],[],[],[],[],[],[],[],[],[],[]
        for i in range(len(self.df_attr.id_comptag.tolist())) :
            for j in (janv, fev, mars) :
                j.extend([self.df_attr.trim1_TV.tolist()[i],self.df_attr.trim1_pcpl.tolist()[i]])
            for k in (avril,mai,juin) :
                k.extend([self.df_attr.trim2_TV.tolist()[i],self.df_attr.trim2_pcpl.tolist()[i]])
            for l in (juil,aout,sept) :
                l.extend([self.df_attr.trim3_TV.tolist()[i],self.df_attr.trim3_pcpl.tolist()[i]])
            for m in (octo,nov,dec) :
                m.extend([self.df_attr.trim4_TV.tolist()[i],self.df_attr.trim4_pcpl.tolist()[i]])
        self.df_attr_mensuel=pd.DataFrame({'id_comptag':list_id_comptag,'donnees_type':donnees_type,'annee':annee_df,'janv':janv,'fevr':fev,'mars':mars,'avri':avril,
                      'mai':mai,'juin':juin,'juil':juil,'aout':aout,'sept':sept,'octo':octo,'nove':nov,'dece':dec})
            
class Comptage_cd17(Comptage) :
    """
    Classe d'ouvertur de fichiers de comptage du CD17
    en entree : 
        fichier : raw string de chemin du fichier
        type_fichier : type e fichier parmi ['brochure', 'permanent_csv',tournant_xls_bochure', 'permanent_csv_formatTmjPl', 'permanent_csv_format3sens', 'ponctuel_csv']
                       permanent_csv_formatTmjPl est un format avec que des données 2 sens confondues, et une par indicateur ; fournie en 2020
                       permanent_csv_format3sens  est un format avec une ligne par sens 'fournie en 2019)
        annee : integer : annee des points de comptages
    """  
            
    def __init__(self,fichier, type_fichier, annee, delimiter=';', skiprows=0):
        Comptage.__init__(self, fichier)
        self.annee = annee
        self.fichier = fichier
        # verif de la valeur du type_fichier
        self.liste_type_fichier = ['brochure_pdf', 'permanent_csv_format3sens','tournant_xls_bochure','ponctuel_xls_bochure',
                                   'permanent_csv_format3sens', 'permanent_csv_formatTmjPl', 'ponctuel_csv']
        O.checkParamValues(type_fichier, self.liste_type_fichier)
        self.type_fichier = type_fichier
        if self.type_fichier == 'brochure_pdf':
            self.fichier_src=self.lire_borchure_pdf()
        elif self.type_fichier in ('permanent_csv_format3sens', 'permanent_csv_formatTmjPl'): 
            self.fichier_src=self.ouvrir_csv(delimiter, skiprows)
        elif self.type_fichier == 'tournant_xls_bochure': 
            self.fichier_src=self.ouvrir_xls_tournant_brochure()
        elif self.type_fichier == 'ponctuel_xls_bochure':
            self.fichier_src = self.ouvrir_xls_ponctuel_brochure()
        elif self.type_fichier == 'ponctuel_csv':
            self.fichier_src = self.ponctuel_csv_attr()
            
        
        
    def lire_borchure_pdf(self):
        """
        Transformer les données contenues dans les borchures pdf en list de données
        POur ça il faut copier-coller les données de la brohcure pdf en l'ouvrant avec firfox puis cole dans notepad puis enregistrer en .txt. 
        Pas d'en tete de colonne
        attention à la dénomination des mois,la vérifier
        """
        with open(self.fichier, encoding="utf-8") as fichier :  # ouvrir l fichier
            liste = [element.replace('\n', ' ').replace('    ', ' ').replace('   ', ' ').replace('  ', ' ') for element in fichier]
            liste_decomposee_ligne = re.split(r'(Janvier|Février|Mars|Avril|Mai|Juin|Juillet|Août|Sept.|Oct.|Nov.|Déc.|Jan.)', "".join(liste))  # permet de conserver le mois
            liste_decomposee_ligne = [liste_decomposee_ligne[i] + liste_decomposee_ligne[i + 1] for i in range(0, len(liste_decomposee_ligne) - 1, 2)]  # necessaier pour repasser le mois dans le string
            liste_decomposee_ligne[0] = ' ' + liste_decomposee_ligne[0]  # uniformité des données
            liste_decomposee_ligne = list(filter(None, liste_decomposee_ligne))
            liste_decomposee_ligne=[e.strip() for e in liste_decomposee_ligne]
        return liste_decomposee_ligne
    
    def brochure_pdf_voie_pr_abs(self):
        """
        extraire les voies, pr et abscisse de la liste decomposee issue de lire_borchure_pdf
        en entree : 
            self.fichier_src : liste des element issu du fichier.txt issu du pdf 
        en sortie : 
            voie : list de string : dénomination de la voie au format D000X
            pr : list d'integer : point de repere
            abs : list d'integer : absice
        """
        return [element.split(' ')[1] for element in self.fichier_src], [element.split(' ')[2]for element in self.fichier_src], [element.split(' ')[3]for element in self.fichier_src]
    
    def brochure_pdf_tmj_pcpl_v85(self):
        """
        extraire le tmj, %pl et v85 de la liste decomposee issue de lire_borchure_pdf
        en entree : 
            self.fichier_src : liste des element issu du fichier.txt issu du pdf
        en sortie : 
            tmj : list de numeric : trafic moyen journalier
            pc_pl : list de numeric : point de repere
            v85 : list de numeric : absice
        """
        # pour le tmj et %PL c'est plus compliqué car la taille de la cellule localisation varie, son délimiteur aussi et les chiffres peuvent être entier ou flottant, don on va se baser sur le fait
        # que la rechreche d'un nombre a virgule renvoi le %PL, sinon la vitesse, et si c'est la vitesse, alors ca créer une value error en faisant le float sur l'element + 1, donc on 
        # sait que c'est la vitesse
        pc_pl, v85, tmj = [], [], []
        for element in self.fichier_src : 
            element_decompose = element.split()
            nombre_a_virgule = re.search('[0-9]{1,}\,[0-9]{1,}', element)  # rechreche un truc avec deux chiffres séparés par une virgule : renvoi un objet match si ok, none sinon
            if nombre_a_virgule : 
                try : 
                    v85.append(float(element_decompose[element_decompose.index(nombre_a_virgule.group()) + 1].replace(',', '.')))  # idem tmj
                    pc_pl.append(float(nombre_a_virgule.group().replace(',', '.'))) 
                    tmj.append(int(element_decompose[element_decompose.index(nombre_a_virgule.group()) - 2]))  # donc on en deduit le tmja selon al position (car seare par un ' ' dans le fichier de base) 
                except ValueError :
                    pc_pl.append(float(element_decompose[element_decompose.index(nombre_a_virgule.group()) - 1].replace(',', '.')))
                    v85.append(float(nombre_a_virgule.group().replace(',', '.')))
                    tmj.append(int(element_decompose[element_decompose.index(nombre_a_virgule.group()) - 3]))            
            else :  # si les deux données sont des entiers
                liste_nombre = []  # la liste des nombres dans la ligne de données :
                for objet in element_decompose : 
                    try :  # comme ça on ne garde que les nombre
                        liste_nombre.append(float(objet))
                    except ValueError : 
                        pass
                tmj, pc_pl, v85 = liste_nombre[-4], liste_nombre[-2], liste_nombre[-1]  
        return tmj, pc_pl, v85
    
    def brochure_pdf_mois_periode(self):
        """
        extraire le mois et la periode de mesure issue de lire_borchure_pdf
        en entree : 
            self.fichier_src : liste des element issu du fichier.txt issu du pdf
        en sortie : 
            tmj : list de numeric : trafic moyen journalier
            pc_pl : list de numeric : point de repere
            v85 : list de numeric : absice
        """
        # plus que les dates de mesure !!
        mois = [element.split()[-1] for element in self.fichier_src]
        periode = [element.split()[-3] + '-' + element.split()[-2] for element in self.fichier_src]
        return mois, periode

    def brochure_pdf_tt_attr(self):
        """
        sort une dataframe des voie, pr, abs, tmj, pc_pl, v85, periode et mois
        """
        voie, pr, absc=self.brochure_pdf_voie_pr_abs()
        tmj, pcpl, v85=self.brochure_pdf_tmj_pcpl_v85()
        mois, periode=self.brochure_pdf_mois_periode()
        return pd.DataFrame({'route': voie, 'pr':pr,'abs':absc,'tmja_'+str(self.annee):tmj, 'pc_pl_'+str(self.annee):pcpl, 'v85':v85,'mois':mois, 'periode':periode}) 
    
    def permanent_csv_attr(self) :
        """
        sort une dataframe des voie, pr, abs, tmj, pc_pl, v85, periode et mois pour les fichiers csv de comptag permanent
        """
        if self.type_fichier == 'permanent_csv_format3sens':
            fichier_src_2sens = self.fichier_src.loc[self.fichier_src['Sens']=='3'].copy()
            liste_attr = ([a for a in fichier_src_2sens.columns if a[:6]=='MJM TV']+['Route','PRC','ABC']+['MJA TV TCJ '+str(self.annee),
                                                                                'MJA %PL TCJ '+str(self.annee),'MJAV85 TV TCJ '+str(self.annee)])
            liste_nom = (['janv', 'fevr', 'mars', 'avri', 'mai', 'juin', 'juil', 'aout', 'sept', 'octo', 'nove', 'dece']+['route', 'pr','abs']+[
                                                                                'tmja_'+str(self.annee), 'pc_pl_'+str(self.annee), 'v85'])
            dico_corres_mois={a:b for a,b in zip(liste_attr,liste_nom)}
            fichier_filtre = fichier_src_2sens[liste_attr].rename(columns=dico_corres_mois).copy()
            fichier_filtre = fichier_filtre.loc[~fichier_filtre['tmja_'+str(self.annee)].isna()].copy()
            fichier_filtre['tmja_'+str(self.annee)] = fichier_filtre['tmja_'+str(self.annee)].apply(lambda x : int(x))
            fichier_filtre['pc_pl_'+str(self.annee)] = fichier_filtre['pc_pl_'+str(self.annee)].apply(lambda x : float(x.strip().replace(',','.')))
            fichier_filtre['route'] = fichier_filtre.route.apply(lambda x : x.split(' ')[1]) 
            fichier_filtre['src'] = self.type_fichier
        elif self.type_fichier == 'permanent_csv_formatTmjPl':
            fichier_filtre = self.fichier_src.drop(cd17_permCsvDropligneDebut)
            fichier_filtre.columns = cd17_permCsvTmjmPlColumns
            fichier_filtre['route'] = fichier_filtre.route.apply(lambda x: O.epurationNomRoute(x))
            fichier_filtre['pr'] = fichier_filtre.pr.astype('int')
            fichier_filtre['abs'] = fichier_filtre['abs'].astype('int')
            fichier_filtre['latitude'] = fichier_filtre['latitude'].astype('float')
            fichier_filtre['longitude'] = fichier_filtre['longitude'].astype('float')
            fichier_filtre['id_comptag'] = fichier_filtre.apply(lambda x: f"17-{x.route}-{x.pr}+{x['abs']}", axis=1)
            fichier_filtre['geom'] = fichier_filtre.apply(lambda x: O.reprojeter_shapely(Point(x.longitude, x.latitude), '4326', '2154')[1], axis=1)
            fichier_filtre['src'] = self.type_fichier
            for c in [c for c in fichier_filtre.columns if 'pc_pl' in c]:
                fichier_filtre[c] = fichier_filtre[c].apply(lambda x: x.replace(',', '.').replace('%', '') if not isinstance(x, float) else x).astype('float')
            for c in [c for c in fichier_filtre.columns if 'tmja' in c]:
                fichier_filtre[c] = fichier_filtre[c].apply(lambda x: int(x) if not isinstance(x, float) else None)
            fichier_filtre = gp.GeoDataFrame(fichier_filtre, geometry='geom')
            fichier_filtre.set_crs('epsg:2154', inplace=True)
        else:
            raise NotImplementedError("format de permanent non traites pour le moment")
        return fichier_filtre
    
    
    def permanent_csv_attrMens(self, fichier_filtre):
        """
        extraire les données mensuelle des données issues de permanent_csv_attr()
        """
        dfMensPerm = fichier_filtre.melt(id_vars='id_comptag',
                                         value_vars=[c for c in fichier_filtre.columns if '_tmja' in c or '_pc_pl' in c],
                                         var_name='indic',
                                         value_name='valeur')
        dfMensPerm['mois'] = dfMensPerm.indic.apply(lambda x: x.split('_')[0])
        dfMensPerm['indicateur'] = dfMensPerm.indic.apply(lambda x: '_'.join(x.split('_')[1:]))
        dfMensPerm.drop('indic', axis=1, inplace=True)
        return dfMensPerm
    
    
    def ponctuel_csv_attr(self):
        """
        a partir du format csv du millésime 2020, mettre en forme les donnees : nom d'attribut, creation d'attribut
        """
        dfBrute = pd.read_csv(self.fichier, skiprows=11, delimiter=';', parse_dates=['Début du comptage', 'Fin du comptage'], dayfirst=True, encoding='LATIN1').rename(columns={'Unnamed: 0': 'nom'})
        dfBrute['pr_abs'] = dfBrute.nom.apply(lambda x: re.search('[0-9]{1,3}\+[|0-9]{3}', x).group(0) if re.search('[0-9]{1,3}\+[|0-9]{3}', x)  else None)
        dfBrute['pr'] = dfBrute.pr_abs.apply(lambda x: x.split('+')[0] if x else None)
        dfBrute['absc'] = dfBrute.pr_abs.apply(lambda x: x.split('+')[1] if x else None)
        dfBrute['nomRoute'] = dfBrute.nom.str.strip().apply(lambda x: x.split()[0])
        dfBrute.rename(columns={'MJA TV S3': 'tmja','%PL S3': 'pc_pl', 'nomRoute': 'route' }, inplace=True)
        dfBrute['periode'] = dfBrute.apply(lambda x: f"{x['Début du comptage'].strftime('%Y/%m/%d')}-{x['Fin du comptage'].strftime('%Y/%m/%d')}" 
                                   if not pd.isnull(x['Début du comptage']) and not pd.isnull(x['Fin du comptage']) else None, axis=1)
        dfBrute['pc_pl'] = dfBrute.pc_pl.apply(lambda x: float(x.replace(',', '.')))
        # filtre des données et cretion de l'identifiant
        dfDonneesLocalisable = dfBrute.loc[dfBrute.route.apply(lambda x: True if re.search('D[0-9]+', x) else False)].copy()
        dfDonneesLocalisable['id_comptag'] = dfDonneesLocalisable.apply(lambda x: f"17-{x.route}-{x.pr}+{x['absc']}", axis=1)
        dfDonneesLocalisable = dfDonneesLocalisable.loc[(~dfDonneesLocalisable.pr.isna()) & (~dfDonneesLocalisable.absc.isna()) &
                                                        (dfDonneesLocalisable.tmja > 0)].copy()
        dfDonneesLocalisable['type_poste'] = 'ponctuel'
        dfDonneesLocalisable['src'] = self.type_fichier
        dfDonneesLocalisable['fichier'] = os.path.basename(self.fichier)
        dfDonneesLocalisable['src'] = self.type_fichier
        dfDonneesLocalisable['type_veh'] = 'tv/pl'
        dfDonneesLocalisable['annee'] = self.annee
        dfDonneesLocalisable['dep'] = '17'
        dfDonneesLocalisable['reseau'] = 'CD'
        dfDonneesLocalisable['gestionnai'] = 'CD17'
        dfDonneesLocalisable['concession'] = 'False'
        dfDonneesLocalisable['src_geo'] = 'pr+abs_gestionnaire'
        dfDonneesLocalisable['fictif'] = False
        dfDonneesLocalisable['src_cpt'] = 'gestionnaire'
        dfDonneesLocalisable['sens_cpt'] = 'double sens'
        dfDonneesLocalisable['en_service'] = True
        return dfDonneesLocalisable

    def ouvrir_xls_tournant_brochure (self):
        """
        ouvrir le fichier et y ajouter un id_comptag
        """
        donnees_brutes=pd.read_excel(self.fichier, skiprows=1)
        donnees_brutes.dropna(axis=1, how='all',inplace=True)
        donnees_brutes.columns=['identifiant','localisation','type','route','pr','tmja_2018','tmja_2017','evol_2017_2018','tmja_2016','tmja_2015','tmja_2014']
        def pr(pr):
            pr=str(pr)
            ptpr=pr.split('.')[0]
            absc=pr.split('.')[1]
            if len(absc)<4 : 
                absc=int(absc+((4-len(absc))*'0'))
            else : 
                absc=int(absc)
            pr=ptpr+'+'+str(absc)
            return pr
        donnees_brutes['pr']=donnees_brutes.apply(lambda x : pr(x['pr']), axis=1)
        donnees_brutes['id_comptag']=donnees_brutes.apply(lambda x : '17-'+O.epurationNomRoute(x['route'][2:].strip())+'-'+x['pr'],axis=1)
        return donnees_brutes
    
    def ouvrir_xls_ponctuel_brochure(self):
        """
        ouvrir le fichier, ajouter un id_comptag et obs, filtrer les points en juillet / aout
        """
        donnees_brutes=pd.read_excel(self.fichier, skiprows=2)
        donnees_brutes.columns=['route','pr','absc','localisation','tmja','pl','pc_pl','v85','agence','zone','debut_periode','fin_periode']
        donnees_brutes['id_comptag']=donnees_brutes.apply(lambda x : '17-'+O.epurationNomRoute(x['route'][2:].strip())+'-'+str(x['pr'])+'+'+str(x['absc']),axis=1)
        donnees_brutes['obs']=donnees_brutes.apply(lambda x : 'nouveau_point,'+x['debut_periode'].strftime('%d/%m/%Y')+'-'+
                                                   x['fin_periode'].strftime('%d/%m/%Y')+',v85_tv '+str(x['v85']), axis=1)
        return donnees_brutes
    
    def conversion_id_comptg_existant_xls_brochure (self):
        """
        prendre den compte si des id_comptages ont unnom differents entre le CD17 et le Cerema
        """
        rqt_corresp_comptg='select * from comptage.corresp_id_comptag'
        with ct.ConnexionBdd(nomConnBddOtv) as c:
            corresp_comptg=pd.read_sql(rqt_corresp_comptg, c.sqlAlchemyConn)
        self.df_attr=self.fichier_src.copy()
        self.df_attr['id_comptag']=self.df_attr.apply(lambda x : corresp_comptg.loc[corresp_comptg['id_gest']==x['id_comptag']].id_gti.values[0] 
                                                    if x['id_comptag'] in corresp_comptg.id_gest.tolist() else x['id_comptag'], axis=1)
    
    def filtrer_periode_ponctuels_xls_brochure(self):
        """
        filtrer des periodesde vacances
        """
        #filtrer pt comptage pendant juillet aout
        self.df_attr=self.df_attr.loc[self.df_attr.apply(lambda x : x['debut_periode'].month not in [7,8] and x['fin_periode'].month not in [7,8], 
                                                         axis=1)].copy()
            
    def carac_xls_brochure(self):
        """
        separer les donnes entre les pt de comptage deja dans la base et les autres (fonctionne pour tout type de comptage)
        """
        #identification des nouveaux points
        self.df_attr_insert=self.df_attr.loc[~self.df_attr['id_comptag'].isin(self.existant.id_comptag.tolist())].copy()
        #identification des points à mettre a jour
        self.df_attr_update=self.df_attr.loc[self.df_attr['id_comptag'].isin(self.existant.id_comptag.tolist())]
  
    def mises_forme_bdd_brochure_pdf(self, dep, type_poste):
        """
        mise en forme et decompoistion selon comptage existant ou non dans Bdd
        in : 
            localisation : facilité pour dire si je suis en tain d'utiliser uen bdd ici ou au boulot
        en sortie : 
            df_attr_insert : df des pt de comptage a insrere
            df_attr_update : df des points de comtage a mettre a jour
        """ 
        # mise en forme
        # RELIQUAT DES TRAITEMENTS AVEC DONNEES AVANT CHANGEMENT DE FORME DE BDD
        if self.type_fichier in ('brochure', 'permanent_csv_format3sens'):
            if self.type_fichier == 'brochure': 
                df_attr = self.brochure_pdf_tt_attr()
                df_attr['type_poste'] = type_poste
                df_attr['obs_'+str(self.annee)] = df_attr.apply(lambda x : 'nouveau point,'+x['periode']+',v85_tv '+str(x['v85']),axis=1)
                df_attr.drop(['v85', 'mois','periode'], axis=1, inplace=True)
            elif self.type_fichier == 'permanent_csv_format3sens' : 
                df_attr = self.permanent_csv_attr()
                df_attr['type_poste']=type_poste
                df_attr['obs_'+str(self.annee)] = df_attr.apply(lambda x : 'v85_tv '+str(x['v85']),axis=1)
                df_attr.drop('v85',axis=1, inplace=True)
            df_attr['id_comptag'] = df_attr.apply(lambda x : '17-'+O.epurationNomRoute(x['route'])+'-'+str(x['pr'])+'+'+str(x['abs']),axis=1)
            df_attr['dep'] = '17'
            df_attr['reseau'] = 'RD'
            df_attr['gestionnai'] = 'CD17'
            df_attr['concession']  ='N'
            df_attr['fichier'] = self.fichier
                #verif que pas de doublons et seprartion si c'est le cas
            existant = compteur_existant_bdd(dep, type_poste)
            self.corresp_nom_id_comptag(df_attr) 
        #ANNEE 2020
        elif self.type_fichier in ('permanent_csv_formatTmjPl'):
            # formatage des données 
            self.fichier_src.columns = ['idcpt_gest', 'indicateur'] + self.fichier_src.columns[2:].tolist()
            self.fichier_src.indicateur.replace({'TMJ': 'tmja', 'PL': 'pl', '%PL': 'pc_pl'}, inplace=True)
            self.fichier_src.replace({'%': '', ',':'.'}, regex=True, inplace=True)
            self.fichier_src['idcpt_gest'] = self.fichier_src.idcpt_gest.str.strip()
            for mois in self.fichier_src.columns[2:]:
                self.fichier_src[mois] = self.fichier_src[mois].astype(float)
            self.fichier_src = renommerMois(self.fichier_src)
            self.fichier_src['fichier'] = os.path.basename(self.fichier)
            #j jointure avec existant
            existant = compteur_existant_bdd(gest='CD17', type_poste='permanent')[['id_comptag', 'id_cpt']]
            existantNotNa = existant.loc[~existant.id_cpt.isna()]
            df_attr = existantNotNa.merge(self.fichier_src, left_on='id_cpt', right_on='idcpt_gest', how='left')
            df_attr = df_attr.loc[df_attr.indicateur.isin(['tmja', 'pc_pl'])].drop(['id_cpt', 'idcpt_gest'], axis=1).copy()
            df_attr.rename(columns={'Annee': 'valeur'}, inplace=True)
            df_attr['annee'] = str(self.annee)
        df_attr_insert = df_attr.loc[~df_attr['id_comptag'].isin(existant.id_comptag.to_list())]
        df_attr_update = df_attr.loc[df_attr['id_comptag'].isin(existant.id_comptag.to_list())]
        self.df_attr, self.df_attr_insert, self.df_attr_update=df_attr, df_attr_insert,df_attr_update  
            
        
    def update_bdd_17(self, schema, table, localisation='boulot'):
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr_update,['id_comptag',f'tmja_{self.annee}',f'pc_pl_{self.annee}', 'src',
                                                                      f'obs_{self.annee}','fichier'])
        dico_attr_modif={f'tmja_{self.annee}':f'tmja_{self.annee}', f'pc_pl_{self.annee}':f'pc_pl_{self.annee}', f'src_{self.annee}':'src',
                         f'obs_{self.annee}':f'obs_{self.annee}','fichier':'fichier'}
        self.update_bdd(schema, table, valeurs_txt,dico_attr_modif)
       
    def insert_bdd_mens(self, schema, table) :
        """
        insérer des données dans la table des comptages mensuels
        en entree : 
            schema : string nom du schema de la table
            table : string : nom de la table
        """ 
        list_attr_mens=['janv', 'fevr', 'mars', 'avri', 'mai', 'juin', 'juil', 'aout', 'sept', 'octo', 'nove', 'dece', 'id_comptag', 'donnees_type', 'annee']
        mens=self.df_attr.copy()
        mens['donnees_type']='tmja' #a travailler plus tatrd si on doit extraire le tmja à partir des noms de colonnes, en lien du coup avec permanent_csv_attr()
        mens['annee']=str(self.annee)
        mens_fin=mens[list_attr_mens].copy()
        with ct.ConnexionBdd(nomConnBddOtv) as c:
            mens_fin.to_sql(table,c.sqlAlchemyConn,schema=schema,if_exists='append', index=False )
        
        
    class CptCd17_typeFichierError(Exception):  
        """
        Exception levee si le type de fcihier n'est pas dans la liste self.liste_type_fichier
        """     
        def __init__(self, type_fichier):
            Exception.__init__(self,f'type de fichier "{type_fichier}" non présent dans {Comptage_cd17.liste_type_fichier} ')

class Comptage_cd19(Comptage):
    """
    Dans le 19 pour le moment on qu'un simple fichier xls des comptages perm ou tourn
    """
    def __init__(self,fichier,annee) :
        Comptage.__init__(self, fichier)
        self.annee=annee
    
    def comptage_forme(self):
        
        def id_comptage(route,pr) : 
            route=str(route).strip()
            pr=str(int(pr.split('+')[0]))+'+0' if int(pr.split('+')[1])==0 else str(int(pr.split('+')[0]))+'+'+str(int(pr.split('+')[1]))
            return '19-D'+route+'-'+pr
        
        donnees_brutes=pd.read_excel(self.fichier, skiprows=6)
        donnees_filtrees=donnees_brutes.rename(columns={' RD':'route','P.R.':'pr',self.annee:f'ann_{self.annee}'})[['route','pr',f'ann_{self.annee}']]
        donnees_filtrees=donnees_filtrees.loc[~donnees_filtrees.pr.isna()].copy()
        donnees_filtrees['route']=donnees_filtrees.route.apply(lambda x : x.strip().replace(' ','') if isinstance(x,str) else x)
        donnees_filtrees['id_comptag']=donnees_filtrees.apply(lambda x : id_comptage(x['route'],x['pr']), axis=1)
        donnees_filtrees['tmja']=donnees_filtrees[f'ann_{self.annee}'].apply(lambda x : 0 if (pd.isna(x) or x=='x') else int(x.split('\n')[0]))
        donnees_filtrees['pc_pl']=donnees_filtrees[f'ann_{self.annee}'].apply(lambda x : 0 if (pd.isna(x) or x=='x') else float(x.split('\n')[1].split('%')[0].replace(',','.')))
        donnees_filtrees['type_poste']='tournant'
        donnees_filtrees['src']='tableur_2019'
        donnees_transfert=donnees_filtrees.loc[donnees_filtrees['tmja']>0].copy()
        self.df_attr= donnees_transfert.copy()
    
    def classer_compteur_update_insert(self,table_cpt,schema_cpt,
                                       schema_temp,nom_table_temp,table_linearisation_existante,
                                       schema_graph, table_graph,table_vertex,id_name,table_pr) : 
        """
        classer les comptage a insrere ou mettre a jour, selon leur correspondance ou non avec des comptages existants.
        se base sur la creation de graph variable selon le type de comptage, pour prendsre en compte ou non certaines categories de vehicules
        in : 
            table_pr : string : schema qualifyed nom de la table de reference des pr
        """
        #creation de l'existant
        self.existant = compteur_existant_bdd(schema=schema_cpt, table=table_cpt,dep='19', type_poste=False)
        #mise a jour des noms selon la table de corresp existnate
        self.corresp_nom_id_comptag(self.df_attr)
        #classement
        self.df_attr_update=self.df_attr.loc[self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        self.df_attr_insert=self.df_attr.loc[~self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        #recherche de correspondance
        corresp_cd19=self.corresp_old_new_comptag(schema_temp,nom_table_temp,table_linearisation_existante,
                                       schema_graph, table_graph,table_vertex,id_name,table_pr)[0]
        #insertion des correspondance dans la table dediee
        self.insert_bdd(schema_cpt,'corresp_id_comptag',corresp_cd19[['id_comptag','id_comptag_lin']].rename(columns={'id_comptag':'id_gest','id_comptag_lin':'id_gti'}))
        #ajout de correspondance manuelle (3, à cause de nom de voie notamment) et creation d'un unique opint a insert
        #nouvelle mise a jour de l'id comptag suivant correspondance
        self.existant = compteur_existant_bdd(schema=schema_cpt, table=table_cpt,dep='19', type_poste=False)
        self.corresp_nom_id_comptag(self.df_attr)
        #puis nouveau classement
        self.df_attr_update=self.df_attr.loc[self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        self.df_attr_insert=self.df_attr.loc[~self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()

    def update_bdd_19(self, schema, table):
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr_update,['id_comptag','tmja','pc_pl', 'src'])
        dico_attr_modif={'tmja_2019':'tmja', 'pc_pl_2019':'pc_pl', 'src_2019':'src'}
        self.update_bdd(schema, table, valeurs_txt,dico_attr_modif)

class Comptage_cd40(Comptage):
    """
    Les donnees du CD40 sont habituelement fournies au format B152, i.e u dossier par site, chaque dossier contenant des sous dossiers qui menent a un fihcier xls
    """
    def __init__(self,dossier, annee, donneesType='B152'): 
        """
        attributs : 
            dossier : chemin complet du dossier des fichiers
            donneesType : texte : valeur parmis B152, B153. par défaut : B152
            annee : string : annee sur 4 caracteres
            liste_fichier : liste de paths vers les fichiers
        """
        self.dossier = dossier
        self.liste_fichier = [os.path.join(chemin, fichier) for chemin, dossier, files in os.walk(dossier) 
                              for fichier in files if fichier.endswith('.xls')]
        if donneesType not in ('B152', 'B153'): 
            raise ValueError('le type de données n\'est pas parmi B152, B153')
        self.donneesType = donneesType
        self.annee = annee
    
    def type_fichier(self,fichier): 
        """
        récupérer le type de fichier selon la case A5
        """
        return pd.read_excel(fichier, header=None, skiprows=1).iloc[3,0]
    
    def verif_liste_fichier(self):
        """
        trier la liste des fichiers fournis pour ne garder que le B152
        """
        return [fichier for fichier in self.liste_fichier if self.type_fichier(fichier) == self.donneesType]
            
    def extraire_donnees_annuelles(self, fichier):
        """
        recuperer les donnes annuelles et de caract des comptage
        """
        dep,gest, reseau, concession, type_poste='40','CD40','D','N','permanent'
        if self.donneesType == 'B152' :
            df = pd.read_excel(fichier, header=None, skiprows=2) #les 1eres lignes mettent le bordel dans la definition des colonnes
            df2 = df.dropna(how='all').dropna(axis=1, how='all')
            compteur = '040.'+df2.loc[0,125].split(' ')[1]
            vma = int(df2.loc[4 ,0].split(' : ')[1][:2])
            voie = O.epurationNomRoute(df2.loc[4, 141].split(' ')[1])
            pr,absice = df2.loc[4, 125].split(' ')[1],df2.loc[4,125].split(' ')[2]
            tmja = df2.loc[18, 107]
            pc_pl = round(df2.loc[19, 107],2)
            tmje = np.nan
            tmjhe = np.nan
        elif self.donneesType == 'B153' :
            df = pd.read_excel(fichier, header=None, skiprows=1) #les 1eres lignes mettent le bordel dans la definition des colonnes
            compteur = df2.loc[1 ,125].split(' ')[2]
            vma = int(df2.loc[5, 0].split(' : ')[1][:2])
            voie = O.epurationNomRoute(df2.loc[5, 125].split(' ')[1])
            pr,absice = df2.loc[5, 141].split(' ')[1],df2.loc[5, 141].split(' ')[3]
            tmja = df2.loc[df2[0] == int(self.annee)][114].values[0]
            pc_pl = float(df2.loc[df2[0] == int(self.annee)][149].values[0].split('\n')[1].replace(' ', '') .replace(',', '.'))
            tmje = df2.loc[df2[0] == int(self.annee)][123].values[0]
            tmjhe = df2.loc[df2[0] == int(self.annee)][132].values[0]
        id_comptag = dep + '-' + voie + '-' + pr + '+' + absice    
        return compteur, vma, voie, pr, absice, dep, gest, reseau, concession, type_poste, id_comptag, tmja, pc_pl, tmje, tmjhe, df2
    
    def remplir_dico(self, dico,datalist, *donnees):
        for t, d in zip(datalist, donnees): 
            dico[t].append(d)
    
    def donnees_mensuelles(self, df2, id_comptag):
        """
        extraire les donnes mens d'un fichiere te les mettre ne forme
        """
        if self.donneesType == 'B152' :
            donnees_mens = df2.loc[[7, 18, 19],[13, 23, 30,
                           37, 44, 51, 58,
                           65, 72, 79, 86,
                           93, 100, 107]].dropna(axis=1, how='all')
            #renommer les colonnes
            donnees_mens.columns = [element.replace('é', 'e').replace('.', '').lower() for element in list(donnees_mens.loc[7])]
                #remplacer l'annee en string et ne conserver 
            donnees_mens = donnees_mens.drop(7).replace(['D.Moy.Jour', '% PL'],['tmja', 'pc_pl'])
            donnees_mens['annee'] = self.annee
            donnees_mens['id_comptag'] = id_comptag
                #rearranger les colonnes
            cols = donnees_mens.columns.tolist()
            cols_arrangees = cols[-1:]+cols[:1]+cols[-2:-1]+cols[1:-2]
            donnees_mens = donnees_mens[cols_arrangees]
            donnees_mens.columns = ['id_comptag', 'donnees_type', 'annee', 'janv', 'fevr', 'mars', 'avri', 'mai', 'juin',
                                    'juil', 'aout', 'sept', 'octo', 'nove', 'dece']
        elif self.donneesType == 'B153':
            #inserer les valeusr qui vont bien
            donnees_mens = df2.loc[df2[0] == 2020][[6, 15, 24, 33, 42, 51, 60, 69, 78, 87, 96, 105]]
            donnees_mens['annee'] = self.annee
            donnees_mens['id_comptag'] = id_comptag
            donnees_mens.columns = ['janv', 'fevr', 'mars', 'avri', 'mai', 'juin', 'juil', 'aout', 'sept', 'octo', 
                                    'nove', 'dece', 'annee', 'id_comptag']   
            donnees_mens['donnees_type'] = 'tmja'
            donnees_mens.replace('Panne', np.NaN, inplace=True)
        return donnees_mens
    
    def comptage_forme(self):
        """
        mise en forme des comptages dans une df pour les donnees annuelles et une autre pour les donnees mensuelles
        """     
        #ouvrir un fichier et modifier son cotenu
        if self.donneesType=='B152' :
                dataList=['compteur', 'vma', 'voie', 'pr', 'absice', 'dep', 'gest', 'reseau', 'concession', 'type_poste', 'annee',
                         'id_comptag','tmja','pc_pl', 'fichier']
        elif self.donneesType=='B153' :
                dataList=['compteur', 'vma', 'voie', 'pr', 'absice', 'dep', 'gest', 'reseau', 'concession', 'type_poste', 'annee',
                         'id_comptag','tmja','pc_pl', 'fichier', 'tmje', 'tmjhe']
        dico_annee={k:[] for k in dataList}
        for i,fichier in enumerate(self.verif_liste_fichier()) :
            #traitement des donnees agregee a l'annee
            print (fichier)
            if self.donneesType=='B152' :
                (compteur, vma, voie, pr, absice, dep, gest, reseau, concession, type_poste,
                 id_comptag, tmja, pc_pl, tmje, tmjhe, df2) = self.extraire_donnees_annuelles(fichier)
                self.remplir_dico(dico_annee, dataList, compteur, vma, voie, pr, absice, dep, gest, reseau, concession, 
                                  type_poste, self.annee, id_comptag, tmja, pc_pl, os.path.basename(fichier))
            elif self.donneesType=='B153' :
                (compteur, vma, voie, pr, absice, dep, gest, reseau, concession, type_poste,
                 id_comptag, tmja, pc_pl, tmje, tmjhe, df2) = self.extraire_donnees_annuelles(fichier)
                self.remplir_dico(dico_annee, dataList, compteur, vma, voie, pr, absice, dep, gest, 
                                  reseau, concession, type_poste, self.annee, id_comptag, tmja, pc_pl, os.path.basename(fichier), tmje, tmjhe)
            #traiteent des donnees mensuelles
            donnees_mens=self.donnees_mensuelles(df2,id_comptag)
            donnees_mens['fichier']=os.path.basename(fichier)
            if i==0 : 
                donnees_mens_tot=donnees_mens.copy()
            else : 
                donnees_mens_tot=pd.concat([donnees_mens_tot,donnees_mens], sort=False, axis=0)
            #print(voie, pr, absice, fichier)
        #attributs finaux
        self.df_attr = pd.DataFrame(dico_annee)
        self.df_attr_mens = donnees_mens_tot.copy()
        
        
    def update_bdd_40(self, schema, table):
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr_update,['id_comptag','tmja','pc_pl'])
        dico_attr_modif={'tmja_2019':'tmja', 'pc_pl_2019':'pc_pl'}
        self.update_bdd(schema, table, valeurs_txt,dico_attr_modif)
    
class Comptage_cd47(Comptage):
    """
    traiter les données du CD47, selon les traitement on creer les attributs :
    -type_cpt
    -dico_perm
    -dico_periodq
    -dico_tempo
    -dico_tot
    PLUS TARD ON POURRA AJOUTER LA RECUPDES DONNEES HORAIRES
    """
    def __init__(self,dossier,type_cpt,annee ) :
        self.dossier=dossier 
        self.annee=annee
        self.liste_type_cpt=['TRAFICS PERIODIQUES','TRAFICS PERMANENTS','TRAFICS TEMPORAIRES']
        if type_cpt in self.liste_type_cpt:
            self.type_cpt=type_cpt
        else : 
            raise Comptage_cd47.CptCd47_typeCptError(type_cpt)
        #liste des dossiers contenant du permanent
    
    def modifier_type_cpt(self,new_type_cpt):
        """
        mettre à jour le type_cpt de l'objet
        """
        if new_type_cpt in self.liste_type_cpt:
            self.type_cpt=new_type_cpt
        else : 
            raise Comptage_cd47.CptCd47_typeCptError(new_type_cpt)
    
    def liste_dossier(self):
        """
        recupérer la liste des dossier contenant les comptages en fonction du type_cpt
        """
        liste_dossiers=[os.path.join(root,directory) for root, dirs, files in 
                 os.walk(self.dossier) 
                 for directory in dirs if 'UD' in root and self.type_cpt.upper() in root]
        return liste_dossiers
    
    def dico_fichier(self,liste_dossiers):
        """
        creer un dico contenant la liste des fichier contenu dans chaque dossier qui contient des fichiers de comptage
        in : 
            liste_dossiers : list des dossiers contenant des fichiers de commtage, cf liste_dossier
        """
        if self.type_cpt.upper()=='TRAFICS PERMANENTS':
            dico_fichiers={dossier : {'fichier' : O.ListerFichierDossier(dossier,'.xlsx')} for dossier in liste_dossiers}
        elif self.type_cpt.upper()=='TRAFICS PERIODIQUES': 
            dico_fichiers={dossier : {'fichier' : [[fichier for fichier in O.ListerFichierDossier(dossier,'.xlsx') 
                        if uniq in fichier  and re.search('(T[1-4]|[1-4]T)',fichier) and (
                            'vitesse' not in fichier.lower() and not re.search(' V (T[1-4]|[1-4]T)', fichier.upper())) ] 
                        for uniq in set([re.split('[T,V][1-4]',fichier)[0] 
                        for fichier in O.ListerFichierDossier(dossier,'.xlsx')]) ] } for dossier in liste_dossiers}
        elif self.type_cpt.upper()=='TRAFICS TEMPORAIRES' :
            dico_fichiers={dossier : {'fichier' : [[fichier for fichier in O.ListerFichierDossier(dossier,'.xlsx') 
                        for uniq in [re.split(' T.xlsx',fichier)[0] 
                        for fichier in O.ListerFichierDossier(dossier,'.xlsx')]
                        if uniq in fichier  and re.search(' T.xlsx',fichier) and 'V.xlsx' not in fichier
                        and re.search('D[ ]{0,1}[0-9]+',fichier)]] } for dossier in self.liste_dossier()}
        return dico_fichiers
    
    def ouverture_fichier(self,dossier,fichier):
        """
        ouvrir un fichier, prendre la bonne feuille, virer les NaN et renommer les colonnes
        in : 
            dossier : raw string du dossier contenant les fichier
            fichier : strin g du nom de fichier, sans el achemin
        """
        colonne = ['jour','type_veh']+[str(i)+'_'+str(i+1)+'h' for i in range(24)]+['total','pc_pl']
        wbFeuilles = load_workbook(os.path.join(dossier,fichier),
              read_only=True, keep_links=False).sheetnames
        try : 
            feuille2Sens = [a for a in wbFeuilles[::-1] if a[:2]=='S_'][0]     
            data=pd.read_excel(os.path.join(dossier,fichier),sheet_name=feuille2Sens, header=None)#on ouvre la premiere feuille ne partany de ladroite qui a un prefixe avec S_
        except IndexError : 
            try :
                feuille2Sens = [a for a in wbFeuilles[::-1] if 'sens 3' in a.lower().strip()][0] 
            except IndexError : 
                print(f'IndexError sur fichier {os.path.join(dossier,fichier)}')    
                #data=pd.read_excel(os.path.join(dossier,fichier),sheet_name=feuille2Sens, header=None)
                return
        data.dropna(axis=0,how='all', inplace=True)
        data.columns=colonne
        return data
    
    def id_comptag(self, data):
        """
        definir l'id_comptag à patir d'une df issue d'un fcihier xls (cf ouverture_fichier)
        """
        localisation = data.loc[1,'jour']
        if 'PR' in localisation.upper() :
            pr_abs=localisation.upper().split(' PR')[1].strip()[:-1].split(' ')[0]
            pr=int(pr_abs.split('+')[0])
            absc=int(pr_abs.split('+')[1])
            route=re.search('D[ ]{0,1}[0-9]+',localisation.split(' PR ')[0].split("(")[1].strip())[0].replace(' ','')
            id_comptag='47-'+route+'-'+str(pr)+'+'+str(absc)
        else : 
            pr=None
            absc=None
            route=re.search('D[ ]{0,1}[0-9]+',localisation)[0]
            id_comptag=localisation
        return id_comptag,pr,absc,route
    
    def donnees_generales(self,data):
        """
        recuperer le pc_pl et le tmja et les periodes de mesures à patir d'une df issue d'un fcihier xls (cf ouverture_fichier)
        """
        
        for c in data.columns : #pour les cas bizrres ou les données ne sont pas à leurs places et  / ou sans calcul du pc_pl
            try : 
                tmja=int(data.loc[data['jour']=='Moyenne journalière : ',c].values[0].split(' ')[0])
                try : 
                    pc_pl=float(data.loc[data.loc[data['jour']=='Moyenne journalière : ',c].index+2,c].values[0].split('(')[1][1:-3].replace(',','.'))
                except IndexError : 
                    pc_pl=int(data.loc[data.loc[data['jour']=='Moyenne journalière : ',c].index+2,c].values[0].split(' ')[0])/tmja*100
                break
            except (AttributeError,ValueError) :
                continue
        
        dico_date={a:b for (a,b) in zip(['janvier','février','mars' ,'avril','mai','juin','juillet','août','septembre','octobre','novembre','décembre'],
                           [a.lower() for a in ['January','February','March', 'April','May','June','July','August','September','October','November','December']])}
        #trouver date de mesures peu importe la lign / colonne
        test=data.loc[data.apply(lambda x : any(['Période ' in a for a in x.values if isinstance(a,str)]), axis=1)]
        b_mask=test.apply(lambda x : 'Période ' in str(x.values), axis=0).to_numpy()
        date_mes=test.to_numpy()[:,b_mask][0][0]
        
        debut_periode=date_mes.lower().split('du')[1].split('au')[0].strip()
        fin_periode=date_mes.lower().split('du')[1].split('au')[1].strip()
        for k,v in dico_date.items() :
            debut_periode=debut_periode.replace(k,v)
            fin_periode=fin_periode.replace(k,v)
        debut_periode=pd.to_datetime(debut_periode)
        fin_periode=pd.to_datetime(fin_periode)
        return tmja, pc_pl,debut_periode,fin_periode
    
    def donnees_horaires(self, data):
        """
        Récuperer les valeurs unitaires de trafics
        """
        data_filtree=data.loc[data['jour']!='jour'].copy().reset_index(drop=True)
        data_filtree['jour'].fillna(method='pad', inplace=True)
        data_horaires=data_filtree.iloc[:data_filtree.loc[data_filtree['jour'].apply(lambda x : x[:7]=='Moyenne' if isinstance(x,str) else False)].index[0]].copy()
        data_horaires.jour=pd.to_datetime(data_horaires.jour)
        data_horaires=data_horaires[data_horaires.columns[:-2]].set_index('jour')
        data_horaires=data_horaires.loc[data_horaires.type_veh.isin(['TV','PL'])].copy()
        return data_horaires
    
    def remplir_dico_fichier(self):
        """
        creer un dico avec les valeusr de tmja et pc_pl par id_comptage
        """
        dico_final={}
        dico=self.dico_fichier(self.liste_dossier())
        if self.type_cpt.upper()=='TRAFICS PERMANENTS' : 
            for k,v in dico.items():
                v['mensuel']={'donnees_type':['tmja', 'pc_pl']}
                v['listfichier']=[]
                for i,fichier in enumerate(v['fichier']):
                    try :
                        data=self.ouverture_fichier(k,fichier)
                    except Exception as e : 
                        print(f'probleme sur le fichier {fichier}')
                        print(e)
                        continue
                    id_comptag,pr,absc,route=self.id_comptag(data)
                    if not 'id_comptag' in v.keys() : 
                        v['id_comptag']=id_comptag
                    tmja, pc_pl,debut_periode=self.donnees_generales(data)[:3]
                    mois=[k for k, v in dico_mois.items() if unidecode(debut_periode.month_name()[:4].lower()) in v][0]
                    v['mensuel'][mois]=[tmja,pc_pl]
                    if i==0 : 
                        v['horaire']=self.donnees_horaires(data.iloc[4:]).assign(fichier=fichier)
                    else : 
                        v['horaire']=pd.concat([v['horaire'],self.donnees_horaires(data.iloc[4:]).assign(fichier=fichier)], sort=False, axis=0)
                    v['listfichier'].append(fichier)
                try :
                    v['tmja']=int(statistics.mean([a[0] for a in v['mensuel'].values() if not isinstance(a[0],str)]))
                except Exception as e :
                    print(f"{k} {v['fichier']} {e}") 
                v['pc_pl']=round(float(statistics.mean([a[1] for a in v['mensuel'].values() if not isinstance(a[0],str)])),1)
                v['pr'], v['absc'],v['route'] =pr, absc, route
            dico_final={v['id_comptag']:{'tmja':v['tmja'],'pc_pl':v['pc_pl'], 'type_poste' : 'permanent','periode':None,
                                         'pr':v['pr'],'absc':v['absc'],'route':v['route'], 'horaire':v['horaire'], 
                                         'mensuel':v['mensuel'], 'fichier': ', '.join(v['listfichier'])} for k,v in dico.items()}
        elif self.type_cpt.upper()=='TRAFICS PERIODIQUES' :
            for k,v in dico.items() :
                for liste_fichier in v['fichier'] :
                    listUniqFichier=[]
                    if liste_fichier :
                        liste_tmja=[]
                        liste_pc_pl=[]
                        listePeriode=[]
                        i=0
                        for fichier in liste_fichier :
                            if 'Vitesse' not in fichier : 
                                data=self.ouverture_fichier(k,fichier)
                                try :
                                    id_comptag,pr,absc,route=self.id_comptag(data) 
                                except Exception as e:
                                    print(f'probleme sur le fichier {fichier}')
                                    print(e)
                                    continue
                                tmja, pc_pl, debut_periode, fin_periode=self.donnees_generales(data)
                                periode=f'{debut_periode.strftime("%Y/%m/%d")}-{fin_periode.strftime("%Y/%m/%d")}'
                                if i==0 : 
                                    horaire=self.donnees_horaires(data.iloc[4:].assign(fichier=fichier))
                                else : 
                                    horaire=pd.concat([horaire,self.donnees_horaires(data.iloc[4:]).assign(fichier=fichier)], sort=False, axis=0)
                                i+=1
                                liste_tmja.append(tmja)
                                liste_pc_pl.append(pc_pl) 
                                listUniqFichier.append(fichier)
                                listePeriode.append(periode)
                        try :      
                            dico_final[id_comptag]={'tmja' : int(statistics.mean(liste_tmja)), 'pc_pl' : round(float(statistics.mean(liste_pc_pl)),1), 
                                            'type_poste' : 'tournant','periode': ' ; '.join(listePeriode),'pr':pr,
                                            'absc':absc,'route':route, 'horaire':horaire, 'fichier': ', '.join(listUniqFichier)}
                        
                        except Exception as e : 
                            print(f'erreur statistique sur {os.path.join(k,fichier)} {e}')
                         
        elif self.type_cpt.upper()=='TRAFICS TEMPORAIRES' : 
            for k,v in dico.items():
                for liste_fichier in v['fichier']:
                    if liste_fichier :
                        if len(liste_fichier)>1 : 
                            for fichier in liste_fichier :
                                data=self.ouverture_fichier(k,fichier)
                                id_comptag, pr, absc,route=self.id_comptag(data)
                                tmja, pc_pl, debut_periode, fin_periode=self.donnees_generales(data)
                                periode=f'{debut_periode.strftime("%Y/%m/%d")}-{fin_periode.strftime("%Y/%m/%d")}'
                                horaire=self.donnees_horaires(data.iloc[4:]).assign(fichier=fichier)
                                dico_final[id_comptag]={'tmja':tmja, 'pc_pl':pc_pl,'type_poste' : 'ponctuel',
                                                        'periode':periode,'pr':pr,'absc':absc,'route':route,'horaire':horaire,
                                                        'fichier': fichier}
                        else : 
                            data=self.ouverture_fichier(k,liste_fichier[0])
                            id_comptag,pr,absc,route=self.id_comptag(data)
                            tmja, pc_pl,debut_periode,fin_periode=self.donnees_generales(data)
                            periode=f'{debut_periode.strftime("%Y/%m/%d")}-{fin_periode.strftime("%Y/%m/%d")}'
                            horaire=self.donnees_horaires(data.iloc[4:]).assign(fichier=liste_fichier[0])
                            dico_final[id_comptag]={'tmja':tmja, 'pc_pl':pc_pl,'type_poste' : 'ponctuel',
                                                        'periode':periode,'pr':pr,'absc':absc,'route':route,'horaire':horaire,
                                                        'fichier': liste_fichier[0]}
        return dico_final

    def regrouper_dico(self):
        """
        regrouper les dico creer dans remplir_dico_fichier et sortir une df
        """
        self.type_cpt='TRAFICS PERMANENTS'
        self.dico_perm=self.remplir_dico_fichier()
        self.modifier_type_cpt('TRAFICS PERIODIQUES')
        self.dico_perdq=self.remplir_dico_fichier()
        self.modifier_type_cpt('TRAFICS TEMPORAIRES')
        self.dico_tempo=self.remplir_dico_fichier()   
        self.dico_tot={}
        for d in [self.dico_perm, self.dico_perdq, self.dico_tempo]:
            self.dico_tot.update(d)
     
    def dataframe_dico(self, dico):
        """
        conversion d'un dico en df
        """
        df_agrege=pd.DataFrame.from_dict({k: {x: y for x, y in v.items() if x not in  ('horaire', 'mensuel')} for k, v in dico.items()},
                                         orient='index').reset_index().rename(columns={'index': 'id_comptag'})
        df_agrege['src']='donnees_xls_sources'
        
        for (i,(k, v)) in enumerate(dico.items()) : 
            for x,y in v.items()  : 
                if x == 'horaire' :
                    df_horaire = y
                    df_horaire['id_comptag'] = k
                    df_horaire.drop('total', axis=1, inplace=True, errors='ignore')
                    if i==0 : 
                        df_horaire_tot=df_horaire.copy()
                    else : 
                        df_horaire_tot=pd.concat([df_horaire_tot,df_horaire], sort=False, axis=0)
        print(df_horaire_tot.columns)
        df_horaire_tot.rename(columns={c: f"h{c.split('_')[0]}_{c.split('_')[1].replace('h','')}" for c in df_horaire_tot.columns
                                       if re.match('[0-9]+_[0-9]+h', c)}, inplace=True)
        """df_horaire_tot.columns=['type_veh','h0_1','h1_2','h2_3','h3_4','h4_5','h5_6','h6_7','h7_8','h8_9','h9_10','h10_11','h11_12','h12_13',
                                'h13_14','h14_15','h15_16','h16_17','h17_18','h18_19','h19_20','h20_21','h21_22','h22_23','h23_24', 'id_comptag', 'fichier']
        """
        print(df_horaire_tot.columns)
        df_horaire_tot.reset_index(inplace=True)
        
        for (i,(k, v)) in enumerate(dico.items()) : 
            for x, y in v.items():
                if x == 'mensuel':
                    df_mensuel=pd.DataFrame(y, index=range(2))
                    df_mensuel['id_comptag'] = k
                    df_mensuel['annee'] = str(self.annee)
                    df_mensuel['fichier'] = v['fichier']
                    if i == 0 : 
                        df_mensuel_tot = df_mensuel.copy()
                    else : 
                        df_mensuel_tot = pd.concat([df_mensuel_tot,df_mensuel], sort=False, axis=0)
        df_mensuel_tot['annee']=str(self.annee)
               
        return df_agrege,df_horaire_tot, df_mensuel_tot   
    
    
    def update_bdd_47(self,schema,table,df):
        """
        mise a jour base sur la fonction update bdd
        """
        val_txt=self.creer_valeur_txt_update(df, ['id_comptag','tmja','pc_pl','src'])
        self.update_bdd(schema,table, val_txt,{f'tmja_{str(self.annee)}':'tmja',f'pc_pl_{str(self.annee)}':'pc_pl',f'src_{str(self.annee)}':'src'})
        
    def mise_en_forme_insert(self):
        """
        ajout des attributs à self.df_attr_insert attendu dans la table comptag avant transfert dans bdd
        in : 
            annee : string : annee sur 4 lettres pour mise enf orme nom attr
        """
        if not len(str(self.annee))==4 : 
            raise TypeError('annee doit un string ou entier sur 4 caracteres')
        self.df_attr_insert['dep']='47'
        self.df_attr_insert['reseau']='RD'
        self.df_attr_insert['gestionnai']='CD47'
        self.df_attr_insert['concession']='N'
        self.df_attr_insert['obs']=self.df_attr_insert.apply(lambda x : f"""nouveau_point,{x['debut_periode'].strftime("%d/%m/%Y")}-{x['fin_periode'].strftime("%d/%m/%Y")}""" if not (pd.isnull(x['debut_periode']) and  pd.isnull(x['fin_periode'])) else None,axis=1)
        self.df_attr_insert.rename(columns={'absc' : 'abs', 'tmja':'tmja_'+str(self.annee),'pc_pl':'pc_pl_'+str(self.annee),'obs':'obs_'+str(self.annee)},inplace=True)
        self.df_attr_insert.drop(['debut_periode','fin_periode'],axis=1,inplace=True)
        
    class CptCd47_typeCptError(Exception):  
        """
        Exception levee si le type de comptage n'est pas dans la liste self.self.type_cpt
        """     
        def __init__(self, type_cpt):
            Exception.__init__(self,f'type de comptage "{type_cpt}" non présent dans {Comptage_cd47.liste_type_cpt} ')


class Comptage_cd87(Comptage):
    """
    les données fournies par le CD87 sont des fichiers fim bruts
    attributs : 
        dossier : nom du dossier contenant les fchiers
        liste_nom_simple : list edes fihciers avce un n om explicite
        liste_nom_foireux : liste des fichiers avec un nom pourri
        annee : string : annee des comptages
    """
    def __init__(self,dossier, annee):
        self.dossier=dossier
        self.annee=annee
        self.liste_fichier=O.ListerFichierDossier(dossier,'.FIM')
    

    def classer_fichiers_par_nom(self):
        """
        classer les fihciers selon le nom : avec RD et PR ou sans
        """
        self.liste_nom_simple=[fichier for fichier in self.liste_fichier 
                  if re.search('[D][( |_)]{0,1}[0-9]{1,5}([A-O]|[Q-Z]|[a-o]|[q-z]){0,3}.*PR[( |_)]{0,1}[0-9]{1,4}',fichier)]
        self.liste_nom_foireux=[fichier for fichier in self.liste_fichier 
                    if re.search('^[0-9]{1,4}([A-Z]|[a-z]){0,3}[( |_)]{0,1}[0-9]{4,5}',fichier) and fichier not in self.liste_nom_simple]

        
    def verif_fichiers_dico(self):
        """
        verifier que les fichiers des listes (cf classer_fichiers_par_nom()) sont tous dans le dico
        """    
        #verif sur le nombre de fichier
        if not len(self.liste_nom_simple)+len(self.liste_nom_foireux)==len([e for a in [b['fichiers'] for a in self.dico_voie.values() for b in a] for e in a]) : 
            nbfichierManquants=abs((len(self.liste_nom_simple)+len(self.liste_nom_foireux))-len([e for a in [b['fichiers'] for a in self.dico_voie.values() for b in a] for e in a]) )
            raise self.CptCd87_ManqueFichierDansDicoError(nbfichierManquants)
        else : print('OK : tous fichiers dans dico')
        
    def supprimer_fichiers_doublons(self):
        """
        supprimer dans les listes de fichiers presentes dans le dico, les fihciers qui sont les mm
        """
        for v in self.dico_voie.values() :
            for e in v :
                liste_fichiers=e['fichiers']
                liste_fichier_dbl=[]
                dico_corresp_fichier_dbl={}
                for i,f_a_tester in enumerate(liste_fichiers) :
                    if i!=0:
                        if f_a_tester in [a for b in dico_corresp_fichier_dbl.values() for a in b] :
                            continue
                    for f_test in liste_fichiers :
                        if f_test!=f_a_tester and filecmp.cmp(os.path.join(self.dossier,f_a_tester),os.path.join(self.dossier,f_test),shallow=True):
                            liste_fichier_dbl.append(f_test)
                    dico_corresp_fichier_dbl[f_a_tester]=liste_fichier_dbl
                    liste_fichier_dbl=[]
                e['fichiers']=[a for a in dico_corresp_fichier_dbl.keys()] 
        
        
    def dico_pt_cptg(self):
        """
        creer un dico de caracterisation pdes pt de comptag : pr, absc, fihciers.
        le dico supprime les fihciers doublons du dossier source
        """
        #recuperer les liste de fichiers : 
        self.classer_fichiers_par_nom()
        #associer les fichiers à des voies_pr_absc pour la liste_nom_simple 
        self.dico_voie={}
        for fichier in self.liste_nom_simple :
            nom_voie=O.epurationNomRoute(re.sub('( |_)','',re.search('[D][( |_)]{0,1}[0-9]{1,5}([A-O]|[Q-Z]|[a-o]|[q-z]){0,3}',fichier).group(0))).upper()
            pr=re.search('(PR[( |_)]{0,1})([0-9]{1,2})([+]{0,1})([0-9]{0,3})',re.sub("(!|\()",'+',fichier)).group(2)
            absc=re.search('(PR[( |+)]{0,1})([0-9]{1,2})([+]{0,1})([0-9]{0,3})',re.sub("(!|\(|_|x)",'+',fichier)).group(4)
            if absc!='' and len(absc)<3 :
                absc, pr = int(pr[-(3-len(absc))] + absc), int(pr[0:-(3-2)])
            elif absc!='' : 
                absc, pr=int(absc),int(pr)
            else :
                pr,absc=int(pr),0
            if nom_voie in self.dico_voie.keys():
                for e in self.dico_voie[nom_voie] :
                    if pr==e['pr'] : 
                        if absc==e['abs'] : 
                            e['fichiers'].append(fichier)
                            break
                        else : 
                            self.dico_voie[nom_voie].append({'pr':pr,'abs':absc,'fichiers':[fichier]})
                            break
                else : 
                    self.dico_voie[nom_voie].append({'pr':pr,'abs':absc,'fichiers':[fichier]})
            else : 
                self.dico_voie[nom_voie]=[{'pr':pr,'abs':absc,'fichiers':[fichier]},]
                
        #pour la liste_nom_foireux : 
        for fichier in self.liste_nom_foireux : 
            nom_voie=O.epurationNomRoute('D'+re.split('( |_)',fichier)[0]).upper()
            absc, pr=int(re.split('( |_)',fichier)[2][-3:]), int(re.split('( |_)',fichier)[2][:-3])
            if nom_voie in self.dico_voie.keys():
                for e in self.dico_voie[nom_voie] :
                    if pr==e['pr'] : 
                        if absc==e['abs'] : 
                            e['fichiers'].append(fichier)
                            break
                        else : 
                            self.dico_voie[nom_voie].append({'pr':pr,'abs':absc,'fichiers':[fichier]})
                            break
                else : 
                    self.dico_voie[nom_voie].append({'pr':pr,'abs':absc,'fichiers':[fichier]})
            else : 
                self.dico_voie[nom_voie]=[{'pr':pr,'abs':absc,'fichiers':[fichier]},]
        
        self.verif_fichiers_dico()#verfi que tous fihciers dans dico
        self.supprimer_fichiers_doublons()
        
    def remplir_indicateurs_dico(self):
        """
        ajouter eu dico issu de dico_pt_cptg les données de tmja, pc_pl, date_debut, date_fin
        """
        for v in self.dico_voie.values(): 
            for e in v : 
                if 'tmja' in e.keys() : #pour pouvoir relancer le traitement sans refaire ce qui estdeja traite
                    print(f"fichier {e['fichiers']} deja traites")
                    continue
                if len(e['fichiers']) == 1: 
                    print(e['fichiers'][0])
                    try:
                        obj_fim=FIM(os.path.join(self.dossier,e['fichiers'][0]), 'Message')
                        if obj_fim.date_debut.year != int(self.annee):
                            print(f"le fichier {e['fichiers'][0]} ne contient pas des données de l'année {self.annee}")
                            continue
                        else :
                            e['tmja'], e['pc_pl'], e['date_debut'], e['date_fin'], e[
                            'periode'], e['horaire'] = (obj_fim.tmja, obj_fim.pc_pl, obj_fim.date_debut,obj_fim.date_fin, obj_fim.periode,
                                                        obj_fim.dfHoraire2Sens)
                    except PasAssezMesureError: 
                        continue
                    except Exception as ex : 
                        print(f"erreur : {ex} \n dans fichier : {e['fichiers'][0]}")   
                elif len(e['fichiers'])>1:
                    list_tmja = []
                    list_pc_pl = []
                    list_dfHoraire = []
                    for f in e['fichiers']:
                        try :  
                            obj_fim=FIM(os.path.join(self.dossier,f), 'Message')
                            if obj_fim.date_debut.year != int(self.annee):
                                print(obj_fim.date_debut, f"le fichier {f} ne contient pas des données de l'année {self.annee}")
                            else:
                                list_tmja.append(obj_fim.tmja)
                                list_pc_pl.append(obj_fim.pc_pl)
                                list_dfHoraire.append(obj_fim.dfHoraire2Sens)
                            print(f)
                        except (PasAssezMesureError,obj_fim.fimNbBlocDonneesError)  : 
                            print('dans except')
                            continue
                        except Exception as ex : 
                            print(f"erreur : {ex} \n dans fichier : {f}")
                    list_pc_pl=[p for p in list_pc_pl if p>0]
                    e['tmja'], e['date_debut'], e['date_fin'] = int(statistics.mean(list_tmja)), np.NaN, np.NaN
                    e['pc_pl'] = round(statistics.mean([p for p in list_pc_pl if p>0]), 2) if list_pc_pl else np.NaN
                    e['horaire'] = pd.concat(list_dfHoraire).drop_duplicates(['jour', 'indicateur'])
    
    def remplir_type_poste_dico(self):
        """
        ajouter eu dico issu de remplir_indicateurs_dico le type de poste selon le nb de fichiers ayant servi à calculer le tmja
        """        
        for v in self.dico_voie.values() : 
            for e in v : 
                if len(e['fichiers']) > 4 :
                    e['type_poste']='permanent'
                elif 1<len(e['fichiers'])<=4 : 
                    e['type_poste']='tournant'
                elif len(e['fichiers'])== 1 :
                    e['type_poste']='ponctuel'
                else : 
                    e['type_poste']='NC'
    
    def dataframe_dico(self):
        """
        obtenir une df à partir du dico issu de remplir_type_poste_dico
        """
        self.df_attr = pd.DataFrame([[k, e['pr'], e['abs'], e['tmja'], e['pc_pl'], e['type_poste'],
                                      e['date_debut'],e['date_fin']] for k, v in self.dico_voie.items() for e in v if 'tmja' in e.keys()],
                                      columns=['route','pr','absc','tmja','pc_pl','type_poste','date_debut','date_fin'])
        self.df_attr['id_comptag']=self.df_attr.apply(lambda x :'87-'+x['route']+'-'+str(x['pr'])+'+'+str(x['absc']), axis=1)
        dfHoraire = pd.concat([e['horaire'].assign(id_comptag=f"87-{k}-{e['pr']}+{e['abs']}") 
                                          for k, v in self.dico_voie.items() for e in v if 'tmja' in e.keys()])
        #suppression des doublons entier sauf fichiers
        dfHoraire = dfHoraire.drop_duplicates([c for c in dfHoraire.columns if c != 'fichier'])
        # verif que pour chaque jour et id_comptage, je n'ai pas plus de 2 type d'indicateur
        dfHoraire['nb_occ'] = dfHoraire.groupby(['jour', 'id_comptag']).indicateur.transform(lambda x: x.count())
        if not dfHoraire.loc[dfHoraire.nb_occ > 2].empty:
            warnings.warn(f" il y a des id_comptage avec plus de 2 indicateurs sur certaines journées. verifiez dfHoraire")
        # verif que les identifiants de comptage entre les données horaire et agreges correspondent:
        if not len(dfHoraire.id_comptag.unique()) == len(self.df_attr.id_comptag.unique()):
            raise ValueError(f"le nombre d'iuddentifiant de comptage entre les données horaire ({len(dfHoraire.id_comptag.unique())}) et agregee ({len(self.df_attr.id_comptag.unique())}) n'est pas équivalent. verifiez")
        self.df_attr_horaire = dfHoraire.drop('nb_occ', axis=1, errors='ignore')
        self.df_attr_mensuel = mensuelDepuisHoraire(self.df_attr_horaire.assign(annee=self.annee))
        
    def filtrer_periode_ponctuels(self):
        """
        filtrer des periodesde vacances
        """
        #filtrer pt comptage pendant juillet aout
        self.df_attr=self.df_attr.loc[self.df_attr.apply(lambda x : x['date_debut'].month not in [7,8] and x['date_fin'].month not in [7,8]
                                                         if not (pd.isnull(x['date_debut']) and pd.isnull(x['date_fin'])) else True, 
                                                         axis=1)].copy()
    
    def dataframe_dico_glob(self):
        """
        fonction globla de création d'une df
        """
        self.remplir_indicateurs_dico()
        self.remplir_type_poste_dico()
        self.dataframe_dico()
        self.filtrer_periode_ponctuels()
    
    def classer_compteur_update_insert(self,table_cpt,schema_cpt,
                                       schema_temp,nom_table_temp,table_linearisation_existante,
                                       schema_graph, table_graph,table_vertex,id_name,table_pr):
        """
        classer la df des comptage (self.df_attr) selon le spoints à mettre à jour et ceux à inserer, en prenant en compte les points diont l'id_comptag
        diffère mais qui sont sur le mm troncon elemnraire
        in :
            table_cpt : string : nom de la table dans la bdd contenar les cpt
            schema_cpt :string : nom du schma contenant la table
            schema_temp : string : nom du schema en bdd opur calcul geom, cf localiser_comptage_a_inserer
            nom_table_temp : string : nom de latable temporaire en bdd opur calcul geom, cf localiser_comptage_a_inserer
            table_linearisation_existante : string : schema-qualified table de linearisation de reference cf donnees_existantes
            schema_graph : string : nom du schema contenant la table qui sert de topologie
            table_graph : string : nom de la table topologie (normalement elle devrait etre issue de table_linearisation_existante
            table_vertex : string : nom de la table des vertex de la topoolgie
            id_name : nom de l'identifiant uniq en integer de la table_graoh
            table_cpt : string : table des comptages
        """
        #fare le tri avec les comptages existants : 
        #recuperer les compmtages existants
        self.existant = compteur_existant_bdd(table_cpt, schema=schema_cpt,dep='87', type_poste=False)
        self.df_attr_update=self.df_attr.loc[self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        self.df_attr_insert=self.df_attr.loc[~self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        #obtenir une cle de correspondace pour les comptages tournants et permanents
        df_correspondance=self.corresp_old_new_comptag(schema_temp,nom_table_temp,table_linearisation_existante,
                                        schema_graph, table_graph,table_vertex,id_name,table_pr,table_cpt)[0]
        #passer la df de correspondance dans le table corresp_id_comptage
        #verifier si cette clé n'existent pas deja dans la table de correspondance et passer les nouvelles dedans
        rqt_corresp_comptg='select * from comptage.corresp_id_comptag'
        with ct.ConnexionBdd(nomConnBddOtv) as c:
            corresp_comptg=pd.read_sql(rqt_corresp_comptg, c.sqlAlchemyConn)
        df_correspondance=df_correspondance[0].loc[~df_correspondance[0]['id_comptag'].isin(corresp_comptg.id_gest.tolist())]
        if not df_correspondance.empty:
            self.insert_bdd('comptage', 'corresp_id_comptag', 
               df_correspondance.rename(columns={'id_comptag_lin':'id_gti','id_comptag':'id_gest'})[['id_gest','id_gti']])
        #faire la correspondance entre les noms de comptage
        self.df_attr = self.corresp_nom_id_comptag(self.df_attr)
        #recalculer les insert et update
        self.df_attr_update=self.df_attr.loc[self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        self.df_attr_insert=self.df_attr.loc[~self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
    
    def update_bdd_d87(self,table_cpt,schema_cpt):
        """
        mettre à jour la bdd avec df_attr_update, en ayant au préalbale traite les NaN
        """
        #mettre en forme pour update
        self.df_attr_update['obs']=self.df_attr_update.apply(lambda x : x['date_debut'].strftime('%d/%m/%Y')+'-'+ x['date_fin'].strftime('%d/%m/%Y') if not pd.isnull(x['date_debut']) else '', axis=1)
        self.df_attr_update.loc[self.df_attr_update.pc_pl.isna(),'obs']='pc_pl inconnu'
        self.df_attr_update.loc[self.df_attr_update.pc_pl.isna(),'pc_pl']=-99
        self.df_attr_update['src']='fichiers FIM'
        self.df_attr_update['fichier']=self.dossier
        #preparer update
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr_update, ['id_comptag','tmja','pc_pl','obs','src','fichier'])
        dico_attr={'tmja_'+self.annee:'tmja','pc_pl_'+self.annee:'pc_pl','obs_'+self.annee:'obs','src_'+self.annee:'src','fichier':'fichier'}
        #update
        self.update_bdd(schema_cpt, table_cpt, valeurs_txt,dico_attr)
        
    def insert_bdd_d87(self,table_cpt,schema_cpt,nom_table_ref,nom_table_pr):
        """
        inserer les point df_attr_insert qui n'était pas à mettre à jour
        in : 
            nom_table_ref : string : nom de table contenant le referntiel (schema qualified)
            nom_table_pr : string : nom de table contenant les pr (schema qualified)
        """
        #mettre en forme le insert
        dbl=self.df_attr_insert.loc[self.df_attr_insert.duplicated('id_comptag', False)].copy()
        ss_dbl=self.df_attr_insert.loc[~self.df_attr_insert.index.isin(dbl.index.tolist())].copy()
        dbl=dbl.dropna()
        dbl_traite=dbl.loc[dbl.tmja==dbl.groupby('id_comptag').tmja.transform(max)].drop_duplicates().copy()
        self.df_attr_insert=pd.concat([dbl_traite,ss_dbl], axis=0, sort=False)
        self.df_attr_insert.pc_pl.fillna(-99, inplace=True)
        self.df_attr_insert['dep']='87'
        self.df_attr_insert['reseau']='RD'
        self.df_attr_insert['gestionnai']='CD87'
        self.df_attr_insert['concession']='N'
        self.df_attr_insert['obs']=self.df_attr_insert.apply(lambda x : f"""nouveau_point,{x['date_debut'].strftime("%d/%m/%Y")}-{x['date_fin'].strftime("%d/%m/%Y")}""" if not (pd.isnull(x['date_debut']) and  pd.isnull(x['date_fin'])) else None,axis=1)
        self.df_attr_insert.rename(columns={'absc' : 'abs', 'tmja':'tmja_'+self.annee,'pc_pl':'pc_pl_'+self.annee,'obs':'obs_'+self.annee},inplace=True)
        self.df_attr_insert.drop(['date_debut','date_fin','route'],axis=1,inplace=True)
        self.insert_bdd(schema_cpt,table_cpt, self.df_attr_insert)
        #mettre à jour la geom
        self.maj_geom(schema_cpt, table_cpt,nom_table_ref,nom_table_pr, dep='87')
    
    class CptCd87_ManqueFichierDansDicoError(Exception):  
        """
        Exception levee si des fichiers des listes ne sont pas presents dans le dico
        """     
        def __init__(self, nbfichierManquants):
            Exception.__init__(self,f'il manque {nbfichierManquants} fichiers présents dans liste_nom_simple ou liste_nom_foireux dans le dico_voie')


class Comptage_cd16(Comptage):
    """
    les données fournies par le CD16 sont des fichiers excel (jusqu'a 2021) ou geoloc filaire (a partir de 2021)
    pour les compteurs permanents , et des fichiers FIM pour les comptages temporaires.
    il y a aussi des données sur PIGMA qu'il a tout prix télécharger, car les données FIM ne sont géolocalisable que par PIGMA
        fichier_b15_perm : nom complet du fichier des comptages permanents
        fichier_cpt_lgn : nom complet du fihcier de ligne contenant tous les comptages issu de pigma
        fichier_station_cpt : nom complet du fihcier de points contenant tous les points comptages issu de pigma
        dossierIrisTv : dossier contenant les csv au format IRIS pour les donnees horaires
        dossierIrisPl : dossier contenant les csv au format IRIS pour les PL
        dossierCptTemp : dossier contenant les fichiers FIM des comptages temporaires
    """
    
    
    def __init__(self,fichier_b15_perm,fichier_cpt_lgn,fichier_station_cpt,annee, dossierIrisTv, dossierIrisPl, dossierCptTemp):
        self.fichier_b15_perm=fichier_b15_perm
        self.fichier_cpt_lgn=fichier_cpt_lgn
        self.fichier_station_cpt=fichier_station_cpt
        self.annee=annee
        self.dossierIrisTv=dossierIrisTv
        self.dossierIrisPl=dossierIrisPl
        self.dossierCptTemp=dossierCptTemp
        
        
    def cpt_perm_xls(self, skiprows, cpt_a_ignorer=()):
        """
        mettre en forme les comptages permannets issu des fichiers de comptages au format xls B15 de route plus
        in : 
            skiprows : nb de lignes du fichiers excel a ignorer
            cpt_a_ignorer : tuple de string des idc_omptag a ignorer au vu des commentaires du CD16
        """
        donnees_brutes=pd.read_excel(self.fichier_b15_perm, skiprows=skiprows)
        donnees_filtrees=donnees_brutes[[a for a in donnees_brutes.columns if not isinstance(a, int) and 'Unnamed' not in a]].copy()
        # traiter et mettre en forme
        tmja=donnees_filtrees.loc[donnees_filtrees .apply(lambda x : x['Identif. Local.'][-2:]==' 3', axis=1)].copy()
        tmja['pr']=tmja['PR+Distance'].apply(lambda x : re.split('(\.|\+|,)',str(x))[0])
        tmja['absc']=tmja['PR+Distance'].apply(lambda x : int(re.split('(\.|\+|,)',str(x))[2]+'0'*(4-len(re.split('(\.|\+|,)',str(x))[2]))) if re.search('(\.|\+|,)', str(x)) else 0)
        tmja['route']=tmja['Route'].apply(lambda x : O.epurationNomRoute(x[3:]))
        tmja['id_comptag']=tmja.apply(lambda x : f"16-{x['route']}-{x['pr']}+{x['absc']}", axis=1)
        tmja_final=tmja[['pr', 'absc', 'route', 'id_comptag', 'Année']].loc[(tmja['Année']!=' ') & (~tmja['Année'].isna())].rename(columns={'Année':'tmja'})
        tmja_mens=tmja[[c for c in tmja.columns if c in [a[2] for a in dico_mois.values()]]+['id_comptag']].copy()
        tmja_mens=tmja_mens.rename(columns={a:k for a in tmja_mens.columns if a!='id_comptag' for k,v in dico_mois.items() if a==v[2] })
        tmja_mens['donnees_type']='tmja'
        tmja_mens['annee']=str(self.annee)
        index_tmja=tmja.index.tolist()
        index_ppl=[a+1 for a in index_tmja]
        ppl=donnees_filtrees.loc[index_ppl].copy()
        ppl['pr']=ppl['PR+Distance'].apply(lambda x : re.split('(\.|\+|,)',str(x))[0])
        ppl['absc']=ppl['PR+Distance'].apply(lambda x : int(re.split('(\.|\+|,)',str(x))[2]) if re.search('(\.|\+|,)', str(x)) else 0)
        ppl['route']=ppl['Route'].apply(lambda x : O.epurationNomRoute(x[3:]))
        ppl['id_comptag']=ppl.apply(lambda x : f"16-{x['route']}-{x['pr']}+{x['absc']}", axis=1)
        ppl['Année']=ppl['Année'].apply(lambda x : round(float(x),2) if x!=' ' else np.NaN)
        ppl_mens=ppl[[c for c in ppl.columns if c in [a[2] for a in dico_mois.values()]]+['id_comptag']].copy()
        ppl_mens=ppl_mens.rename(columns={a:k for a in ppl_mens.columns if a!='id_comptag' for k,v in dico_mois.items() if a==v[2] })
        ppl_mens['donnees_type']='pc_pl'
        ppl_mens['annee']=str(self.annee)
        ppl_mens=ppl_mens.applymap(lambda x : round(x,2) if isinstance(x,float) else x)
        ppl_final=ppl[['pr', 'absc', 'route', 'id_comptag', 'Année']].rename(columns={'Année':'pc_pl'})
        df_trafic=tmja_final.merge(ppl_final[['pc_pl','id_comptag']], on='id_comptag')
        donnees_mens=pd.concat([tmja_mens,ppl_mens], axis=0, sort=False)
        #filtrer selon comm CD16
        df_compt_perm=df_trafic.loc[~df_trafic.id_comptag.isin(cpt_a_ignorer)].copy()
        donnees_mens=donnees_mens.loc[~donnees_mens.id_comptag.isin(cpt_a_ignorer)].copy()
        df_compt_perm['type_poste']='Per'
        df_compt_perm['src']='tableau B15'
        return df_compt_perm, donnees_mens
    
    
    def cpt_pigma(self):
        """
        mettre en forme les comptages issu de pigma
        """
        
        
        def trouverVmas(vma):
            """
            fonction permettant de déduireles vitesses maximales autorisées depuis l'attributqui va bien
            """
            # déduire les valeusr de vitesse
            if '/' in vma and vma[-1] != '/':
                vma_vl = int(vma.split('/')[0])
            else: 
                vma_vl = int(vma)
            # vérif et corrections des valeurs
            if vma_vl in valeursVmaAdmises:
                vma_pl = correspVmaVlVmaPl[vma_vl]
            else:
                raise ValueError(f"La vitesse VL {vma_vl} n'est pas présente dans la liste des valeurs de VMA admise dans le dictionnaire de correspondace vma_vl-vma_pl")
            return vma_vl, vma_pl
        
        
        donnees_brutes_lgn = gp.read_file(self.fichier_cpt_lgn)
        donnees_brutes_pt = gp.read_file(self.fichier_station_cpt)
        
        donnees_brutes = donnees_brutes_pt[cd16_columnsFichierPtPigma].merge(donnees_brutes_lgn[cd16_columnsFichierLgnPigma],
                                                                left_on=['AXE','PLOD','ABSD'], right_on=['AXE','PRC','ABC'])
        donnees_filtrees = donnees_brutes.loc[donnees_brutes.ANNEE_COMP.astype(str) == self.annee].rename(
            columns=cd16_dicoCorrespNomColums).replace({'type_poste': cd16_dicoCorrespTypePoste,'techno': cd16_dicoCorrespTechno})
        donnees_filtrees['id_comptag'] = donnees_filtrees.apply(lambda x : f"16-{x['route']}-{x['pr']}+{x['abs']}", axis=1)
        donnees_filtrees['src'] = 'sectionnement'
        donnees_filtrees['periode'] = donnees_filtrees.apply(lambda x: f"{x.PERIODE_DE.replace('-', '/')}-{x.PERIODE_FI.replace('-', '/')}" if 
                                                     not pd.isnull(x.PERIODE_DE) and not pd.isnull(x.PERIODE_FI) else None, axis=1)
        donnees_filtrees['type_veh'] = 'tv/pl'
        # recuperation des vitesse maximales autorisées
        donnees_filtrees[['vma_vl', 'vma_pl']] = donnees_filtrees.apply(lambda x: trouverVmas(x.vma), axis=1, result_type='expand')
        # remplacement des valeurs à 0 par du None
        donnees_filtrees.replace({'tmja': {0: np.nan}, 'tmje': {0: np.nan}, 'pc_pl_e': {0: np.nan}, 'tmjhe': {0: np.nan}, 'pc_pl_he': {0: np.nan},
                                  'vmoy': {0: np.nan}, 'v85': {0: np.nan}}, inplace=True)
        donnees_filtrees.drop(cd16_columnsASuppr, errors='ignore', axis=1, inplace=True)
        dfCompteurPigma = donnees_filtrees[[c for c in donnees_filtrees.columns if c in attBddCompteur]]
        dfComptagePigma = donnees_filtrees[[c for c in donnees_filtrees.columns if c in attrComptage]]
        dfIndicAgregePigma = donnees_filtrees[['id_comptag']+cd16_attrIndicAgregePigma]
        return dfCompteurPigma, dfComptagePigma, dfIndicAgregePigma
    
    
    def donnees_horaires(self, donnees_pigma):
        """
        à partir des FIM et du fichier geolocalise des comptages, creer le fichier horaire par pt de comptag
        in : 
            donnees_pigma : donnees de concatenation des comptages issu des fichier PIGMA.issu de cpt_pigma()
        """
        listDfHoraire=[]
        with os.scandir(self.dossierCptTemp) as it:
            for e in it:
                if e.is_file() and e.name.lower().endswith('.fim'):
                    print(e.name)
                    cpt=FIM(e.path, gest='CD16', verifQualite='Message')
                    listDfHoraire.append(cpt.dfHoraire2Sens.assign(section_cp=cpt.section_cp, 
                                                                   qualite=cpt.qualite, 
                                                                   periode=cpt.periode,
                                                                   sens_cpt='sens unique' if cpt.sens_uniq else 'double sens'))
        dfHoraire = pd.concat(listDfHoraire)        dfHoraireId = donnees_pigma[['id_sect', 'id_comptag']].merge(dfHoraire, left_on='id_sect', right_on='section_cp', how='right')
        # verif si des fichiers de comptages temporaires sont inconnues dans la bdd
        if dfHoraireId.id_comptag.isna().any() : 
            warnings.warn("au moins un comptage temporaire n'est pas présent dans le fihcier pigma. recherche dans la bdd...")
            with ct.ConnexionBdd(nomConnBddOtv) as c:
                refIdComptag = pd.read_sql("select id_comptag, id_sect from comptage.compteur where gestionnai = 'CD16'", c.sqlAlchemyConn)
            dfIdComptagInconnus = refIdComptag.merge(dfHoraireId.loc[dfHoraireId.id_comptag.isna()]
                                                     [[c for c in dfHoraireId.columns if c not in ('id_comptag', 'id_sect')]], 
                                                     left_on='id_sect', right_on='section_cp', how='right')
            if dfIdComptagInconnus.id_comptag.empty:
                warnings.warn("au moins un comptage temporaire n'est pas présent dans la bdd. créer les compteurs avant toute insertion")
            dfHoraireId = pd.concat([dfHoraireId.loc[dfHoraireId.id_comptag.notna()], dfIdComptagInconnus])   
        #dfHoraireId.drop(['id_sect','section_cp'], axis=1, errors='ignore', inplace=True)
        return dfHoraireId
        
    
    def comptage_forme(self, skiprows, cpt_a_ignorer, dossier_cpt_fim):
        """
        fusion des données de cpt_pigma() et cpt_perm_xls()
        in : 
            cf fonction cpt_perm_xls
        """
        df_compt_perm, donnees_mens=self.cpt_perm_xls(skiprows, cpt_a_ignorer)
        donnees_filtrees=self.cpt_pigma()
        self.df_attr=pd.concat([df_compt_perm,donnees_filtrees],sort=False, axis=0)
        self.df_attr_mens=donnees_mens
        self.donnees_horaires(donnees_filtrees)
        
    def extraireEnteteIRIS(self, fichier) : 
        """
        depuis un fichier IRIS, en extraire une df d'entete et l'id_comptag
        in : 
            fichier : raw string de chemin complet
        out : 
            entete : dataframe de l'entete brut
            id_comptag : id_comptag du point de mesure
            sens : string en sens1 ou sens2
        """
        entete=pd.read_csv(fichier, sep=';', nrows=9, header=None, encoding='LATIN1')
        id_comptag=f'16-D{entete.loc[4,16]}-{entete.loc[5,16]}'
        sens=f'sens {entete.loc[2,22][-1:]}'
        return entete, id_comptag, sens
    
    
    def creerDataIris(self, fichier, indicateur) : 
        """
        a partir d'un fichier format IRIS, creer les donnees selon le type de fichier PL ou TV
        in : 
            fichier : rawstring : chemin completdu fichier
            indicateur : TV ou PL
        """
        O.checkParamValues(indicateur, ('TV', 'PL'))
        dateparser=lambda x : dt.datetime.strptime(x,'%d/%m/%Y')
        data=pd.read_csv(fichier, sep=';', skiprows=10,skipfooter=11,  encoding='LATIN1', engine='python', parse_dates=['Jours'], date_parser=dateparser, usecols=[f'H{i}' for i in range(1,25)]+['Jours'])
        data.columns=['jour']+[f"h{int(c.split('H')[1])-1}_{c.split('H')[1]}" for c in data.columns if 'H' in c]
        data['annee']=self.annee
        data['indicateur']=indicateur
        data['fichier']=os.path.basename(fichier)
        return data
    
    
    def creerDfTtJoursIris(self, dossier, indicateur):
        """
        creer la df contenanttoute les données horaires TV
        in : 
            dossier : rawstrin du dossier PL ou TV
            indicateur : 'TV' ou 'PL' selon le dossier
        """
        listeDf=[]
        for files in os.listdir(dossier) : 
            fichier=os.path.join(dossier, files)
            if os.path.isfile(fichier) and fichier.endswith('.csv') and '_Voie' not in fichier: 
                id_comptag, sens=self.extraireEnteteIRIS(fichier)[1:]
                data=self.creerDataIris(fichier, indicateur)
                data['id_comptag']=id_comptag
                data['sens']=sens
                listeDf.append(data)
        return pd.concat(listeDf)
    
    
    def creerDfsTtjoursTtIndicIris(self):
        """
        creer les dfs de comptage, indic_agrege, indic_mensuel, indic_horaire contenant tout les jours et tout les indicateurs des données IRIS
        """
        # réation des données Horaires completes
        dataHoraireCompelete = pd.concat([self.creerDfTtJoursIris(self.dossierIrisTv, 'TV'), self.creerDfTtJoursIris(self.dossierIrisPl, 'PL')])
        # filtre des jours non conforme
        dfHoraireFichierFiltre = verifValiditeFichier(dataHoraireCompelete, 24)[0]
        # verif sur la concordance des 2 sens de circulation
        comparer2Sens(dfHoraireFichierFiltre, attributSens='sens')[1]
        # concatener les deux sens
        dfHoraireConcat = concatIndicateurFichierHoraire(dfHoraireFichierFiltre,
                                                         'indicateur')
        # calcul des TMJAs
        indic_agrege = tmjaDepuisHoraire(dfHoraireConcat.assign(annee=self.annee))
        # calcul du mensuel
        tmjMens = mensuelDepuisHoraire(dfHoraireConcat.assign(annee=self.annee))
        tmjMens = tmjMens.loc[~tmjMens.valeur.isna()].copy()
        # verif que tous les id_comptag existent deja en bdd
        dfIdsConnus, dfIdsInconnus = scinderComptagExistant(indic_agrege, self.annee, dep='16')
        if not dfIdsInconnus.empty : 
            raise ValueError(f"""les compteurs {', '.join(dfIdsInconnus.id_comptag.tolist())} ne sont pas présent dans la BDD. 
                                                          ils doivent l'etre pour pouvoir inserer les donnees IRIS""")
        # creation dela dataframe des comptages finale
        dfComptage = dfIdsConnus.drop_duplicates(['id_comptag', 'annee'])[['id_comptag', 'annee']].assign(
            src='donnees horaire IRIS', type_veh='tv/pl') 
        return dfComptage, indic_agrege, tmjMens, dfHoraireConcat
    

class Comptage_cd86(Comptage):
    """
    les données fournies par le CD86 sont des fichiers excel pour les compteurs permanents et secondaires,
    il y a un petit pb sur les compteurs permanents entre la donnees pr+abs chez nous et la leur dans le tableau, donc il faut la premiere fois (2018) tout passer dans la table de correspondance
    si un seul fichier avec 2 feuilles : renseigner le mm nom de fichier pour perm etsecondaire, et indiuqer les feuills
        fichier_perm : nom complet du fichier des comptages permanents
        fichier_secondaire : nom complet du fihcier des comptages secondaires
        annee : integer annee des comptages
        feuil_perm : si c le mm fihciers pour les comptages perm et secondaiere, nom de la feuille avec les comptages perm 
        feuil_secondaire : si c le mm fihciers pour les comptages perm et secondaiere, nom de la feuille avec les comptages secondaires
    """
    def __init__(self,fichier_perm,fichier_secondaire,annee, feuil_perm=None, feuil_secondaire=None):
        self.fichier_perm=fichier_perm
        self.fichier_secondaire=fichier_secondaire
        self.annee=annee
        self.feuil_perm=feuil_perm
        self.feuil_secondaire=feuil_secondaire
    
    def ouvrir_cpt_perm_xls(self):
        """
        mettre en forme les comptages permannets
        """
        donnees_brutes_perm=pd.read_excel(self.fichier_perm)
        donnees_brutes_perm['route']=donnees_brutes_perm['AXE'].apply(lambda x : O.epurationNomRoute(x[3:]))
        donnees_brutes_perm['pr']=donnees_brutes_perm.apply(lambda x : int(x['CUMULD']//1000) if not pd.isnull(x['CUMULD']) else x['PLOD'], axis=1)
        donnees_brutes_perm['absc']=donnees_brutes_perm.apply(lambda x : int(x['CUMULD']%1000) if not pd.isnull(x['CUMULD']) else 0, axis=1)
        donnees_brutes_perm['id_gest']=donnees_brutes_perm.apply(lambda x : f"86-{x['route']}-{x['pr']}+{x['absc']}", axis=1)
        return donnees_brutes_perm

    def corresp_perm(self, table):
        """
        creer le dico de corresp des compt perm fouri en 2018 avec ceux existants
        """
        self.existant = compteur_existant_bdd(table, schema='comptage',dep='86')
        corresp=self.existant[['id_comptag','route','pr','abs']].merge(self.ouvrir_cpt_perm_xls(), left_on=['route','pr'], right_on=['route','PLOD'], how='right').sort_values('id_comptag')
        corresp['id_comptag']=corresp.apply(lambda x : x['id_comptag'] if not pd.isnull(x['id_comptag']) else f"86-{x['route']}-{x['pr_y']}+{x['absc']}", axis=1)
        #le dico de correspondance à inserer dans corresp_id_comptage
        corresp_id_comptag=corresp[['id_comptag', 'id_gest']].loc[corresp['id_comptag']!=corresp['id_gest']].rename(columns={'id_comptag':'id_gti'})
        return corresp_id_comptag
    
    def forme_cpt_perm_xls(self):
        #la future df_attr, avec que les perm
        donnees_brutes_perm=self.ouvrir_cpt_perm_xls()
        df_cpt_perm=donnees_brutes_perm[['id_gest','TMJA','POURCENTAGE_PL','TYPE_POSTE','route','pr','absc']].rename(columns={'id_gest':'id_comptag','TMJA':'tmja','POURCENTAGE_PL':'pc_pl','TYPE_POSTE':'type_poste'})
        df_cpt_perm['type_poste']='permanent'
        return df_cpt_perm

    def cpt_second_xls(self):
        """
        mettre en forme les comptages secondaires
        """  
        donnees_brutes_second=pd.read_excel(self.fichier_secondaire)
        donnees_brutes_second['route']=donnees_brutes_second.apply(lambda x : 'D'+str(x['RD']).strip(), axis=1)
        donnees_brutes_second['pr']=donnees_brutes_second.apply(lambda x : re.split('(\.|\+)',str(x['PR']))[0], axis=1)
        donnees_brutes_second['absc']=donnees_brutes_second.apply(lambda x : int(re.split('(\.|\+|,)',str(x['PR']))[2]+'0'*(3-len(re.split('(\.|\+|,)',str(x['PR']))[2]))) if re.search('(\.|\+)',str(x['PR'])) else 0, axis=1)
        donnees_brutes_second['id_comptag']=donnees_brutes_second.apply(lambda x : f"86-{x['route']}-{x['pr']}+{x['absc']}", axis=1)
        donnees_brutes_second['type_poste']='tournant'
        donnees_brutes_second['% PL']=donnees_brutes_second['% PL']*100
        df_cpt_second=donnees_brutes_second.rename(columns={'TV':'tmja','% PL':'pc_pl'}).drop(['LIEUX', 'RD','PR','n° de compteur','PL'], axis=1)     
        return  df_cpt_second
    
    def ouvrir_fichier_unique(self) : 
        donnees_brutes_perm=pd.read_excel(self.fichier_perm, sheet_name=self.feuil_perm)
        donnees_brutes_perm=donnees_brutes_perm.loc[(donnees_brutes_perm['ANNEE_TRAFIC']==self.annee) & (~donnees_brutes_perm['PLOD'].isna()) &
                            (~donnees_brutes_perm['ABSD'].isna()) & (~donnees_brutes_perm['TMJA'].isna()) & (~donnees_brutes_perm['POURCENTAGE_PL'].isna())].copy()
        donnees_brutes_perm['route']=donnees_brutes_perm['AXE'].apply(lambda x : O.epurationNomRoute(x[3:]))
        donnees_brutes_perm['pr_cumuld']=donnees_brutes_perm.apply(lambda x : int(x['CUMULD']//1000) if not pd.isnull(x['CUMULD']) else np.NaN, axis=1)
        donnees_brutes_perm['absc_cumuld']=donnees_brutes_perm.apply(lambda x : int(x['CUMULD']%1000) if not pd.isnull(x['CUMULD']) else np.NaN, axis=1)
        donnees_brutes_perm['id_gest_cumuld']=donnees_brutes_perm.apply(lambda x : f"86-{x['route']}-{x['pr_cumuld']}+{x['absc_cumuld']}", axis=1)
        donnees_brutes_perm['pr_plod']=donnees_brutes_perm.apply(lambda x : int(x['PLOD']) if not pd.isnull(x['PLOD']) else np.NaN, axis=1)
        donnees_brutes_perm['absc_absd']=donnees_brutes_perm.apply(lambda x : int(x['ABSD']) if not pd.isnull(x['ABSD']) else np.NaN, axis=1)
        donnees_brutes_perm['id_gest_plod']=donnees_brutes_perm.apply(lambda x : f"86-{x['route']}-{x['pr_plod']}+{x['absc_absd']}", axis=1)
        donnees_brutes_perm['POURCENTAGE_PL']=donnees_brutes_perm['POURCENTAGE_PL']*100
        donnees_brutes_second=pd.read_excel(self.fichier_perm, sheet_name='secondaires', names=donnees_brutes_perm.columns, header=None)
        donnees_brutes_second=donnees_brutes_second.loc[(donnees_brutes_second['ANNEE_TRAFIC']==2019) & (~donnees_brutes_second['PLOD'].isna()) &
                            (~donnees_brutes_second['ABSD'].isna()) & (~donnees_brutes_second['TMJA'].isna()) & (~donnees_brutes_second['POURCENTAGE_PL'].isna())].copy()
        donnees_brutes_second['route']=donnees_brutes_second['AXE'].apply(lambda x : O.epurationNomRoute(x[3:]))
        donnees_brutes_second['pr_cumuld']=donnees_brutes_second.apply(lambda x : int(x['CUMULD']//1000) if not pd.isnull(x['CUMULD']) else np.NaN, axis=1)
        donnees_brutes_second['absc_cumuld']=donnees_brutes_second.apply(lambda x : int(x['CUMULD']%1000) if not pd.isnull(x['CUMULD']) else np.NaN, axis=1)
        donnees_brutes_second['id_gest_cumuld']=donnees_brutes_second.apply(lambda x : f"86-{x['route']}-{x['pr_cumuld']}+{x['absc_cumuld']}", axis=1)
        donnees_brutes_second['pr_plod']=donnees_brutes_second.apply(lambda x : int(x['PLOD']) if not pd.isnull(x['PLOD']) else np.NaN, axis=1)
        donnees_brutes_second['absc_absd']=donnees_brutes_second.apply(lambda x : int(x['ABSD']) if not pd.isnull(x['ABSD']) else np.NaN, axis=1)
        donnees_brutes_second['id_gest_plod']=donnees_brutes_second.apply(lambda x : f"86-{x['route']}-{x['pr_plod']}+{x['absc_absd']}", axis=1)
        donnees_brutes_second['POURCENTAGE_PL']=donnees_brutes_second['POURCENTAGE_PL']*100
        donnees_concat=pd.concat([donnees_brutes_perm,donnees_brutes_second],axis=0, sort=False)
        dico_corresp_type_poste={'Permanent':'permanent', 'Tournant':'ponctuel', 'Secondaire':'tournant'}
        donnees_concat.TYPE_POSTE=donnees_concat.TYPE_POSTE.apply(lambda x : dico_corresp_type_poste[x])
        return donnees_brutes_perm,donnees_brutes_second, donnees_concat
    
    def equivalence_id_comptag(self,table_cpt, df) : 
        """
        pour conserver soit l'id_comptag issu du plod, soit celui issu du cumul, si il est présent dans la liste des id_coptage ou dans liste des id_cgest de la table corresp
        """
        rqt_corresp_comptg='select * from comptage.corresp_id_comptag'
        rqt_id_comtg=f"select * from comptage.{table_cpt} where dep='86'"
        with ct.ConnexionBdd(nomConnBddOtv) as c:
            list_corresp_comptg=pd.read_sql(rqt_corresp_comptg, c.sqlAlchemyConn).id_gest.tolist()
            list_id_comptg=gp.GeoDataFrame.from_postgis(rqt_id_comtg, c.sqlAlchemyConn, geom_col='geom',crs={'init': 'epsg:2154'}).id_comptag.tolist()
        
        def maj_id_comptg(id_cpt_plod,id_cpt_cumuld, list_id_comptag,list_id_gest) : 
            for a in  (id_cpt_plod,id_cpt_cumuld) : 
                if a in list_id_comptag or a in list_id_gest : 
                    return a
            else : return id_cpt_plod
    
        df['id_comptag']=df.apply(lambda x : maj_id_comptg(x['id_gest_plod'], x['id_gest_cumuld'], list_id_comptg,list_corresp_comptg ), axis=1)
    
    def comptage_forme(self, table_cpt) : 
        """
        creer le df_attr a partir des donnees secondaire et permanent
        2 cas de figuer pour le moment : en 2018 ajout des pr + abs issus de plod et abscd dans la liste de correspondance, en 2019 conservation des 2 colonnes et recherche si existants
        """
        if not self.feuil_perm : 
            self.df_attr=pd.concat([self.forme_cpt_perm_xls(),self.cpt_second_xls()], axis=0, sort=False)
            self.df_attr=self.df_attr.loc[~self.df_attr.isna().any(axis=1)].copy()
        else : 
            donnees_concat=self.ouvrir_fichier_unique()[2]
            self.equivalence_id_comptag(table_cpt, donnees_concat)
            self.df_attr=donnees_concat[['TMJA','POURCENTAGE_PL','TYPE_POSTE','route','id_comptag']].rename(columns={a:a.lower() for a 
                                        in donnees_concat[['TMJA','POURCENTAGE_PL','TYPE_POSTE','route','id_comptag']].columns}).rename(columns={'pourcentage_pl':'pc_pl'})
            self.df_attr['pr']=self.df_attr.id_comptag.apply(lambda x : int(x.split('-')[2].split('+')[0]))
            self.df_attr['absc']=self.df_attr.id_comptag.apply(lambda x : int(x.split('-')[2].split('+')[1]))
        self.df_attr['src']='tableur'
    
        
    def update_bdd_86(self, schema, table):
        """
        mettre à jour la table des comptages dans le 16
        """
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr_update,['id_comptag','tmja','pc_pl','src'])
        dico_attr_modif={f'tmja_{str(self.annee)}':'tmja', f'pc_pl_{str(self.annee)}':'pc_pl',f'src_{str(self.annee)}':'src'}
        self.update_bdd(schema, table, valeurs_txt,dico_attr_modif)
    
    def insert_bdd_86(self):   
        self.df_attr_insert['dep']='86'
        self.df_attr_insert['reseau']='RD'
        self.df_attr_insert['gestionnai']='CD86'
        self.df_attr_insert['concession']='N'
        self.df_attr_insert['obs']="nouveau_point 2019, denominationCD86 ='tournant'"
        self.df_attr_insert.rename(columns={'absc' : 'abs', 'tmja':'tmja_'+str(self.annee),'pc_pl':'pc_pl_'+str(self.annee),'obs':'obs_'+str(self.annee), 'src':'src_'+str(self.annee)},inplace=True)
        self.insert_bdd(schemaComptage,tableComptage, self.df_attr_insert)
        #mettre à jour la geom
        self.maj_geom(schemaComptage, tableComptage, dep='86')
        
        
class Comptage_cd24(Comptage):
    """
    pour le moment on ne traite que les cpt perm et tournants de 2020
    """ 
    def __init__(self,fichier_perm, annee):
        self.fichier_perm = fichier_perm
        Comptage.__init__(self, fichier_perm)
        self.annee=annee
    
    def cpt_perm_csv(self):
        def creerPeriode(serie):
            listPeriode = []
            for i in range(1,5):
                if not pd.isnull(serie[f'DDPeriode{i}']) and not pd.isnull(serie[f'DFPeriode{i}']):
                    listPeriode.append(f"{pd.to_datetime(serie[f'DDPeriode{i}'], dayfirst=True).strftime('%Y/%m/%d')}-{pd.to_datetime(serie[f'DFPeriode{i}'], dayfirst=True).strftime('%Y/%m/%d')}")
            if not listPeriode:
                return None
            return ' ; '.join(listPeriode)
        
        donnees_brutes = self.ouvrir_csv()
        donnees_traitees = donnees_brutes.loc[~donnees_brutes.MJA.isna()].copy()
        donnees_traitees['Data'] = donnees_traitees.Route.apply(lambda x: x.split(' ')[1] if not pd.isnull(x) else None)
        donnees_traitees = donnees_traitees.loc[donnees_traitees['Sens']=='3'].rename(columns={'MJA' : 'tmja', 'MJAPPL':'pc_pl'}).copy()
        donnees_traitees['id_comptag'] = donnees_traitees.apply(lambda x : f"24-{x['Data']}-{x['PRC']}+{x['ABC']}", axis=1)
        donnees_traitees['Latitude'] = donnees_traitees.Latitude.apply(lambda x : float(x.replace(',','.')))
        donnees_traitees['Longitude'] = donnees_traitees.Longitude.apply(lambda x : float(x.replace(',','.')))
        donnees_traitees['tmja'] = donnees_traitees.tmja.apply(lambda x : int(x))
        donnees_traitees['pc_pl'] = donnees_traitees.pc_pl.apply(lambda x : float(x.replace(',','.')))
        donnees_traitees.rename(columns={'Data':'route', 'PRC':'pr', 'ABC':'abs'}, inplace=True)
        gdf_finale = gp.GeoDataFrame(donnees_traitees, geometry=gp.points_from_xy(donnees_traitees.Longitude, donnees_traitees.Latitude), crs='epsg:4326')
        gdf_finale = gdf_finale.to_crs('epsg:2154')
        gdf_finale['type_poste'] = gdf_finale.apply(lambda x : 'permanent' if x['Type'].lower() == 'per' else 'tournant', axis=1)
        gdf_finale['geometry'] = gdf_finale.apply(lambda x : 0 if x['Latitude'] == 0 else x['geometry'], axis=1)
        gdf_finale['src'] = f'tableau export SIG {str(self.annee)}'
        gdf_finale['fichier'] = os.path.basename(self.fichier)
        gdf_finale['annee'] = self.annee
        gdf_finale.loc[gdf_finale.type_poste == 'tournant','periode'] = gdf_finale.apply(lambda x : creerPeriode(x), axis=1)
        gdf_finale.periode.fillna('', inplace=True)
        return gdf_finale
    
    def comptage_forme(self):
        donnees_finales = self.cpt_perm_csv()
        #mensuel permanent
        donnees_mens_perm = donnees_finales.loc[donnees_finales.type_poste == 'permanent'][[a for a in donnees_finales.columns if a in [m for e in dico_mois.values() for m in e]]+['id_comptag']].copy()
        donnees_mens_perm.rename(columns={c:k for c in donnees_mens_perm.columns if c != 'id_comptag' for k, v in dico_mois.items() if c in v}, inplace=True)
        donnees_mens_perm['donnees_type'] = 'tmja'
        #mensuel tournant
        donnees_mens_tour = pd.DataFrame({'id_comptag':[],'donnees_type':[]})
        for j,e in enumerate(donnees_finales.loc[donnees_finales.type_poste == 'tournant'].itertuples()) : 
            for  i in range(1,5):
                if not pd.isnull(getattr(e,f'DDPeriode{i}')) and not pd.isnull(getattr(e,f'DFPeriode{i}')):
                    mois = [k for k, v in dico_mois.items() if v[0] == Counter(pd.date_range(pd.to_datetime(getattr(e,f'DDPeriode{i}'), dayfirst=True)
                                                            ,pd.to_datetime(getattr(e,f'DFPeriode{i}'), dayfirst=True)).month).most_common()[0][0]][0]
                    donnees_mens_tour.loc[j,mois] = getattr(e,f'MJP{i}')
                    donnees_mens_tour.loc[j,'donnees_type'] = 'tmja'
                    donnees_mens_tour.loc[j+1000,mois] = float(getattr(e,f'MJPPL{i}').replace(',','.'))
                    donnees_mens_tour.loc[j,'id_comptag'] = getattr(e,f'id_comptag')
                    donnees_mens_tour.loc[j+1000,'id_comptag'] = getattr(e,f'id_comptag')
                    donnees_mens_tour.loc[j+1000,'donnees_type'] = 'pc_pl'
        #global        
        self.df_attr_mens = pd.concat([donnees_mens_tour,donnees_mens_perm],axis=0).reset_index(drop=True)
        self.df_attr = donnees_finales[['id_comptag', 'tmja', 'pc_pl','route', 'pr', 'abs','src', 'geometry', 'type_poste', 'fichier', 'periode']].copy()
        
    def classer_compteur_update_insert(self):
        """
        attention, on aurait pu ajouter un check des comptages deja existant et rechercher les correspondances comme dans le 87,
        mais les données PR de l'IGN sont trop pourries dans ce dept, dc corresp faite à la main en amont
        """
        corresp_nom_id_comptag(self.df_attr)
        self.existant = compteur_existant_bdd(dep='24')
        self.df_attr_update = self.df_attr.loc[self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        self.df_attr_insert = self.df_attr.loc[~self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        """
        #on peut tenter un dico de correspondance, mais les données PR de l'IGN sont trop fausses pour faire confaince
        dico_corresp = cd24.corresp_old_new_comptag('local_otv_station_gti', 'public','cd24_perm', 'lineaire.traf2017_bdt24_ed17_l',
             'referentiel','troncon_route_bdt24_ed17_l','troncon_route_bdt24_ed17_l_vertices_pgr','id')
        """
 
class Comptage_cd33(Comptage):
    """
    Cette classe se base sur la fourniture par le CD33 : 
    en 2019 : de tableau des compteurs permanents et tournants, associé au sectionnement.
    en 2020 : de tableau des compteurs permanents et tournants et enquete, associé au sectionnement.
    en 2021 : de tableau des compteurs permanents et tournant + fichiers shape format points permanents, tournants, enquetes
    A noter que Vincent récupérait aussi des données sur le site internet du CD33 https://www.gironde.fr/deplacements/les-routes-et-ponts#comptage-routier
    """ 
    def __init__(self, dossier, fichierPermanentExcel, fichierTournantExcel, fichierPermanentShape,
                 fichierTournantShape, fichierEnqueteShape, annee, epsg='EPSG:2154'):
        """
        attributs : 
            dossier : dossier contenant les fichiers excel et géoréférencés
            fichierPermanentExcel : nom du fichier (sans le chemin) excel qui contient les comptages permanents
            fichierTournantExcel : nom du fichier (sans le chemin) excel qui contient les comptages tournants
            fichierPermanentShape : nom du fichier (sans le chemin) shape qui contient les point de comptages permanents
            fichierTournantShape : nom du fichier (sans le chemin) shape qui contient les point de comptages tournants
            fichierEnqueteShape : nom du fichier (sans le chemin) shape qui contient les point de comptages enquetes (geoloc + datas)
            epsg : string : de la forme 'EPSG:numEpsg'
            annee : string : annee sur 4 caractères
        pour info
        attributs olds: 
            self.fichier=fichier fichier : string : chemin du tableur des comptages permanents et tournants
            self.annee=annee annee : int : annee dur 4 characters
            self.sectionnement=sectionnement sectionnement : chemin du fichier de sectionnement
        """
        self.fichierPermanentExcel = fichierPermanentExcel
        self.fichierTournantExcel = fichierTournantExcel
        self.fichierPermanentShape = fichierPermanentShape
        self.fichierTournantShape = fichierTournantShape
        self.fichierEnqueteShape = fichierEnqueteShape
        self.epsg = epsg
        self.annee = annee
        
    
    def ouvertureFichier(self):
        """
        ouverture des fchiers excel est shape
        """    
        dfPermanentExcel = pd.read_excel(self.fichierPermanentExcel)
        dfTournantExcel = pd.read_excel(self.fichierTournantExcel, converters={'ddpériode1': lambda x: pd.to_datetime(x, dayfirst=True),
                                                                               'dfpériode1': lambda x: pd.to_datetime(x, dayfirst=True),
                                                                               'ddpériode2': lambda x: pd.to_datetime(x, dayfirst=True),
                                                                               'dfpériode2': lambda x: pd.to_datetime(x, dayfirst=True),
                                                                               'ddpériode3': lambda x: pd.to_datetime(x, dayfirst=True),
                                                                               'dfpériode3': lambda x: pd.to_datetime(x, dayfirst=True),
                                                                               'ddpériode4': lambda x: pd.to_datetime(x, dayfirst=True),
                                                                               'dfpériode4': lambda x: pd.to_datetime(x, dayfirst=True)})
        gdfPermanentShape = gp.read_file(self.fichierPermanentShape, crs=self.epsg)
        gdfTournantShape = gp.read_file(self.fichierTournantShape, crs=self.epsg)
        gdfEnqueteShape = gp.read_file(self.fichierEnqueteShape, crs=self.epsg)
        return dfPermanentExcel, dfTournantExcel, gdfPermanentShape, gdfTournantShape, gdfEnqueteShape
    
    
    def miseEnFormeExcel(self, dfPermanentExcel, dfTournantExcel):
        """
        filtrer les colonnes, les renommer, vérifier les doublons, mettre en forme la période des comptages tournants
        in : 
            dfPermanentExcel : dataframe, voir ouvertureFichier()
            dfTournantExcel : dataframe, voir ouvertureFichier()
        out : 
            dfPermExcelFiltre
            dfTournExcelFiltre
        """
        dfTournantExcelColsOk = dfTournantExcel.drop('annee', axis=1, errors='ign').drop(
            [c for c in dfTournantExcel.columns if c not in [e.lower() for e in cd33_dicoAttrTournExcel.keys()] + [
                v.lower() for v in cd33_dicoAttrTournExcel.values()]], axis=1, errors='ignore').rename(columns={k.lower(): v for k, v in cd33_dicoAttrTournExcel.items()})
        dfPermanentExcelColsOk = dfPermanentExcel.drop(
            [c for c in dfPermanentExcel.columns if c not in cd33_dicoAttrPermExcel.keys()], axis=1, errors='ignore').rename(
            columns={k: v for k, v in cd33_dicoAttrPermExcel.items()})
        # limitation des données à l'année en cours, sur le sens désiré, sans valeur nulle dans le tmja
        dfPermExcelFiltre = dfPermanentExcelColsOk.loc[(dfPermanentExcelColsOk.annee == int(self.annee)) & (dfPermanentExcelColsOk.sens == 3) &
                                                       (dfPermanentExcelColsOk.tmja.notna())].copy()
        dfTournExcelFiltre = dfTournantExcelColsOk.loc[(dfTournantExcelColsOk.annee == int(self.annee)) & (dfTournantExcelColsOk.sens == 3) &
                                                       (dfTournantExcelColsOk.tmja.notna())].copy()
        # vérification des doublons
        if not dfPermExcelFiltre.loc[dfPermExcelFiltre.duplicated('troncon')].empty:
            raise ValueError('doublons dans le fichier excel des comptages permanent')
        if not dfTournExcelFiltre.loc[dfTournExcelFiltre.duplicated('troncon')].empty:
            raise ValueError('doublons dans le fichier excel des comptages tournants')
        # mise en forme de la periode pour les comptages tournants
        dfTournExcelFiltre['periode'] = dfTournExcelFiltre.apply(
            lambda x: " ; ".join(['-'.join([c[0].strftime('%Y/%m/%d'),c[1].strftime('%Y/%m/%d')]) 
                                  for c in [[x.debut_periode1, x.fin_periode1], [x.debut_periode2, x.fin_periode2],
                                            [x.debut_periode3, x.fin_periode3], [x.debut_periode4, x.fin_periode4]]
                                  if not pd.isnull(c).all()])
            if not all([pd.isnull(c).all() for c in [[x.debut_periode1, x.fin_periode1], [x.debut_periode2, x.fin_periode2],
                                                     [x.debut_periode3, x.fin_periode3], [x.debut_periode4, x.fin_periode4]]])
            else None, axis=1)
        dfTournExcelFiltre['fichier'] = os.path.basename(self.fichierTournantExcel)
        dfPermExcelFiltre['fichier'] = os.path.basename(self.fichierPermanentExcel)
        return dfPermExcelFiltre, dfTournExcelFiltre
    
    
    def miseEnFormeShape(self, gdfPermanentShape, gdfTournantShape, gdfEnqueteShape):
        """
        mettre en forme les attributs des fichiers shape
        in : 
            gdfPermanentShape : dataframe, voir ouvertureFichier()
            gdfTournantShape : dataframe, voir ouvertureFichier()
            gdfEnqueteShape : dataframe, voir ouvertureFichier()
        out : 
            gdfPermanentFiltre
            gdfTournantFiltre
            gdfEnqueteFiltre
            gdfEnqueteColsOk
        """
        gdfPermanentColsOk = gdfPermanentShape.drop([c for c in gdfPermanentShape.columns if c not in cd33_dicoAttrPermTournShape.keys()
                                                     ], axis=1, errors='ignore').rename(columns=cd33_dicoAttrPermTournShape
                                                                                        ).assign(type_poste='permanent', src='shape + excel')
        gdfTournantColsOk = gdfTournantShape.drop([c for c in gdfTournantShape.columns if c not in cd33_dicoAttrPermTournShape.keys()
                                                   ], axis=1, errors='ignore'
                                                  ).rename(columns=cd33_dicoAttrPermTournShape).assign(
                                                      type_poste='tournant', src='shape + excel')
        gdfEnqueteColsOk = gdfEnqueteShape.drop(
            [c for c in gdfEnqueteShape if c not in list(cd33_dicoAttrPermTournShape.keys()) + list(cd33_dicoAttrEnqueteShape.keys())],
            axis=1, errors='ignore').rename(columns=cd33_dicoAttrPermTournShape).rename(columns=cd33_dicoAttrEnqueteShape).assign(
                                                    type_poste='ponctuel', src='enquete')
        for e in (gdfPermanentColsOk, gdfTournantColsOk, gdfEnqueteColsOk):
            if e.empty:
                continue
            e['route'] = e.route.apply(lambda x: O.epurationNomRoute(x) if not pd.isnull(x) else None)
            e['id_comptag'] = e.apply(lambda x: f'33-{x.route}-{int(x.pr)}+{int(x["abs"])}'
                                      if x[['route', 'pr', 'abs']].notna().all().all() else None, axis=1)
            e['dep'] = '33'
            e['reseau'] = 'RD'
            e['gestionnai'] = 'CD33'
            e['src_geo'] = 'coordonnees_gestionnaire'
            e['convention'] = True
            e['src_cpt'] = 'convention gestionnaire'
            e['techno'] = e.techno.replace(cd33_dicoCorrespTechno)
            e['concession'] = False
            if 'vma' in e.columns:
                e['vma'] = e.vma.apply(lambda x: int(x.split('/')[0]) if '/' in x else None)
            if 'troncon' in e.columns:
                e['id_sect'] = e.troncon
            e['obs_supl'] = e.rattacheme.apply(lambda x: f"rattaché à l'id_cpt {x}" if not pd.isnull(x) else None)
            # print(e)
            e['identifiant_modif'] = e['id_cpt'].str[:-2]
            dfTronconValueCount = e.drop_duplicates(['identifiant_modif', 'sens']).identifiant_modif.value_counts(
                ).rename('nb_sens').reset_index().rename(columns={'index': 'identifiant_modif'})
            e.loc[e.identifiant_modif.isin(dfTronconValueCount.loc[
                dfTronconValueCount.nb_sens >= 3].identifiant_modif.tolist()), 'sens_cpt'] = 'double sens'
            e.loc[e.identifiant_modif.isin(dfTronconValueCount.loc[
                dfTronconValueCount.nb_sens < 3].identifiant_modif.tolist()), 'sens_cpt'] = 'sens unique'
        if not gdfPermanentColsOk.empty:
            gdfPermanentFiltre = gdfPermanentColsOk.loc[(gdfPermanentColsOk.sens == 3) & (gdfPermanentColsOk.geometry.notna()) &
                                                        (gdfPermanentColsOk.id_comptag.notna())].copy()
            gdfPermanentFiltre = O.gp_changer_nom_geom(gdfPermanentFiltre, 'geom')
        else:
            gdfPermanentFiltre = None
        if not gdfTournantColsOk.empty: 
            gdfTournantFiltre = gdfTournantColsOk.loc[(gdfTournantColsOk.sens == 3) & (gdfTournantColsOk.geometry.notna()) & 
                                                      (gdfTournantColsOk.id_comptag.notna())].copy()
            gdfTournantFiltre = O.gp_changer_nom_geom(gdfTournantFiltre, 'geom')
        else:
            gdfTournantFiltre = None       
        if not gdfEnqueteColsOk.empty:                                                                         
            gdfEnqueteFiltre = gdfEnqueteColsOk.loc[(gdfEnqueteColsOk.sens == 3) & (gdfEnqueteColsOk.geometry.notna()) &
                                                    (gdfEnqueteColsOk.annee == int(self.annee)) &
                                                    (gdfEnqueteColsOk.id_comptag.notna())].copy()
            # periode pour les enquetes
            gdfEnqueteFiltre['periode'] = gdfEnqueteFiltre.apply(
                lambda x: f"{pd.to_datetime(x.debut_periode).strftime('%Y/%m/%d')}-{pd.to_datetime(x.fin_periode).strftime('%Y/%m/%d')}"
                if not pd.isnull([x.debut_periode, x.fin_periode]).all() else None, axis=1)
            gdfEnqueteFiltre = O.gp_changer_nom_geom(gdfEnqueteFiltre, 'geom')
            gdfEnqueteFiltre['fichier'] = os.path.basename(self.fichierEnqueteShape)
        else:
            gdfEnqueteFiltre = None
        return gdfPermanentFiltre, gdfTournantFiltre, gdfEnqueteFiltre, gdfEnqueteColsOk
    
    
    def miseEnFormeTournant(self, dfTournExcelFiltre):
        """
        passer les données de trafic relatives à des périodes vers la structure en mois et indicateurs
        in :
            dfTournExcelFiltre : dataframe, issue de miseEnFormeExcel()
        """
        donnees_mens_tour = pd.DataFrame({'troncon': [], 'donnees_type':[]})
        for j, e in enumerate(dfTournExcelFiltre.itertuples()):
            for i in range(1, 5):
                deb = getattr(e, f'debut_periode{i}')
                fin = getattr(e, f'fin_periode{i}')
                if pd.isnull(deb) or pd.isnull(fin):
                    continue
                mois = [k for k, v in dico_mois.items() if v[0] == Counter(pd.date_range(deb, fin).month).most_common()[0][0]][0]
                # print(e)
                donnees_mens_tour.loc[j, mois] = getattr(e, f'tmja_periode{i}')
                donnees_mens_tour.loc[j, 'donnees_type'] = 'tmja'
                donnees_mens_tour.loc[j+1000, mois] = getattr(e, f'pc_pl_periode{i}')
                donnees_mens_tour.loc[j, 'troncon'] = getattr(e, f'troncon')
                donnees_mens_tour.loc[j+1000, 'troncon'] = getattr(e, f'troncon')
                donnees_mens_tour.loc[j+1000, 'donnees_type'] = 'pc_pl'
        return donnees_mens_tour
    
    
    def syntheseNouvellesDonnees(self, dfPermExcelFiltre, gdfPermanentFiltre, dfTournExcelFiltre, gdfTournantFiltre,
                                 gdfEnqueteFiltre, gdfEnqueteColsOk):
        """
        rassembler les données des fichiers excel et shape dans une seule dataframe.
        Attention ! les données mensuelles tournants ne sont pas jointes, car pas la même structure de données
        """
        gdfPermTrafic = dfPermExcelFiltre.merge(gdfPermanentFiltre, on='troncon')
        if len(gdfPermTrafic) != len(dfPermExcelFiltre):
            raise ValueError(
                f"""les nombres de données géoréférencées avec id_comptag {len(gdfPermTrafic)} et excel {len(dfPermExcelFiltre)} aprés jointure est différent. 
                chercher doublons ou manque ou pb id_comptag""")
        gdfTournTrafic = dfTournExcelFiltre.merge(gdfTournantFiltre, on='troncon')
        if len(gdfTournTrafic) != len(dfTournExcelFiltre):
            raise ValueError(
                f"""les nombres de données géoréférencées {len(gdfTournTrafic)} et excel {len(dfTournExcelFiltre)} aprés jointure est différent. 
                chercher doublons ou manque ou pb id_comptag""")
        if len(gdfEnqueteFiltre) != len(gdfEnqueteColsOk.loc[(gdfEnqueteColsOk.sens == 3) & (gdfEnqueteColsOk.geometry.notna()) & 
                                                             (gdfEnqueteColsOk.annee == int(self.annee))]):
            raise ValueError(
                f"""les nombres de données géoréférencées {len(gdfEnqueteFiltre)} et excel {
                len(gdfEnqueteColsOk.loc[(gdfEnqueteColsOk.sens == 3) & (gdfEnqueteColsOk.geometry.notna()) 
                & (gdfEnqueteColsOk.annee == int(self.annee))])} 
                aprés jointure est différent. chercher doublons ou manque ou pb id_comptag""")
        dfTtSources = gp.GeoDataFrame(pd.concat([e.drop([
            c for c in e.columns if c not in attBddCompteur + attrComptage + attrIndicAgrege + list(dico_mois.keys()) + 
            enumIndicateur + ['troncon', 'libelle']], axis=1, errors='ignore') for e in (gdfPermTrafic, gdfTournTrafic, gdfEnqueteFiltre)
            ]), geometry='geom', crs=self.epsg)
        return dfTtSources
    
    
    def miseEnFormeGenerale(self):
        """
        chainer les fonctions de mise en forme préalables des données
        """ 
        dfPermanentExcel, dfTournantExcel, gdfPermanentShape, gdfTournantShape, gdfEnqueteShape = self.ouvertureFichier()
        dfPermExcelFiltre, dfTournExcelFiltre = self.miseEnFormeExcel(dfPermanentExcel, dfTournantExcel)
        gdfPermanentFiltre, gdfTournantFiltre, gdfEnqueteFiltre, gdfEnqueteColsOk = self.miseEnFormeShape(
            gdfPermanentShape, gdfTournantShape, gdfEnqueteShape)
        donnees_mens_tour = self.miseEnFormeTournant(dfTournExcelFiltre)
        dfTtSources = self.syntheseNouvellesDonnees(dfPermExcelFiltre, gdfPermanentFiltre, dfTournExcelFiltre, gdfTournantFiltre,
                                         gdfEnqueteFiltre, gdfEnqueteColsOk)
        return dfTtSources, donnees_mens_tour
        
        
        
    def old_analysePerm(self, sheet_name='Permanents'):
        perm=pd.read_excel(self.fichier, sheet_name=sheet_name)
        gdfPerm = gp.GeoDataFrame(perm, geometry=gp.points_from_xy(perm.Longitude, perm.Latitude), crs='EPSG:4326')
        gdfPerm=gdfPerm.to_crs('EPSG:2154')
        gdfPerm.rename(columns={'Tronçon':'troncon','Data':'route'},inplace=True )
        gdfPerm.rename(columns={c:c.replace('MJA TV TCJ ','tmja_') if 'TV' in c else c.replace('MJA %PL TCJ ','pc_pl_')  for c in gdfPerm.columns if any([a in c for a in ('MJA TV TCJ','MJA %PL TCJ')])},inplace=True)
        gdfPerm.rename(columns={c:[k for k, v in dico_mois.items() if int(c.split()[3][:2]) in v][0] for c in gdfPerm.columns if 'MJM TV TCJ' in c}, inplace=True)
        gdfPerm=gdfPerm.loc[~gdfPerm.tmja_2019.isna()].copy()
        gdfPerm['fichier']=self.fichier
        gdfPerm[f'src_{self.annee}']='tableur cpt permanent / tournant'
        gdfPerm['convention']=True
        gdfPerm['type_poste']='permanent'
        return gdfPerm
    
    
    def old_trierPermConnus(self, table, localisation):
        """
        trouver les comptages permanents connus a partir de l'annee precedente
        """
        gdfPerm=self.analysePerm()
        #on va récupérer les données existantes : 
        self.existant = compteur_existant_bdd(table, schema='comptage',localisation=localisation,dep='33', gest='CD33')
        #et tenter une jointure sur les donnees 2018 :  tous y sont, donc c'est bon
        GdfPerm=gdfPerm[['troncon','route', f'tmja_{self.annee}', f'tmja_{self.annee-1}', f'pc_pl_{self.annee}', 'geometry',
                                  'fichier',f'src_{self.annee}','convention','type_poste']+[k for k in dico_mois.keys()]].merge(
            self.existant.loc[~self.existant[f'tmja_{self.annee-1}'].isna()][['id_comptag', f'tmja_{self.annee-1}']], how='left')
        gdfPermConnus=GdfPerm.loc[~GdfPerm.id_comptag.isna()].copy()
        gdfPermInconnus=GdfPerm.loc[GdfPerm.id_comptag.isna()].copy()
        return GdfPerm,gdfPermConnus, gdfPermInconnus
   
    
    def old_assignerCptInconnus(self, dicoAssigne, gdfInconnus):
        """
        assigner manuellement les comptages inconnus, en place dans la df
        in  : 
            dicoAssigne : dico avec en cle l'identificant de troncon de la df des cpt perm inconnu en vlaue la valeur d'id_comptag
            gdfInconnus : df des permanents inconnus isue de trierPermConnus
        """
        for k, v in dicoAssigne.items() : 
            gdfInconnus.loc[gdfInconnus.troncon==k,'id_comptag']=v
            
            
    def old_analyseTourn(self, sheet_name='Tournants'):
        perm=pd.read_excel(self.fichier, sheet_name=sheet_name)
        gdfTourn = gp.GeoDataFrame(perm, geometry=gp.points_from_xy(perm.Longitude, perm.Latitude), crs='EPSG:4326')
        gdfTourn=gdfTourn.to_crs('EPSG:2154')
        gdfTourn.rename(columns={'Tronçon':'troncon','Data':'route'},inplace=True )
        gdfTourn.rename(columns={c:c.replace('MJA TV TCJ ','tmja_') if 'TV' in c else c.replace('MJA %PL TCJ ','pc_pl_')  for c in gdfTourn.columns if any([a in c for a in ('MJA TV TCJ','MJA %PL TCJ')])},inplace=True)
        #gdfTourn.rename(columns={c:[k for k, v in dico_mois.items() if int(c.split()[3][:2]) in v][0] for c in gdfTourn.columns if 'MJM TV TCJ' in c}, inplace=True)
        gdfTourn=gdfTourn.loc[~gdfTourn.tmja_2019.isna()].copy()
        gdfTourn[f'obs_{self.annee}']=gdfTourn.apply(lambda x : ' ; '.join([f"{x[f'DDPériode{i}'].date()}-{x[f'DFPériode{i}'].date()}" for i in range(1,5) if not (pd.isnull(x[f'DDPériode{i}']) and pd.isnull(x[f'DFPériode{i}']))]), axis=1)
        gdfTourn['fichier']=self.fichier
        gdfTourn[f'src_{self.annee}']='tableur cpt permanent / tournant'
        gdfTourn['convention']=True
        gdfTourn['type_poste']='tournant'
        return gdfTourn
    
    
    def old_donneesMensTournant(self,gdfTourn):
        """
        ajouter les donnees mensuelles aux donnees de comptage tournant et renvoyer une nouvelle df
        """
        donnees_mens_tour=pd.DataFrame({'troncon':[],'donnees_type':[]})
        for j,e in enumerate(gdfTourn.itertuples()) : 
            for  i in range(1,5):
                try :
                    mois=[k for k, v in dico_mois.items() if v[0]==Counter(pd.date_range(pd.to_datetime(getattr(e,f'DDPériode{i}'), dayfirst=True)
                                                        ,pd.to_datetime(getattr(e,f'DDPériode{i}'), dayfirst=True)).month).most_common()[0][0]][0]
                except ValueError : 
                    continue
                #print(e)
                donnees_mens_tour.loc[j,mois]=getattr(e,f'MJPTV{i}')
                donnees_mens_tour.loc[j,'donnees_type']='tmja'
                donnees_mens_tour.loc[j+1000,mois]=round(getattr(e,f'MJPPL{i}')/getattr(e,f'MJPTV{i}')*100,2)
                donnees_mens_tour.loc[j,'troncon']=getattr(e,f'troncon')
                donnees_mens_tour.loc[j+1000,'troncon']=getattr(e,f'troncon')
                donnees_mens_tour.loc[j+1000,'donnees_type']='pc_pl'
        return donnees_mens_tour
            
            
    def old_correspondanceTournant(self, localisation, rqtPpvCptTournant, rqtCptExistant, rqtGeomReferentielEpure, rqtPpvCptBdd,
                               schema, table_graph, table_vertex, dfTournant ):
        """"
        Comme on a aucun PR + abs, on fait une correspondanec en cherchant les id_ign commun avec une determination de troncon elementaire,
        un peu commme pour le 87 mais en ayant simplifie la table de recherche au prealable
        """
        #fonction d'association en fonction du dico de correspondance
        def pt_corresp(id_ign_lin,id_ign_cpt_new,dico_corresp) : 
            if id_ign_lin in dico_corresp[id_ign_cpt_new] : 
                return True
            else : return False
            
        with ct.ConnexionBdd(localisation) as c : 
            list_lgn=pd.read_sql(rqtPpvCptTournant, c.sqlAlchemyConn).id_ign.tolist()
            cpt_existant=pd.read_sql(rqtCptExistant, c.sqlAlchemyConn)
            dfDepartementale=gp.read_postgis(rqtGeomReferentielEpure, c.sqlAlchemyConn)
            ppvIdComptag=pd.read_sql(rqtPpvCptBdd, c.sqlAlchemyConn)
        
        #dico des troncons elementaire
        simplification=self.troncon_elemntaires(schema, table_graph, table_vertex,liste_lignes=list_lgn,id_name='id')
        #plus proche voisin des comptages tournants
        ppvTournIgn=O.plus_proche_voisin(dfTournant[['troncon','geometry']],dfDepartementale,20,'troncon','id_ign' )
        #synthese des plus proche voisin
        ppv_final=ppvTournIgn.merge(cpt_existant, on='id_ign', how='left').merge(ppvIdComptag, on='id_comptag', how='left').rename(columns={'id_ign_x':'id_ign_cpt_new', 'id_ign_y':'id_ign_lin'})
        #ajout d'un attribut d'identification des correpsondance
        ppv_final['correspondance']=ppv_final.apply(lambda x : pt_corresp(x['id_ign_lin'],x['id_ign_cpt_new'],simplification),axis=1)
        ppv_final.drop_duplicates(['troncon','id_comptag'], inplace=True)
        #identification
        correspTournant=ppv_final.loc[ppv_final['correspondance']]
        inconnuTournant=ppv_final.loc[~ppv_final['correspondance']]
        return correspTournant,inconnuTournant
    
    
    def old_correctionCorrespondanceTournant(self,dicoCorrespTourn,tournant,correspTournant):
        """
        suite a la determination des correspondance de comptage tournants, on ajoute les correction manuelle et on produit
        la table de synthese : 
        in :
           dicoCorrespTourn : dico avec en cle l'identificant de troncon de la df des cpt perm inconnu en vlaue la valeur d'id_comptag
           tournant : df issue de analyseTourn()
           correspTournant : df issue de correspondanceTournant
        """
        cptTournantAffecte=tournant.merge(correspTournant[['troncon','id_comptag','correspondance']], on='troncon', how='left')
        self.assignerCptInconnus(dicoCorrespTourn,cptTournantAffecte)
        cptTournantAffecte.loc[~cptTournantAffecte.id_comptag.isna(),'correspondance']=True
        cptTournantAffecte.loc[cptTournantAffecte.id_comptag.isna(),'correspondance']=False
        return cptTournantAffecte
    
    
    def old_creerNouveauPointTournants(self,cptTournantAffecte,dicoNewCpt):
        """
        apres xorrection de la correspondance, creer les nouveaux points a inserer
        in : 
            dicoNewCpt : dico avec en cle le numero de troncon, le pr, l'abscisse et en valeur les listes de valeur
        """
        cptTournantInconnu=cptTournantAffecte.loc[cptTournantAffecte.correspondance==False]
        df_attr_insert=cptTournantInconnu.merge(pd.DataFrame(dicoNewCpt).assign(src_geo='tableur millesime 2019', dep='33', reseau='RD', 
                                                gestionnai='CD33', concession='N', obs_geo='pr et abscisse aproximatifs'), on='troncon')
        df_attr_insert['route']=df_attr_insert.route.apply(lambda x : O.epurationNomRoute(x))
        df_attr_insert['id_comptag']=df_attr_insert.apply(lambda x : f"33-{x['route']}-{x['pr']}+{x['absc']}", axis=1)
        df_attr_insert['x_l93']=round(df_attr_insert.geometry.x,3)
        df_attr_insert['y_l93']=round(df_attr_insert.geometry.x,3)
        df_attr_insert[f'obs_{self.annee}']='nouveau_point, '+df_attr_insert[f'obs_{self.annee}']
        return df_attr_insert[['id_comptag','route','type_poste','src_geo','obs_geo','dep','reseau','gestionnai','concession','x_l93','y_l93',f'tmja_{self.annee}',
                f'pc_pl_{self.annee}', f'obs_{self.annee}', f'src_{self.annee}', 'fichier', 'convention', 'geometry', 'troncon']]
 
  
class Comptage_cd79(Comptage):
    """
    données issues d'un fichier SIG et concernant des comptages tournants
    attributs : 
        fichier : raw string du nom complet du fichier source
        annee : sur 4 caractère : année de référence des comptages
    """
    
    
    def __init__(self, fichier, annee):
        self.fichier = fichier
        self.annee = annee
        self.df_attr = self.ouvertureMiseEnForme()
    
    def ouvertureMiseEnForme(self):
        """
        ouvrir et mettre en forme les données SIG
        """
        gdfBrute = gp.read_file(self.fichier)
        gdfAnnee = gdfBrute.loc[gdfBrute.Annee_der_ == self.annee].copy()
        gdfAnnee['pc_pl'] = gdfAnnee.PL / gdfAnnee.TMJA *100
        gdfAnnee['id_comptag'] = gdfAnnee.apply(lambda x: f"79-{x.Axe}-{x.PR_dernie_.split('.')[0]}+{int(x.PR_dernie_.split('.')[1])}",
                                                axis=1)
        gdfAnnee['date_deb'] = gdfAnnee.Semaine.apply(lambda x: dt.datetime.strptime(f"2021-W{int(x)}-1", '%G-W%V-%u'
                                                                                     ).strftime('%Y/%m/%d'))
        gdfAnnee['date_fin'] = gdfAnnee.Semaine.apply(lambda x: dt.datetime.strptime(f"2021-W{int(x)}-7", '%G-W%V-%u'
                                                                                     ).strftime('%Y/%m/%d'))
        gdfAnnee['periode'] = gdfAnnee.date_deb + '-' + gdfAnnee.date_fin
        gdfAnnee['fichier'] = os.path.basename(self.fichier)
        gdfAnnee.Type_mate_.replace(cd79_dicoCorrespMaterielTechno, inplace=True)
        gdfAnnee.rename(columns={'TMJA': 'tmja', 'Type_mate_': 'techno'}, inplace=True)
        return gdfAnnee
    
    
    def miseAJourTechno(self, dfSource):
        """
        Mettre a jour la table des compteur Bdd, attribut techno, avec les valeurs aclcul lors de l'ouvereture
        in : 
            dfSource : la df qui contient les données. Normalement c'est self.df_attr
        """
        # mettre à jour la techno
        with ct.ConnexionBdd() as c:
            dfSource.drop('geometry', axis=1).to_sql(f'cd79_{self.annee}_temp', c.sqlAlchemyConn, schema='public')
            rqtMaJTechno = f"update comptage.compteur ce set techno = cd.techno from public.cd79_{self.annee}_temp cd where cd.id_comptag = ce.id_comptag"
            c.curs.execute(rqtMaJTechno)
            c.connexionPsy.commit()
        
    
class Comptage_vinci(Comptage):
    """
    inserer les donnees de comptage de Vinci
    POur info il y a une table de correspondance entre les donnees fournies par Vinci et les notre dans la base otv, scham source table asf_otv_tmja_2017
    """  
    def __init__(self,fichier_perm, annee):
        self.fichier_perm=fichier_perm
        self.annee=annee
        self.comptage_forme()
    
    def ouvrir_fichier(self):
        donnees_brutes=pd.read_excel(self.fichier_perm).rename(columns={'(*) PR début':'pr_deb'})
        donnees_brutes=donnees_brutes.loc[~donnees_brutes.pr_deb.isna()][:-1].copy()
        donnees_brutes['pr_deb']=donnees_brutes.pr_deb.astype(float)
        return donnees_brutes
    
    def importer_donnees_correspondance(self):
        with ct.ConnexionBdd(nomConnBddOtv) as c :
            rqt=f"""select * from source.asf_otv_tmja_2017"""
            base=pd.read_sql(rqt, c.sqlAlchemyConn).rename(columns={'(*) PR début':'pr_deb'})
            base=base.loc[~base.pr_deb.isna()].copy()
            base['pr_deb']=base.pr_deb.apply(lambda x : float(x.strip()))
        return base
    
    def comptage_forme(self):
        donnees_brutes=self.ouvrir_fichier()
        base=self.importer_donnees_correspondance() 
        self.df_attr=donnees_brutes[['pr_deb',f'TMJA {str(self.annee)}',f'Pc PL {str(self.annee)}', 'Vitesse moyenne annuelle']].merge(base[['pr_deb','ID']], on='pr_deb', how='left').rename(columns=
                                                            {'ID':'id_comptag',f'TMJA {str(self.annee)}':'tmja',f'Pc PL {str(self.annee)}':'pc_pl', 'Vitesse moyenne annuelle':'vmoy'})
        self.df_attr['pc_pl']=self.df_attr.pc_pl.apply(lambda x : round(x,2))
        self.df_attr['src']='tableur'
        self.df_attr['fichier']=os.path.basename(self.fichier_perm)
        self.df_attr['annee']=self.annee
        
        tmjm_mens=donnees_brutes[[a for a in donnees_brutes.columns if a =='pr_deb' or 'TMJM '+str(self.annee) in a]].merge(base[['pr_deb','ID']], on='pr_deb').rename(columns=
                                                            {'ID':'id_comptag'}).drop('pr_deb', axis=1)
        tmjm_mens.rename(columns={c:k for c in tmjm_mens.columns if c!='id_comptag' for k, v in dico_mois.items()  if v[0]==int(c[-2:])}, inplace=True)
        tmjm_mens=tmjm_mens.applymap(lambda x : int(x) if isinstance(x, float) else  x)
        tmjm_mens['donnees_type']='tmja'
        pc_pl_mens=donnees_brutes[[a for a in donnees_brutes.columns if a == 'pr_deb' or re.search('Pc PL [0-9]{2}$',a)]].merge(base[['pr_deb','ID']], on='pr_deb').rename(columns=
                                                                    {'ID':'id_comptag'}).drop('pr_deb', axis=1)
        pc_pl_mens.rename(columns={c:k for c in pc_pl_mens.columns if c!='id_comptag' for k, v in dico_mois.items()  if v[0]==int(c[-2:])}, inplace=True)
        pc_pl_mens['donnees_type']='pc_pl'
        self.df_attr_mens=pd.concat([tmjm_mens,pc_pl_mens], axis=0, sort=False).sort_values('id_comptag')
        self.df_attr_mens['annee']=str(self.annee)
        self.df_attr_mens['fichier']=os.path.basename(self.fichier_perm)
        
    def classer_compteur_update_insert(self, table_cpt):
        """
        uniquement pour verif, car normalemnet df_attr=df_atr_update et df_attr_insertest vide
        """
        self.existant = compteur_existant_bdd(table_cpt)
        self.df_attr_update=self.df_attr.loc[self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        self.df_attr_insert=self.df_attr.loc[~self.df_attr.id_comptag.isin(self.existant.id_comptag.tolist())].copy()
        
        
    def update_bdd_Vinci(self, schema, table):
        val_txt=self.creer_valeur_txt_update(self.df_attr, ['id_comptag','tmja','pc_pl', 'src'])
        self.update_bdd(schema, table, val_txt,{f'tmja_{str(self.annee)}':'tmja',f'pc_pl_{str(self.annee)}':'pc_pl', f'src_{str(self.annee)}':'src'})
        
class Comptage_alienor(Comptage):    
    def __init__(self,fichier_perm, annee):
        self.fichier_perm=fichier_perm
        self.annee=annee
    
    def ouvrir_et_separe_donnees(self):
        donnees=pd.read_excel(self.fichier_perm).rename(columns={'ID':'id_comptag'})
        donnees_agregees=donnees[[a for a in donnees.columns if a in ('id_comptag','TMJA_'+str(self.annee),'Pc_PL_'+str(self.annee))]].copy()
        donnees_agregees.rename(columns={a:a.lower() for a in donnees_agregees.columns}, inplace=True)
        donnees_agregees['tmja_'+str(self.annee)]=donnees_agregees['tmja_'+str(self.annee)].apply(lambda x : int(x))
        donnees_agregees['src']='tableur'
        donnees_agregees['fichier']=os.path.basename(self.fichier_perm)
        tmjm_mens=donnees[[a for a in donnees.columns if a == 'id_comptag' or 'TMJM' in a]].copy()#or re.search('Pc_PL_[0-9]{4}_[0-9]{2}',a)
        tmjm_mens.rename(columns={c:k for c in tmjm_mens.columns if c!='id_comptag' for k, v in dico_mois.items()  if v[0]==int(c[-2:])}, inplace=True)
        tmjm_mens=tmjm_mens.applymap(lambda x : int(x) if isinstance(x, float) else  x)
        tmjm_mens['donnees_type']='tmja'
        pc_pl_mens=donnees[[a for a in donnees.columns if a == 'id_comptag' or re.search('Pc_PL_[0-9]{4}_[0-9]{2}',a)]].copy()
        pc_pl_mens.rename(columns={c:k for c in pc_pl_mens.columns if c!='id_comptag' for k, v in dico_mois.items()  if v[0]==int(c[-2:])}, inplace=True)
        pc_pl_mens['donnees_type']='pc_pl'
        donnees_mens=pd.concat([tmjm_mens,pc_pl_mens], axis=0, sort=False).sort_values('id_comptag')
        donnees_mens['annee']=str(self.annee)
        donnees_mens['fichier']=os.path.basename(self.fichier_perm)
        self.df_attr=donnees_agregees.copy()
        self.df_attr_mens=donnees_mens.copy()
    
    def update_bdd_Alienor(self, schema, table):
        val_txt=self.creer_valeur_txt_update(self.df_attr, ['id_comptag','tmja_'+str(self.annee),'pc_pl_'+str(self.annee), 'src'])
        self.update_bdd(schema, table, val_txt,{'tmja_'+str(self.annee):'tmja_'+str(self.annee),'pc_pl_'+str(self.annee):'pc_pl_'+str(self.annee), 
                                                     'src_'+str(self.annee):'src'})    

class Comptage_atlandes(Comptage):    
    def __init__(self,fichier_perm, annee):
        self.fichier_perm=fichier_perm
        self.annee=annee  
        
    def miseEnForme(self):
        """
        traiter le fichier pour renommer et preparer les colonnes
        """ 
        donnees=pd.read_excel(self.fichier_perm)
        donnees=donnees.loc[(~donnees['TMJA 2 sens confondus'].isna()) & (~donnees['TMJM VL'].isna())].copy()
        val_l0=donnees.iloc[0].values
        col_tmjm_vl=donnees.columns[donnees.columns.tolist().index('TMJM VL'):donnees.columns.tolist().index('TMJM PL')]
        dico_nom_tmjm_vl={a:'tmjm_vl_'+b for a, b in zip(col_tmjm_vl,val_l0[donnees.columns.tolist().index('TMJM VL'):donnees.columns.tolist().index('TMJM PL')])}
        col_tmjm_pl=donnees.columns[donnees.columns.tolist().index('TMJM PL'):]
        dico_nom_tmjm_pl={a:'tmjm_pl_'+b for a, b in zip(col_tmjm_pl,val_l0[donnees.columns.tolist().index('TMJM VL'):donnees.columns.tolist().index('TMJM PL')])}
        donnees=donnees.rename(columns=dico_nom_tmjm_vl).rename(columns=dico_nom_tmjm_pl).rename(columns={'TMJA 2 sens confondus':'tmja_vl','Unnamed: 8':'tmja_pl' })
        donnees=donnees.iloc[1:].copy()
        donnees['id_comptag']=donnees.apply(lambda x : f"40-A63-{str(x['Echangeur de début de section Trafic'])}+{str(x['Unnamed: 2'])}"if x['Echangeur de début de section Trafic'] != 36 else
                                          f"33-A63-{str(x['Echangeur de début de section Trafic'])}+{str(x['Unnamed: 2'])}", axis=1)
        donnees['fichier']=os.path.basename(self.fichier_perm)
        donnees['src']='tableur gestionnaire'
        donnees['annee']=str(self.annee)
        return donnees
    
    def donneesAgregees(self):
        """
        creer la df format bdd
        """
        donneesAgrege=self.miseEnForme()[['id_comptag','tmja_vl','tmja_pl', 'fichier', 'annee', 'src' ]].copy()
        donneesAgrege['type_veh']='vl/pl'
        donneesAgrege['tmja']=(donneesAgrege.tmja_vl+donneesAgrege.tmja_pl).astype(int)
        donneesAgrege['pc_pl']=donneesAgrege.tmja_pl/donneesAgrege.tmja*100
        return donneesAgrege
        
        
    def donnees_mens(self):
        """
        donnees=pd.read_excel(self.fichier_perm)
        donnees=donnees.loc[(~donnees['TMJA 2 sens confondus'].isna()) & (~donnees['TMJM VL'].isna())].copy()
        val_l0=donnees.iloc[0].values
        col_tmjm_vl=donnees.columns[donnees.columns.tolist().index('TMJM VL'):donnees.columns.tolist().index('TMJM PL')]
        dico_nom_tmjm_vl={a:'tmjm_vl_'+b for a, b in zip(col_tmjm_vl,val_l0[donnees.columns.tolist().index('TMJM VL'):donnees.columns.tolist().index('TMJM PL')])}
        col_tmjm_pl=donnees.columns[donnees.columns.tolist().index('TMJM PL'):]
        dico_nom_tmjm_pl={a:'tmjm_pl_'+b for a, b in zip(col_tmjm_pl,val_l0[donnees.columns.tolist().index('TMJM VL'):donnees.columns.tolist().index('TMJM PL')])}
        donnees=donnees.rename(columns=dico_nom_tmjm_vl).rename(columns=dico_nom_tmjm_pl)
        donnees=donnees.iloc[1:].copy()
        donnees['id_comptag']=donnees.apply(lambda x : f"40-A63-{str(x['Echangeur de début de section Trafic'])}+{str(x['Unnamed: 2'])}"if x['Echangeur de début de section Trafic'] != 36 else
                                          f"33-A63-{str(x['Echangeur de début de section Trafic'])}+{str(x['Unnamed: 2'])}", axis=1)
        """
        donnees=self.miseEnForme()
        colFixe=['id_comptag', 'fichier', 'annee']
        donnees_mens=donnees[[a for a in donnees.columns if a in colFixe or 'tmjm' in a ]].copy()
        #calcul du tmjm_tv
        for m in dico_mois.values():
            donnees_mens[f'tmjm_tv_{m[1]}']=donnees_mens[f'tmjm_vl_{m[1]}']+donnees_mens[f'tmjm_pl_{m[1]}']
            donnees_mens[f'tmjm_pc_pl_{m[1]}']=donnees_mens.apply(lambda x : round(x[f'tmjm_pl_{m[1]}']/x[f'tmjm_tv_{m[1]}']*100,2), axis=1)
        tmjm_tv_mens=donnees_mens[[a for a in donnees_mens.columns if 'tv' in a or a in colFixe]].copy()
        tmjm_tv_mens.rename(columns={c:k for c in tmjm_tv_mens.columns if c!='id_comptag' for k, v in dico_mois.items()  if v[1]==c.split('_')[-1]}, inplace=True)
        tmjm_tv_mens['donnees_type']='tmja'
        tmjm_pl_mens=donnees_mens[[a for a in donnees_mens.columns if 'pc_pl_' in a or a in colFixe]].copy()
        tmjm_pl_mens.rename(columns={c:k for c in tmjm_pl_mens.columns if c!='id_comptag' for k, v in dico_mois.items()  if v[1]==c.split('_')[-1]}, inplace=True)
        tmjm_pl_mens['donnees_type']='pc_pl'
        donnees_mens_final=pd.concat([tmjm_tv_mens,tmjm_pl_mens], axis=0, sort=False).sort_values('id_comptag')
        return donnees_mens_final 
    
class Comptage_Cofiroute(Comptage):
    """
    classs rapide surtout pour traiter le mensuel
    """
    def __init__(self, annee, fichier):
        """
        attributs:
            annee : string 4 caracteres
            fichier : raw string full path   
            dfFichier : dataframe issue de fichier et corrigée par ouvrirMiseEnForme()
            colTmjaMens : liste des colonne de TMJM
            colPcplMens : liste des colonees de Pcpl mensuel
        """
        self.annee = annee
        self.fichier = fichier
        self.dfFichier, self.colTmjaMens, self.colPcplMens = self.ouvrirMiseEnForme()
        
    def ouvrirMiseEnForme(self):
        dfFichier = pd.read_excel(self.fichier)
        # params mensuel et correction pc_pl
        colTmjaMens = [f'TMJM {str(i).zfill(2)}' for i in range (1,13)]
        colPcplMens = [f'Pc PL {str(i).zfill(2)}' for i in range (1,13)]
        for c in colPcplMens:
            dfFichier[c] = dfFichier[c]*100
        dfFichier[f'Pc PL {self.annee}'] = dfFichier[f'Pc PL {self.annee}'] * 100
        return dfFichier, colTmjaMens, colPcplMens
    
    def insererDonneesComptageBdd(self, inserer=False):
        dfComptage = creer_comptage(self.dfFichier.id_comptag, self.annee, 'tableau excel', 'tv/pl')
        if inserer: 
            insererSchemaComptage(dfComptage, 'comptage')
        return dfComptage
    
    def insererDonneesAgregeBdd(self, inserer=False):
        dfAgrege = structureBddOld2NewForm(self.dfFichier.rename(
            columns={f'TMJA {self.annee}': 'tmja', f'Pc PL {self.annee}': 'pc_pl','Vitesse moyenne annuelle (km/h)': 'vmoy'
                     }).assign(fichier=os.path.basename(self.fichier), annee=self.annee), 
            ['id_comptag', 'fichier', 'annee'], ['tmja', 'pc_pl', 'vmoy'], 'agrege')
        if inserer : 
            insererSchemaComptage(dfAgrege, 'indicAgrege')
        return dfAgrege
    
    def insererDonneesMensuelleBdd(self, inserer=False):
        tmjm_tv_mens = self.dfFichier[self.colTmjaMens+['id_comptag']].copy()
        tmjm_tv_mens = tmjm_tv_mens.rename(columns={c: k for c in tmjm_tv_mens.columns if c != 'id_comptag' 
                                                    for k, v in dico_mois.items() if str(v[0]).zfill(2) == c.split(' ')[-1]}).assign(donnees_type='tmja')
        tmjm_pl_mens = self.dfFichier[self.colPcplMens+['id_comptag']].copy()
        tmjm_pl_mens = tmjm_pl_mens.rename(columns={c: k for c in tmjm_pl_mens.columns if c != 'id_comptag' 
                                                    for k, v in dico_mois.items()  if str(v[0]).zfill(2) == c.split(' ')[-1]}).assign(donnees_type='pc_pl')
        donnees_mens_final = pd.concat([tmjm_tv_mens,tmjm_pl_mens]).sort_values('id_comptag').assign(annee=self.annee, fichier=os.path.basename(self.fichier))
        dfMensuel = structureBddOld2NewForm(
            donnees_mens_final, ['id_comptag', 'annee', 'donnees_type', 'fichier'],
            list(dico_mois.keys()), 'mensuel')
        if inserer : 
            insererSchemaComptage(dfMensuel, 'indicMensuel')
        return dfMensuel
        
        
class Comptage_GrandPoitiers(Comptage):
    def __init__(self,dossier, annee):
        self.dossier=dossier
        self.fichiers=O.ListerFichierDossier(dossier, 'xlsx')
        self.annee=annee
        
    def donnnees_agregees(self, fichier):
        donnees_agregees=pd.read_excel(os.path.join(self.dossier,fichier), sheet_name='Informations')  
        tmja=int(donnees_agregees.iloc[10,5])
        pc_pl=round(donnees_agregees.iloc[10,8]*100,2)
        pl=int(donnees_agregees.iloc[10,7])
        rue=donnees_agregees.iloc[1,1].split('(')[0].strip()
        return donnees_agregees, rue, tmja, pc_pl, pl
    
    def donnees_horaires(self, fichier,start=8, step=34):
        donnees_detaillees=pd.read_excel(os.path.join(self.dossier,fichier), sheet_name='Tableau', header=None)  
        donnees_detaillees=donnees_detaillees.loc[:,[1,26,27]].copy()
        donnees_detaillees.columns=['heure','tv','pl']
        for i,j in enumerate(range(start,len(donnees_detaillees),step)):
            date=donnees_detaillees.iloc[j,0]
            donnees=donnees_detaillees.iloc[j+2:j+26,1:3]
            date_index=pd.date_range(date, periods=24, freq='H')
            donnees.index=date_index
            donnees=donnees.reset_index().rename(columns={'index':'date'})
            donnees['jour']=donnees['date'].apply(lambda x : pd.to_datetime(x.strftime('%Y-%m-%d')))
            donnees['heure']=donnees['date'].apply(lambda x : f'h{x.hour}_{x.hour+1 if x.hour!=24 else 24}')
            donnees=pd.concat([donnees[['jour', 'heure', 'tv']].pivot(index='jour',columns='heure', values='tv'),
            donnees[['jour', 'heure', 'pl']].pivot(index='jour',columns='heure', values='pl')], axis=0, sort=False).reset_index()
            donnees['type_veh']=['TV','PL']
            if i==0 : 
                donnees_horaire=donnees.copy()
            else : 
                donnees_horaire=pd.concat([donnees_horaire,donnees], axis=0, sort=False)
            donnees_horaire.index.name='index' #juste pour lisibilite
            donnees_horaire.columns.name=None
        return donnees_horaire
    
    def cumul_sens(self, ref_pt, id_comptag):
        """
        faire la somme des deux sens quand on a une liste des deux fichiers de comptage
        in : 
            ref_pt : string des points relatif à un id_comptag, issu de la Bdd: comptag.gp_tmja_2015_p.id_comp
            id_comptag : id_comptag auquel associe les fichiers
        """
        liste_tmja=[]
        liste_pl=[]
        liste_donnees_horaire=[]
        if len(ref_pt)==1 :
            fichiers_sens=[f for f in self.fichiers if f'Poste {ref_pt[0]} - sens' in f] if ref_pt[0]!='20' else [f for f in self.fichiers if re.search('Poste 20[A,B] - sens',f)]
        elif len(ref_pt)==2 : 
            fichiers_sens=[f for f in self.fichiers if (f'Poste {ref_pt[0]} - sens' in f or f'Poste {ref_pt[1]} - sens' in f)]
        else  : 
            print(f'pb nb de fichier : {ref_pt}')
            raise Exception
        
        if not fichiers_sens : 
            raise self.PasDeFichierError(ref_pt)
        
        for f in fichiers_sens : 
            rue, tmja, pc_pl,pl=self.donnnees_agregees(f)[1:]
            liste_tmja.append(tmja)
            liste_pl.append(pl)
            liste_donnees_horaire.append(self.donnees_horaires(f))
        tmja=sum(liste_tmja) 
        pc_pl=round(100*sum(liste_pl)/tmja, 2)
        donnees_horaires_f=liste_donnees_horaire[0].set_index('jour').add(liste_donnees_horaire[1].set_index('jour')).reset_index() if len(fichiers_sens)>1 else liste_donnees_horaire[0]
        donnees_horaires_f.type_veh=donnees_horaires_f.type_veh.apply(lambda x : x[:2])
        donnees_horaires_f['id_comptag']=id_comptag
        return tmja,pc_pl, donnees_horaires_f,fichiers_sens
    
    def comptage_forme(self):
        with ct.ConnexionBdd(nomConnBddOtv) as c : 
            rqt="select * from comptage.gp_tmja_2015_p"
            points_existants=gp.read_postgis(rqt,c.sqlAlchemyConn)
        i=0
        dico_cpt={}
        for p in points_existants.point_comp.tolist():
            if isinstance(p, str) : #ne pas prendre en compte les points sans références
                ref_pt=p.split('_')
                id_comptag=points_existants.loc[points_existants['point_comp']==p].id_comptag.to_numpy()[0]
                try :
                    tmja,pc_pl, donnees_horaires_f,fichiers_sens=self.cumul_sens(ref_pt,id_comptag)
                except self.PasDeFichierError : 
                    continue
                dico_cpt[id_comptag]={'tmja':tmja, 'pc_pl':pc_pl, 'src':f"tableurs : {', '.join(fichiers_sens)}"}
                if i==0 :
                    donnees_horaires_tot=donnees_horaires_f.copy()
                else : donnees_horaires_tot=pd.concat([donnees_horaires_tot, donnees_horaires_f], axis=0, sort=False)
                i+=1
        self.df_attr=pd.DataFrame.from_dict(dico_cpt, orient='index').reset_index().rename(columns={'index':'id_comptag'})
        self.df_attr_horaire=donnees_horaires_tot.copy()
        
    def update_bdd_grdPoi(self, schema, table):
        val_txt=self.creer_valeur_txt_update(self.df_attr, ['id_comptag','tmja','pc_pl', 'src'])
        self.update_bdd(schema, table, val_txt,{f'tmja_{str(self.annee)}':'tmja',f'pc_pl_{str(self.annee)}':'pc_pl', f'src_{str(self.annee)}':'src'})
    
    class PasDeFichierError(Exception): 
        def __init__(self, ref_pt):
            Exception.__init__(self,f'pas de fihcier pour les postes de comptage {ref_pt} ')  
        

class Comptage_Niort(Comptage): 
    def __init__(self,dossier, annee) : 
        """
        fichiers csv fourni par sens, les fichiers csv fournis ressemble aux fichiers mdb de la ville d'Anglet et Angouleme
        on part comme criter de VL ou PL un longueur à 7,2 m
        # voiture : <5,2m
        # camionnete : >=5.2 à <7.2
        # bus et PL : >=7.2 à <10.9
        # PL remorque : >=10.9 à <22
        """
        self.dossier=dossier
        self.annee=annee
        
    def creer_dico(self, dico_id_comptag, dicoFormat):
        """
        creeru dico avec en cle les id_comptag et en value les fchiers concernes
        in:
            dico_id_comptag : dico de correspondance entre les dossier fournis par la ville de Niort et les id_comptag
            dicoFormat : dico de description des types et format de données dans chaque dossier fournis par la ville de Niort.
                         format du dico : {dossier: {'type': 'agrege' ou 'horaire', 'format': 'csv' ou 'xls' ou 'mdb'},}
        """
        dico_fichier = {}
        for v in dicoFormat.values():
            if v['format'] not in niort_formatFichierAccepte:
                raise FormatError(f"le format {v['format']} n'est pas pris en charge dans le code")
        for chemin, dossier, files in os.walk(self.dossier):
            for d in dossier:
                if d not in dicoFormat:
                    warnings.warn(f"le dossier {d} n'est pas présent dans le dico des format. A ajouter avant nouvelle itération")
                else:
                    dico_fichier[d] = [chemin+os.sep+d+os.sep+f for f in O.ListerFichierDossier(
                        os.path.join(chemin, d), dicoFormat[d]['format'])]
                
            #pour les dossier cotenat plus de 2 fichiers csv (i.e plus de 1 pt de comptage), erreur levee et correction mano
        dico_fichier_separe={}
        for k, v in dico_fichier.items():
            if len(v)>2:
                raise NotImplementedError(f"""le dossier {k} contient plus de 2 comptages. 
                    a separer dans deux dossiers differents. 
                    mettre les id_comptag en cohérence avec les nouveaux nom de dossier""")
            else:
                dico_fichier_separe[k] = v  
        dico_fichiers_final = {}
        for k, v in dico_fichier_separe.items():
            for c, val in dico_id_comptag.items():
                if k == val['dossier']:
                    dico_fichiers_final[c] = v
        return dico_fichiers_final
    
    def formater_fichier_csv(self, fichier):
        """
        à partir des fichiers csv, obtenir une df avec modification de la date si le fichier comporte uniquement 8 jours, i.e : 6 jours plein et 2 demi journée
        """    
        df_fichier = pd.read_csv(fichier)
        df_fichier.DateTime = pd.to_datetime(df_fichier.DateTime)
        nb_jours_mesure = len(df_fichier.set_index('DateTime').resample('D'))
        jourMin = df_fichier.DateTime.apply(lambda x: x.dayofyear).min()
        jourMax = df_fichier.DateTime.apply(lambda x: x.dayofyear).max()
        if nb_jours_mesure == 8: #si le nombre total de jours est inférieur à 9, il faudra regrouper les jours de début et de fin pour avoir une journée complete, ou alors c'est qu'une erreur a ete commise lors de la mesure de trafic (duree < 7 jours)
            df_fichier.loc[df_fichier.DateTime.apply(
                lambda x: x.dayofyear == jourMin), 'DateTime'] = df_fichier.loc[df_fichier.DateTime.apply(
                    lambda x: x.dayofyear == jourMin)].DateTime.apply(lambda x: x + pd.Timedelta('7D'))
        elif nb_jours_mesure > 8: 
            df_fichier = df_fichier.loc[df_fichier.DateTime.apply(lambda x: (x.dayofyear != jourMin) & (x.dayofyear != jourMax))].copy()
        elif nb_jours_mesure == 7: 
            df_fichier.loc[df_fichier.DateTime.apply(lambda x : x.dayofyear == jourMin), 'DateTime'
                           ] = df_fichier.loc[df_fichier.DateTime.apply(lambda x: x.dayofyear==jourMin)].DateTime.apply(
                               lambda x: x + pd.Timedelta('6D'))
        else : 
            warnings.warn(f"attention, moins de 7 jours dans le fichier {fichier}")
        return df_fichier
             
    def csv_identifier_type_veh(self,df_brute):
        """
        separer les vl des pl à partir de la df de données brutes creer avec formater_fichier_csv
        """        
        vl = df_brute.loc[(df_brute['Length']<7.2) & (df_brute['AdviceCode']!=128)].copy()
        pl = df_brute.loc[(df_brute['Length']>=7.2) & (df_brute['Length']<22) & (df_brute['AdviceCode']!=128) ].copy()
        return vl,pl
    
    def csv_agrege_donnees_brutes(self,vl,pl):
        """
        prendre les donnees issues de identifier_type_veh et les regrouper par heure et jour
        """
        vl_agrege_h = vl.set_index('DateTime').resample('H').AdviceCode.count().reset_index().rename(columns={'AdviceCode':'nb_veh'})
        vl_agrege_j = vl.set_index('DateTime').resample('D').AdviceCode.count().reset_index().rename(columns={'AdviceCode':'nb_veh'})
        pl_agrege_h = pl.set_index('DateTime').resample('H').AdviceCode.count().reset_index().rename(columns={'AdviceCode':'nb_veh'})
        pl_agrege_j = pl.set_index('DateTime').resample('D').AdviceCode.count().reset_index().rename(columns={'AdviceCode':'nb_veh'})
        return vl_agrege_h, vl_agrege_j, pl_agrege_h, pl_agrege_j
    
    def csv_indicateurs_agrege(self, vl_agrege_j, pl_agrege_j ):
        """
        calcul des indicateurs tmja et nb_pl
        """
        nb_vl = round(vl_agrege_j.mean(numeric_only=True).values[0])
        nb_pl = round(pl_agrege_j.mean(numeric_only=True).values[0])
        tmja = nb_vl + nb_pl
        pc_pl = round(nb_pl * 100  /tmja, 2)
        tmjo = round(vl_agrege_j.loc[vl_agrege_j.DateTime.dt.dayofweek.isin(range(5))].nb_veh.mean()
                      + pl_agrege_j.loc[pl_agrege_j.DateTime.dt.dayofweek.isin(range(5))].nb_veh.mean())
        nb_pl_o = pl_agrege_j.loc[pl_agrege_j.DateTime.dt.dayofweek.isin(range(5))].nb_veh.mean()
        pc_pl_o = nb_pl_o / tmjo * 100
        
        return nb_vl, nb_pl, tmja, pc_pl, tmjo, pc_pl_o, nb_pl_o
    
    def csv_donnees_horaires(self, vl_agrege_h, pl_agrege_h):
        """
        renvoyer la df des donnees horaires en tv et pl
        """
        vl_agrege_h['jour'] = vl_agrege_h['DateTime'].apply(lambda x: pd.to_datetime(x.strftime('%Y-%m-%d')))
        vl_agrege_h['heure'] = vl_agrege_h['DateTime'].apply(lambda x: f"h{x.hour}_{x.hour+1 if x.hour != 24 else 24}")
        pl_agrege_h['jour'] = pl_agrege_h['DateTime'].apply(lambda x: pd.to_datetime(x.strftime('%Y-%m-%d')))
        pl_agrege_h['heure'] = pl_agrege_h['DateTime'].apply(lambda x: f'h{x.hour}_{x.hour+1 if x.hour != 24 else 24}')
        
        pl_agrege_h = pl_agrege_h[['jour', 'heure', 'nb_veh']].pivot(index='jour', columns='heure', values='nb_veh')
        calcul_agreg_h = pd.concat([vl_agrege_h[['jour', 'heure', 'nb_veh']].pivot(index='jour',columns='heure', values='nb_veh'),
                  pl_agrege_h], axis=0, sort=False)
        tv_agrege_h = calcul_agreg_h.groupby('jour').sum()
        tv_agrege_h['type_veh'] = 'TV'
        pl_agrege_h['type_veh'] = 'PL'
        donnees_horaires = pd.concat([tv_agrege_h,pl_agrege_h], axis=0, sort=False).fillna(0).reset_index()
        return donnees_horaires
    
    def csv_indic_agreg_2sens(self, id_comptag, fichiers):
        """
        calculer les indicateurs agreges pour un fichier de type csv horaire
        """
        list_resultat=[]
        for f in fichiers : 
            try :
                list_resultat.append(self.traiter_csv_sens_unique(f))
            except PasAssezMesureError :
                return pd.DataFrame([])
        tmja, nb_pl, tmjo, nb_pl_o, dateMin, dateMax = ([list_resultat[i][2] for i in range(len(fichiers))],
                                                        [list_resultat[i][1] for i in range(len(fichiers))],
                                                        [list_resultat[i][4] for i in range(len(fichiers))],
                                                        [list_resultat[i][6] for i in range(len(fichiers))],
                                                        [list_resultat[i][8] for i in range(len(fichiers))],
                                                        [list_resultat[i][9] for i in range(len(fichiers))])
        tmja = sum(tmja)
        tmjo = sum(tmjo)
        pc_pl = round(sum(nb_pl) / tmja * 100, 2)
        pc_pl_o = round(sum(nb_pl_o) / tmjo * 100, 2)
        periode = f"{min(dateMin).strftime('%Y/%m/%d')}-{max(dateMax).strftime('%Y/%m/%d')}"
        return gp.GeoDataFrame({'id_comptag':[id_comptag,],'tmja':[tmja,],'pc_pl':pc_pl,
                                       'tmjo': tmjo, 'pc_pl_o': pc_pl_o, 'geometry': geomFromIdComptagCommunal(id_comptag),
                                       'fichier': ', '.join([os.path.basename(f) for f in fichiers]),
                                       'periode': periode})
    
    def traiter_csv_sens_unique(self, fichier):
        """
        concatenation des ofnctions permettant d'aboutir aux donnees agrege et horaire d'un fichier csv de comptage pour un sens
        """
        f = self.formater_fichier_csv(fichier)
        vl, pl = self.csv_identifier_type_veh(f)
        vl_agrege_h, vl_agrege_j, pl_agrege_h, pl_agrege_j = self.csv_agrege_donnees_brutes(vl, pl)
        nb_vl, nb_pl, tmja, pc_pl, tmjo, pc_pl_o, nb_pl_o = self.csv_indicateurs_agrege(vl_agrege_j, pl_agrege_j)
        donnees_horaires = self.csv_donnees_horaires(vl_agrege_h, pl_agrege_h)
        donnees_horaires['fichier'] = os.path.basename(fichier)
        dateMin = min(pl_agrege_j.DateTime.min(), vl_agrege_j.DateTime.min())
        dateMax = max(pl_agrege_j.DateTime.max(), vl_agrege_j.DateTime.max())
        return nb_vl, nb_pl, tmja, pc_pl, tmjo, pc_pl_o, nb_pl_o, donnees_horaires, dateMin, dateMax
    
    
    def xlsCpevAgrege(self, fichier, id_comptag):
        """
        à partir du nom de fichier, récupérer les infos des tables compteurs, comptage et indic_agrege de la BDD
        in :
            fichier : raw string du nom complet du fichier
            id_comptag : id_comptage associé
        out : 
            gdfAgrege : Geodataframe d'une seule ligne
        """
        dfBrute = pd.read_excel(fichier)
        geom = geomFromIdComptagCommunal(id_comptag)
        periode = '-'.join([pd.to_datetime(e, dayfirst=True).strftime('%Y/%m/%d') for e in re.findall('[0-9]{1,2}/[0-1][0-9]/20[0-2][0-9]', dfBrute.iloc[1, 16])])
        vma = int(re.search('([0-9]{2})( km/h)', dfBrute.iloc[6, 0]).group(1))
        tmjo = round(dfBrute.iloc[7, 12])
        pc_pl_o = round(dfBrute.iloc[9, 14]*100, 2)
        tmja = round(dfBrute.iloc[10, 12])
        pc_pl = round(dfBrute.iloc[12, 14]*100, 2)
        vmoy = round(dfBrute.iloc[10, 16], 2)
        v85 = round(dfBrute.iloc[10, 19], 2)
        gdfAgrege = gp.GeoDataFrame({'periode': periode, 'vma': vma, 'vmoy': vmoy, 'v85': v85, 'tmjo': tmjo, 'pc_pl_o': pc_pl_o, 'tmja': tmja, 'pc_pl': pc_pl,
                                     'geometry': geom, 'id_comptag': id_comptag, 'fichier': os.path.basename(fichier)},
                                     crs='EPSG:2154', index=[0])
        return gdfAgrege

    
    def horaire_2_sens(self, id_comptag, list_df_sens):
        """
        creer une df des donnes horaires pour les 2 sens d'un id_comptage
        """

        df_finale = pd.concat(list_df_sens, axis=0, sort=False) 
        comparer2Sens(df_finale.assign(id_comptag=id_comptag), attributSens='sens', attributIndicateur='type_veh',
                      facteurComp=10000, TauxErreur=0.00001)
        df_finale = df_finale.groupby(['jour', 'type_veh']).sum().reset_index()
        df_finale['id_comptag'] = id_comptag
        return df_finale
    
    
    def xlsCpevHoraireDebit(self, fichier):
        """
        à partir d'un fichier CPEV débit/vitesse tout sens confondus, extraire les débits horaires TV et PL
        """
        dfBrute = pd.read_excel(fichier)
        dfBruteHoraire = pd.concat(
            [dfBrute.iloc[niort_ligneDebutDebitHoraireTv: niort_ligneDebutDebitHoraireTv + niort_nbJoursHoraireCpev,
                          niort_colonneDebutDebitHoraire: niort_colonneFinDebitHoraire].assign(type_veh='TV'),
                          dfBrute.iloc[niort_ligneDebutDebitHorairePl: niort_ligneDebutDebitHorairePl + niort_nbJoursHoraireCpev,
                                       niort_colonneDebutDebitHoraire: niort_colonneFinDebitHoraire].assign(type_veh='PL')])
        dfBruteHoraire.columns = ['jour'] + attributsHoraire + ['type_veh']
        dfBruteHoraire['jour'] = dfBruteHoraire.jour.apply(lambda x: pd.to_datetime(f"{x.split('.')[1]}/{self.annee}", dayfirst=True))
        return dfBruteHoraire
    
    
    def xlsCpevHoraireVitesse(self, fichier):
        """
        à partir d'un fichier CPEV débit/vitesse tout sens confondus, extraire les vitesses horaires TV et PL
        """
        listVmoyHoraire = []
        dfBrute = pd.read_excel(fichier)
        for e in range(niort_vmoyHoraireVlStartCpev, niort_vmoyHoraireVlStartCpev + (niort_vmoyHoraireVlPasCpev * niort_nbJoursHoraireCpev
                                                                                     ), niort_vmoyHoraireVlPasCpev):
            dfJourVmoyHoraireVl = dfBrute.iloc[e: e + 24, [niort_colonneVmoyHoraire]].T.copy()
            dfJourVmoyHorairePl = dfBrute.iloc[e + niort_vmoyHorairePlPasCpev: e + niort_vmoyHorairePlPasCpev + 24,
                                               [niort_colonneVmoyHoraire]].T.copy()
            dfJourVmoyHoraireVl.columns = attributsHoraire
            dfJourVmoyHorairePl.columns = attributsHoraire
            dfJourVmoyHoraireVl['type_veh'] = 'vmoy_vl'
            dfJourVmoyHorairePl['type_veh'] = 'vmoy_pl'
            listVmoyHoraire.append(pd.concat([dfJourVmoyHoraireVl, dfJourVmoyHorairePl]).assign(
                jour=pd.to_datetime(dfBrute.iloc[e+1, niort_colonneVmoyHoraireJour], dayfirst = True)))
        return pd.concat(listVmoyHoraire)
    
    
    def xlsCpevHoraire(self, fichier, id_comptag):
        return pd.concat([self.xlsCpevHoraireDebit(fichier), self.xlsCpevHoraireVitesse(fichier)]
                         ).assign(fichier = os.path.basename(fichier), id_comptag=id_comptag).reset_index(drop=True)


    def calculAgregeHoraireTouteSource(self, dico_fichiers_final, dico_id_comptag, dicoFormat):
        """
        pour l'ensembledes points de comptage, fournir une df agrege et Horaire.
        in : 
            dico_fichiers_final : dico créée par creer_dico()
            dico_id_comptag : dico de correspondance entre les dossier fournis par la ville de Niort et les id_comptag
            dicoFormat : dico de description des types et format de données dans chaque dossier fournis par la ville de Niort.
                         format du dico : {dossier: {'type': 'agrege' ou 'horaire', 'format': 'csv' ou 'xls' ou 'mdb'},}
        out:
            dfAgregeFinale : df avec un entrée par compteur et les attributs id_comptag', 'tmja', 'pc_pl', 'tmjo', 'pc_pl_o', 'geometry', 'fichier',
                             'periode', 'src', 'annee', 'vma', 'vmoy', 'v85'
            dfHoraireFinale : df avec les attributs de la Bdd Horaire, (hormis id_uniq_cpt remplacé par id_comptag et annee)
        """
        listGdfAgrege, listDfHoraire, list_2sens = [], [], []
        for k, v in dico_fichiers_final.items():
            # if k != 'Niort-32_Rue_de_l_Aerodrome--0.4260;46.3247':
                # continue
            # récupération de la typologie des données
            doss = dico_id_comptag[k]['dossier']
            if dicoFormat[doss]['type'] == 'horaire':
                src = dicoFormat[doss]['type'] + '_' + dicoFormat[doss]['format']
                if dicoFormat[doss]['format'] == 'xls':
                    listGdfAgrege.append(self.xlsCpevAgrege(v[0], k).assign(src=src, annee=self.annee))
                    listDfHoraire.append(self.xlsCpevHoraire(v[0], k).assign( annee=self.annee))
                elif dicoFormat[doss]['format'] == 'csv':
                    for i, f in enumerate(v, 1):
                        list_2sens.append(self.traiter_csv_sens_unique(f)[7].assign(sens=f'sens {i}', 
                                                                                    sens_cpt=dico_id_comptag[k]['sens_cpt'],))
                    if not list_2sens:
                        warnings.warn(f'des donnees sont vides pour le comptage {k} dans le dossier {doss}')
                        list_2sens = []
                        continue
                    listDfHoraire.append(self.horaire_2_sens(k, list_2sens).assign(fichier=os.path.basename(v[0]),
                                                                                   annee=self.annee))
                    listGdfAgrege.append(gp.GeoDataFrame(self.csv_indic_agreg_2sens(k, v).assign(src=src, annee=self.annee),
                                                         geometry=[geomFromIdComptagCommunal(k)], crs='EPSG:2154'))
                    list_2sens = []
                elif dicoFormat[doss]['format'] == 'mdb':
                    if doss[:5].lower() == 'gare_':
                        intSensAUtiliser = int(doss.split('_')[1])
                        numSensAUtiliser = [intSensAUtiliser] if intSensAUtiliser <= 10 else [int(str(intSensAUtiliser)[0]), int(str(intSensAUtiliser)[1])]
                        cpt = MHCorbin(v[0], numSensAUtiliser=numSensAUtiliser)
                    else:
                        cpt = MHCorbin(v[0])
                    dfHoraire = cpt.formaterDonneesHoraires(cpt.formaterDonneesIndiv(cpt.dfAgreg2Sens)).assign(
                        id_comptag=k).reset_index().assign(annee=self.annee, fichier=os.path.basename(v[0]))
                    dfAgrege = gp.GeoDataFrame(tmjaDepuisHoraire(dfHoraire).pivot(
                        index=['id_comptag', 'annee', 'fichier'], columns='indicateur', values='valeur').reset_index().assign(
                        periode=periodeDepuisHoraire(dfHoraire).periode[0], src='horaire_mdb'), 
                        geometry=[geomFromIdComptagCommunal(k)], crs='EPSG:2154')
                    dfHoraire.rename(columns={'indicateur': 'type_veh'}, inplace=True)
                    listDfHoraire.append(dfHoraire)
                    listGdfAgrege.append(dfAgrege)
                else:
                    raise FormatError(dicoFormat[doss]['format'])
        return pd.concat(listGdfAgrege).reset_index(drop=True), pd.concat(listDfHoraire).reset_index(drop=True)
         
         
    def miseEnFormeCompteur(self, dfAgregeFinale, listRd, listSensUnique):
        """
        formater la df issue de calculAgregeHoraireTouteSource() pour pouvoir creer des compteurs.
        in : 
            dfAgregeFinale : df issue de calculAgregeHoraireTouteSource()
            listRd : list des id_comptag situés sur une RD
            listSensUnique : list des id_comptag situé sur des sens uniques
        """    
        return creerCompteur(
            dfAgregeFinale.assign(type_poste='ponctuel',
                                  src_geo='carto_gestionnaire',
                                  pr=None,abs=None,route=dfAgregeFinale.id_comptag.apply(
                                      lambda x: re.sub('[0-9]{1,3}[a-z]{0,1}_', '', x.split('-')[1]).replace('_', ' ')),
                                  src_cpt=dfAgregeFinale.id_comptag.apply(
                                      lambda x: 'collectivite territoriale' if x in listRd else 'convention gestionnaire'),
                                  convention=True,
                                  sens_cpt=dfAgregeFinale.id_comptag.apply(
                                      lambda x: 'sens unique' if x in listSensUnique else 'double sens')),
            'geometry',
            79,
            dfAgregeFinale.id_comptag.apply(
                lambda x: 'RD' if x in listRd else 'VC'),
            dfAgregeFinale.id_comptag.apply(lambda x: 'CD79' if x in listRd else 'Niort'), False)
    
        
class Comptage_GrandDax(Comptage): 
    def __init__(self, fichiers_pt_comptages):
        """
        pour le moment je me base sur un fichier de geolocalisation creer à la main qui reprend aussi les principales info 
        pour creer l'di_comptag et les noms de fichiers a utiliser pour creer les donnees
        """
        self.fichiers_pt_comptages=gp.read_file(fichiers_pt_comptages, encoding='UTF-8')
        self.MaJ_fichiers_pt_comptages()
    
    def MaJ_fichiers_pt_comptages(self):
        """
        ajouter les attributs necessaires au fichier de pt de comptage
        """
        self.fichiers_pt_comptages['x_wgs84']=self.fichiers_pt_comptages.geometry.to_crs({'proj':'longlat', 'ellps':'WGS84', 'datum':'WGS84'}).apply(lambda x: round(x.x,4))
        self.fichiers_pt_comptages['y_wgs84']=self.fichiers_pt_comptages.geometry.to_crs({'proj':'longlat', 'ellps':'WGS84', 'datum':'WGS84'}).apply(lambda x: round(x.y,4))
        self.fichiers_pt_comptages['id_comptag']=self.fichiers_pt_comptages.apply(lambda x : f'GrdDax-{x.nom_rte}-{str(x.x_wgs84)};{str(x.y_wgs84)}', axis=1)
        self.fichiers_pt_comptages['dep']='40'
        self.fichiers_pt_comptages['reseau']='VC'
        self.fichiers_pt_comptages['gestionnai']='GrdDax'
        self.fichiers_pt_comptages['concession']='N'
        self.fichiers_pt_comptages['type_poste']='ponctuel'
        self.fichiers_pt_comptages['src_geo']='plan pdf'
        self.fichiers_pt_comptages['x_l93']=self.fichiers_pt_comptages.geometry.apply(lambda x : round(x.x,3))
        self.fichiers_pt_comptages['y_l93']=self.fichiers_pt_comptages.geometry.apply(lambda x : round(x.y,3))
     
    def ouvrir_feuille_infos(self, dossier, fichier):
        donnees_generale=pd.read_excel(os.path.join(dossier, fichier), sheet_name='Informations')
        donnees_gal_sansNa=donnees_generale.dropna(how='all').dropna(axis=1,how='all')
        return donnees_generale,donnees_gal_sansNa 
    
    def recup_infos(self,donnees_gal_sansNa, lgn_date_deb, lgn_date_fin):
        #dates et duree
        def dates_cpt(ligne, donnees_gal_sansNa):
            txt=donnees_gal_sansNa.loc[ligne,'Unnamed: 7'].split()
            for v in dico_mois.values():
                if txt[2] in [a.lower() for a in v if isinstance(a, str)] : 
                    mois=str(v[0])
            return pd.to_datetime('-'.join([txt[1],mois,txt[3]]), dayfirst=True)
    
        date_debut=dates_cpt(lgn_date_deb,donnees_gal_sansNa)
        date_fin=dates_cpt(lgn_date_fin,donnees_gal_sansNa)
        duree_cpt=int(donnees_gal_sansNa.loc[15,'Unnamed: 29'])
        annee=str(date_debut.year)
        return date_debut, date_fin, duree_cpt, annee
    
    def recup_donnees_agregees(self,donnees_gal_sansNa):
        #donnees_trafic
        tmja_vl=round(donnees_gal_sansNa.loc[19,'Unnamed: 12'],0)
        tmjo_vl=round(donnees_gal_sansNa.loc[20,'Unnamed: 12'],0)
        tmja_pl=round(donnees_gal_sansNa.loc[19,'Unnamed: 23'],0)
        tmjo_pl=round(donnees_gal_sansNa.loc[20,'Unnamed: 23'],0)
        pc_pl_ja=round((donnees_gal_sansNa.loc[19,'Unnamed: 32'])*100,1)
        pc_pl_jo=round((donnees_gal_sansNa.loc[20,'Unnamed: 32'])*100,1)
        tmja=tmja_vl+tmja_pl
        tmjo=tmjo_vl+tmjo_pl
        return tmja_vl,tmjo_vl, pc_pl_ja, tmjo_pl, tmjo, pc_pl_jo, tmja, tmja_pl
        
    def creer_df_agrege(self):
        """
        creer la df de tout les points de comptage avec indicateur agreges
        """
        #importer fichiers georeferencedes pt de comptag contenat id_comptag et reference vers le dossier et ou fichier      
        dico={}
        dico['id_comptag']=[]
        dico['obs_geo']=[]
        for a in range(2011,2021) : 
            dico[f'tmja_{str(a)}']=[]
            dico[f'pc_pl_{str(a)}']=[]
            dico[f'src_{str(a)}']=[]
            
        for dossier, fichiers, id_comptag in zip(self.fichiers_pt_comptages.dossier.tolist(),self.fichiers_pt_comptages.nom_fich.tolist(),
                                                 self.fichiers_pt_comptages.id_comptag.tolist()) : 
            dico['id_comptag'].append(id_comptag)
            dico['obs_geo'].append(dossier)
            if fichiers==None : #si pas de fichiers renseigne normelent seulement 2 IFX dans le dossier
                fichiers=O.ListerFichierDossier(dossier,extension='.IFX')
            else : 
                fichiers=fichiers.split(';')
    
            for f in fichiers : # si j'ai oublie l'extension 
                if not f.endswith('.IFX') : 
                    fichiers[fichiers.index(f)]+='.IFX'
    
            # recup des infos de dates
            donnees_gal_sansNa=self.ouvrir_feuille_infos(dossier, fichiers[0])[1]
            annee=self.recup_infos(donnees_gal_sansNa, 10, 11)[3]
    
    
            donnees_agregees=[self.recup_donnees_agregees(self.ouvrir_feuille_infos(dossier, f)[1])[6:] for f in fichiers]
            tmja=sum(i[0] for i in donnees_agregees)
            pc_pl=round(sum(i[1] for i in donnees_agregees)/tmja*100,1)
    
            for a in range(2011,2021) : 
                if int(annee)==a : 
                    dico[f'tmja_{a}'].append(tmja)
                    dico[f'pc_pl_{a}'].append(pc_pl)
                    dico[f'src_{a}'].append(';'.join([os.path.join(dossier, f) for f in fichiers]))
                else : 
                    dico[f'tmja_{a}'].append(np.NaN)
                    dico[f'pc_pl_{a}'].append(np.NaN)
                    dico[f'src_{a}'].append(np.NaN)
            
        self.df_attr=self.fichiers_pt_comptages[['id_comptag','geometry','dep','reseau','gestionnai','concession','type_poste','src_geo','x_l93','y_l93']].merge(pd.DataFrame.from_dict(dico), on='id_comptag')
        
    def creer_donnees_source(self, dossier, fichier,date_debut,duree_cpt):
        return [(pd.read_excel(os.path.join(dossier, fichier), sheet_name=f'Tableau {i+1}'),date_debut+pd.Timedelta(f'{i}D')) for i in range(duree_cpt)]
    
    def creer_df_horaire_brutes(self,donnees_horaires_brutes, date) : #pour 1 journee
        donnees=([donnees_horaires_brutes,'VL','Unnamed: 29'], [donnees_horaires_brutes,'PL','Unnamed: 30'])
        
        def df_horaire_brute(donnees_horaires_brutes,type_veh,attr) : #pourcahque type de veh
            test=donnees_horaires_brutes.loc[15:38,['TABLEAU HORAIRE DE SYNTHESE DES RESULTATS JOURNALIERS',attr]].copy()
            test['type_veh']=type_veh
            test.rename(columns={'TABLEAU HORAIRE DE SYNTHESE DES RESULTATS JOURNALIERS':'heure',attr:'nb_veh'}, inplace=True)
            return test
        
    
        test=pd.concat([df_horaire_brute(*d) for d in donnees], axis=0, sort=False).reset_index(drop=True)
        df_horaire=pd.concat([test.loc[test.type_veh==v[1]].pivot(index='type_veh',columns='heure', values='nb_veh') for v in donnees], axis=0, sort=False)
        df_horaire.loc['TV',:]=df_horaire.sum(axis=0)
        df_horaire.reset_index(inplace=True)
        df_horaire.drop(0, inplace=True)
        df_horaire['jour']=date
        return df_horaire
    
    def creer_df_horaire_tt_jour(self, donnees_tout_jour):
        return pd.concat([self.creer_df_horaire_brutes(*d) for d in donnees_tout_jour], axis=0, sort=False)
    
    #pour 2 fihciers (2sens)
    def creer_df_horaire_2_sens(self, dossier,liste_fichiers, id_comptag):
        if liste_fichiers==None : #si pas de fichiers renseigne normelent seulement 2 IFX dans le dossier
            liste_fichiers=O.ListerFichierDossier(dossier,extension='.IFX')
        else : 
            liste_fichiers=liste_fichiers.split(';')
        for f in liste_fichiers : # si j'ai oublie l'extension 
            if not f.endswith('.IFX') : 
                liste_fichiers[liste_fichiers.index(f)]+='.IFX'     
        donnees_gal_sansNa=self.ouvrir_feuille_infos(dossier, liste_fichiers[0])[1]
        date_debut, date_fin, duree_cpt, annee=self.recup_infos(donnees_gal_sansNa, 10, 11)
        
        donnees_tout_jour=(self.creer_donnees_source(dossier, f,date_debut,duree_cpt) for f in liste_fichiers)
        df_2_sens = pd.concat([self.creer_df_horaire_tt_jour(d) for d in donnees_tout_jour], axis=0, sort=False)
        df_2_sens=df_2_sens.groupby(['jour', 'type_veh']).sum().reset_index()
        df_2_sens['id_comptag']=id_comptag
        df_2_sens.rename(columns={k:f"h{k.replace('H','').replace('-','_')}" for k in df_2_sens.columns if k[-1]=='H'},inplace=True )
        return df_2_sens
    
    def df_horaire_final(self):
        self.df_attr_horaire=pd.concat([self.creer_df_horaire_2_sens(dossier,fichiers, id_comptag) 
                     for dossier, fichiers, id_comptag in zip(self.fichiers_pt_comptages.dossier.tolist(),self.fichiers_pt_comptages.nom_fich.tolist(),
                                                              self.fichiers_pt_comptages.id_comptag.tolist())],
                     axis=0, sort=False)

class Comptage_Limoges_Metropole(Comptage): 
    def __init__(self, fichiers_shape, fichier_zone) : 
        """
        je me base sur un fichier shape généré à partir de qgis, depuis le service arcgis ici : https://siglm.agglo-limoges.fr/servernf/rest/services/_TRANSPORTS/transports_consult/featureServer
        """
        self.fichiers_shape = fichiers_shape
        df_trafic = gp.read_file(fichiers_shape, encoding='UTF8').set_index('objectid')
        df_trafic.drop_duplicates(['date_deb', 'date_fin','tv_moyjour','position','direction','type','nom_voie','codcomm' ], inplace=True)
        self.df_trafic = df_trafic.loc[(df_trafic.geometry!=None) & (~df_trafic.tv_moyjour.isna()) & 
                                     (df_trafic['type_capt']!='Manuel')].copy()
        self.fichier_zone = fichier_zone
        self.df_trafic['geom_wgs84'] = self.df_trafic.geometry.to_crs(epsg='4326')
        self.df_trafic['geom_wgs84_x'] = self.df_trafic.geom_wgs84.apply(lambda x : str(round(x.x, 4)))
        self.df_trafic['geom_wgs84_y'] = self.df_trafic.geom_wgs84.apply(lambda x : str(round(x.y, 4)))
        
        
    #fonction de recherche des points regroupable
    def groupe_point(self):
        """
        regroupe les points selon un fichier de zones creer mano
        ATTENTION : le fichier zone peut etre fauix sur certains sens : il peut y avoir 2 geométries unique
        au sein d'une zone, mais qui désigne 2 points fait das le mm sens sur 2 années différentes. 
        on est donc dans ce cas sur des sens uniques.
        in : 
           fichier_zone : str : raw_strinf de localisation du fichier
        """
        zone_grp=gp.read_file(self.fichier_zone)
        #jointure
        self.df_trafic=gp.sjoin(self.df_trafic,zone_grp, predicate='within')
    
    def isoler_zone(self, num_zone):
        """
        pour tout les points d'une zone, les grouper par type et caracteriser les types selon le nombre de 
        géométries differentes presentes
        in : 
            num_zone : integer présent dans le fichier des zones
        """
        zone_cbl=self.df_trafic.loc[self.df_trafic['id_groupe']==num_zone][['type', 'geometry', 'date_deb','date_fin', 'tv_moyjour',
                'pl_pourcen', 'id_groupe', 'nom_voie', 'position','geom_wgs84_x', 'geom_wgs84_y']].sort_values(['type', 'date_deb']).copy()
        zone_cbl['annee']=zone_cbl.date_deb.apply(lambda x : x[:4])
        #filtre vacances DEPRECATED depuis nouvelle structure donnees
        """zone_cbl=zone_cbl.loc[(zone_cbl.apply(lambda x : pd.to_datetime(x.date_deb).month not in [7,8] if not pd.isnull(x.date_deb) else True, axis=1)) & 
                    (zone_cbl.apply(lambda x : pd.to_datetime(x.date_fin).month not in [7,8] if not pd.isnull(x.date_fin) else True, axis=1))].copy()"""
        zone_cbl_grp_type=zone_cbl.groupby('type')
        carac_geom_type=zone_cbl_grp_type.geometry.unique().apply(lambda x : len(x))
        return zone_cbl, zone_cbl_grp_type, carac_geom_type
    
    def calcul_attributs_commun(self,format_cpt,geom,nom_voie, wgs84_geom_x, wgs84_geom_y,num_zone):
        #geometry
        format_cpt=gp.GeoDataFrame(format_cpt, geometry=[geom])
        format_cpt['src_geo']='export Webservice, cf atribut fichier'
        format_cpt['fichier']=os.path.basename(self.fichiers_shape)
        #reference au fichier cree et à la geolocalisation
        format_cpt['obs_supl']=f'numero de zone dans fichier geoloc : {num_zone}'
        format_cpt['id_comptag']=f'LimMet-{nom_voie.lower()}-{wgs84_geom_x};{wgs84_geom_y}' if nom_voie else f'LimMet-??-{wgs84_geom_x};{wgs84_geom_y}'
    
    def simplifier1PtComptage(self,zone_cbl,geom, wgs84_geom_x, wgs84_geom_y, num_zone, typeCpt):
        """
        convertir les données issue de isoler_zone() vers le format des données de la Bdd OTR
        in : 
           zone_cbl : df issue de  isoler_zone()
           geom : geometry shapely du point
           wgs84_geom_x, wgs84_geom_y : coordonnes en wgs84 arrondi à 10-3,
           num_zone : integer : numero de la zone issu du fichier self.fichier_zone
           typeCpt : string : 'Simple',  'Double', 'DETECTEUR_PC' : issu du type de comptage dans le fichier de Limoge Metropole
        """
        def conversion_date(date_deb, date_fin):
            if date_deb and date_fin : 
                return f"{pd.to_datetime(date_deb).strftime('%Y/%m/%d')}-{pd.to_datetime(date_fin).strftime('%Y/%m/%d')}"
            elif date_deb : 
                return f"{pd.to_datetime(date_deb).strftime('%Y/%m/%d')}-??"
            elif date_fin : 
                return f"??-{pd.to_datetime(date_fin).strftime('%Y/%m/%d')}"
            else : 
                return "pas de date de mesure connues"
            
        dbl=zone_cbl.loc[zone_cbl['type']==typeCpt]
        #gestion duc cas où plusieurs comptages dans l'année
        dbl_max=dbl.loc[dbl['tv_moyjour']==dbl.groupby('annee').tv_moyjour.transform(max)].copy()
        dbl_max['obs']=dbl_max.apply(lambda x : conversion_date(x['date_deb'], x['date_fin']), axis=1)
        format_cpt=dbl_max[['type', 'annee','tv_moyjour','pl_pourcen','id_groupe', 'obs']].pivot(index='id_groupe', columns='annee',
                            values=['tv_moyjour','pl_pourcen', 'obs'])
        #pour convertir les noms de colonnes
        annees=format_cpt.columns.levels[1]
        colonnes_name=['tmja_'+a for a in annees]+['pc_pl_'+a for a in annees]+['obs_'+a for a in annees]
        format_cpt.columns=format_cpt.columns.droplevel()
        format_cpt.columns=colonnes_name
        for a in annees:
            format_cpt[f'src_{a}']='Webservice Limoges Metropole'
        #nom de voies
        nom_voie=zone_cbl.loc[zone_cbl['type']==typeCpt].apply(lambda x : x.nom_voie if not pd.isnull(x.nom_voie) else x.position, axis=1).unique()[0]
        self.calcul_attributs_commun(format_cpt,geom,nom_voie, wgs84_geom_x, wgs84_geom_y,num_zone )
        #regrouper les années antérieure à 2010 dans une colonne 'autre'
        dico_toutes_annee_autre={a:{'tmja':format_cpt[f'tmja_{a}'].values[0],'pc_pl':format_cpt[f'pc_pl_{a}'].values[0],
                    'obs':format_cpt[f'obs_{a}'].values[0],'src':format_cpt[f'src_{a}'].values[0]} for a in annees if a<'2010'}
        if dico_toutes_annee_autre : 
            dico_annee_autre_max=dico_toutes_annee_autre[max(dico_toutes_annee_autre.keys())]
            format_cpt['ann_autr']=max(dico_toutes_annee_autre.keys())
            for c in ('tmja', 'pc_pl', 'obs','src') : 
                format_cpt[f'{c}_autr']=dico_annee_autre_max[c]
            format_cpt.drop([f'{c}_{a}' for c in ('tmja', 'pc_pl', 'obs', 'src') for a in dico_toutes_annee_autre.keys()],axis=1, inplace=True)
        return format_cpt
    
    def df_intermediaire(self,zone_cbl, zone_cbl_grp_type, carac_geom_type, reprojection, num_zone):
        """
        creer une df concatenable à partir des données de Limoges metropole. va dépendre du type de comptage, des zones et des nombre 
        de géométries différents
        in : 
            reprojection : vecteur de transformation selon la methode de shapely
            num_zone : integer : numero de la zone issu du fichier self.fichier_zone
        """
        if 'Double' in carac_geom_type.index : 
            geom, wgs84_geom_x, wgs84_geom_y=self.creer_geom('Double',carac_geom_type,zone_cbl_grp_type,zone_cbl, reprojection )
            df=self.simplifier1PtComptage(zone_cbl, geom, wgs84_geom_x, wgs84_geom_y,num_zone,'Double')
            df['obs_supl']=df['obs_supl']+' ; comptage 2 sens'
        elif (zone_cbl['type']=='Simple').all() : 
            geom, wgs84_geom_x, wgs84_geom_y=self.creer_geom('Simple',carac_geom_type,zone_cbl_grp_type,zone_cbl, reprojection )
            if carac_geom_type['Simple']==1 : 
                df=self.simplifier1PtComptage(zone_cbl, geom, wgs84_geom_x, wgs84_geom_y,num_zone,'Simple')
                df['obs_supl']=df['obs_supl']+' ; comptage 1 sens'
            else : 
                if carac_geom_type['Simple']==2 :
                    zone_cbl=zone_cbl.loc[zone_cbl['type']=='Simple'][['type','geometry','date_deb','date_fin','id_groupe','nom_voie','position','geom_wgs84_x','geom_wgs84_y','annee']].merge(
                                zone_cbl.loc[zone_cbl['type']=='Simple'].groupby('annee')['tv_moyjour'].sum(), left_on='annee', right_index=True).merge(
                                zone_cbl.loc[zone_cbl['type']=='Simple'].groupby('annee')['pl_pourcen'].mean(), left_on='annee', right_index=True).drop_duplicates(
                                ['annee','tv_moyjour'])
                    df=self.simplifier1PtComptage(zone_cbl, geom, wgs84_geom_x, wgs84_geom_y,num_zone,'Simple')
                    df['fictif']='t'
                    df['obs_supl']=df['obs_supl']+' ; comptage 2 sens'
                else : 
                    raise ValueError('cas non traite')
        elif (zone_cbl['type']=='DETECTEUR_PC').all() :
            geom, wgs84_geom_x, wgs84_geom_y=self.creer_geom('DETECTEUR_PC',carac_geom_type,zone_cbl_grp_type,zone_cbl, reprojection ) 
            if carac_geom_type['DETECTEUR_PC']==1 :
                df=self.simplifier1PtComptage(zone_cbl, geom, wgs84_geom_x, wgs84_geom_y,num_zone,'DETECTEUR_PC')
                df['obs_supl']=df['obs_supl']+' ; comptage 1 sens'
            else : 
                raise ValueError('cas non traite')
        else : 
            geom, wgs84_geom_x, wgs84_geom_y=self.creer_geom('Mix',carac_geom_type,zone_cbl_grp_type,zone_cbl, reprojection )
            if len(zone_cbl)==1 : 
                format_cpt=zone_cbl[['geometry','tv_moyjour','pl_pourcen']].copy()
                annee=zone_cbl.annee.values[0]
                format_cpt.rename(columns={'tv_moyjour':f'tmja_{annee}', 'pl_pourcen':f'pc_pl_{annee}'}, inplace=True)
                df=self.calcul_attributs_commun(format_cpt,geom, zone_cbl.nom_voie.values[0], wgs84_geom_x, wgs84_geom_y,num_zone)
                df['obs_supl']=df['obs_supl']+' ; comptage 1 sens' 
            else : 
                if carac_geom_type['Simple']==1 : 
                    df=self.simplifier1PtComptage(zone_cbl, geom, wgs84_geom_x, wgs84_geom_y,num_zone,'Simple')
                    df['obs_supl']=df['obs_supl']+' ; comptage 1 sens'
                elif carac_geom_type['Simple']==2:
                    zone_cbl=zone_cbl.loc[zone_cbl['type']=='Simple'][['type','geometry','date_deb','date_fin','id_groupe','nom_voie','position','geom_wgs84_x','geom_wgs84_y','annee']].merge(
                                zone_cbl.loc[zone_cbl['type']=='Simple'].groupby('annee')['tv_moyjour'].sum(), left_on='annee', right_index=True).merge(
                                zone_cbl.loc[zone_cbl['type']=='Simple'].groupby('annee')['pl_pourcen'].mean(), left_on='annee', right_index=True).drop_duplicates(
                                ['date_deb','date_fin','id_groupe','nom_voie','tv_moyjour','pl_pourcen']).drop_duplicates(['annee','tv_moyjour'])
                    df=self.simplifier1PtComptage(zone_cbl, geom, wgs84_geom_x, wgs84_geom_y,num_zone,'Simple')
                    df['fictif']='t'
                    df['obs_supl']=df['obs_supl']+' ; comptage 2 sens'
                else : 
                    raise ValueError('cas non traite')
        return df
    
    def creer_geom(self,cas, carac_geom_type,zone_cbl_grp_type,zone_cbl, reprojection):
        """
        creer la geometrie : differe si le type de cpt est 'Double','Simple','DETECTEUR_PC', ou 'Mix'. depend aussi du nombre de geom differentes
        in : 
            cas : string : 'Double','Simple','DETECTEUR_PC', ou 'Mix'
            carac_geom_type,zone_cbl_grp_type,zone_cbl : iisu de isoler_zone()
            reprojection : vecteur de transformation selon la methode de shapely
        """
        if (cas=='Double' or (cas in ('Simple','DETECTEUR_PC') and carac_geom_type[cas]==1) 
            or (cas=='Mix' and carac_geom_type['Simple']==1))  : 
            if cas=='Mix' and carac_geom_type['Simple']==1 : 
                cas='Simple'
            geom=zone_cbl_grp_type.geometry.unique()[cas][0]
            wgs84_geom_x=zone_cbl_grp_type.geom_wgs84_x.unique()[cas][0]
            wgs84_geom_y=zone_cbl_grp_type.geom_wgs84_y.unique()[cas][0]
        elif (cas=='Simple' and carac_geom_type[cas]==2) or (cas=='Mix' and carac_geom_type['Simple']==2) :
            list_geom=zone_cbl.loc[zone_cbl['type']=='Simple'].geometry.unique()
            geom=LineString(list_geom).centroid
            wgs84_geom_x=str(round(transform(reprojection, geom).x,3))
            wgs84_geom_y=str(round(transform(reprojection, geom).y,3))
        elif cas=='Mix' and len(zone_cbl)==1 : 
            geom=zone_cbl.geometry.values[0]
            wgs84_geom_x=str(round(transform(reprojection, geom).x,3))
            wgs84_geom_y=str(round(transform(reprojection, geom).y,3))
        return geom, wgs84_geom_x, wgs84_geom_y
    
    def df_regroupee(self):
        """
        produire une df globale de l'ensemble des points de comptages représentatifs
        """
        list_dfs=[]
        reprojection=O.reprojeter_shapely(None, '2154', '4326')[0]
        for i in self.df_trafic.id_groupe.unique() :
            zone_cbl, zone_cbl_grp_type, carac_geom_type=self.isoler_zone(i)
            if zone_cbl.empty : 
                continue
            list_dfs.append(self.df_intermediaire(zone_cbl, zone_cbl_grp_type, carac_geom_type, reprojection, i))
        return list_dfs
    
    def df_regroupee_complete(self, list_dfs):
        """
        ajouter à la df creee par df_regroupee() les attributs généraux. creer le fichier self.df_attr
        """
        def type_compteur(ligne, list_colonnes):
            date_range = range(2015, 2021)
            if (all([f'tmja_{annee}' in list_colonnes for annee in date_range]) and 
                all([not pd.isnull(ligne[f'tmja_{annee}']) for annee in (str(i) for i in date_range)]))  : 
                return 'permanent'
            elif any([f'tmja_{annee}' in list_colonnes for annee in date_range]): 
                i = 0
                for col in [f'tmja_{annee}' for annee in date_range if f'tmja_{annee}' in list_colonnes] : 
                    if not pd.isnull(ligne[col]) : 
                        i += 1
                if i >= 3 :
                    return 'tournant'
                else : 
                    return 'ponctuel'
            else :
                return 'ponctuel'
        gdf = gp.GeoDataFrame(pd.concat(list_dfs, axis=0, sort=False))
        gdf['route'] = gdf.id_comptag.apply(lambda x : x.split('-')[1])
        gdf['type_poste'] = gdf.apply(lambda x : type_compteur(x, gdf.columns), axis=1)
        self.df_attr = gdf.assign(dep='87',reseau='VC', gestionnai='Limoges Metropole',concession='N',
                       x_l93=lambda x : round(x.geometry.x,3),y_l93=lambda x : round(x.geometry.y,3))
        
class Comptage_LaRochelle(Comptage):  
    """
    pour LaRochelle n dispose de fichierde comptage .xls de chez sterela, à combiner pour obtenir les deux sens
    POur ce fair on s'appuie sur une création mano de l agéométrie des points de comptage, qui recencse dans l'attr id_cpt les noms
    des fchiers concernés
    """ 
    def __init__(self, dossier):  
        """
        in : 
            dossier : chemin du dossier contenant tous les fchiers
        """  
        self.dossier=dossier
    
    def extraireDonneesSterela(self,fichier):
        """
        sortir le tmja, pc_pl et les df horaires PL et TV d'unfichier de comptage Sterela 
        """
        print(fichier)
        df=pd.read_excel(os.path.join(self.dossier,fichier))
        tmja=round(df.loc[38,'Unnamed: 6'])
        pl=round(df.loc[42,'Unnamed: 6'])
        periode=df.iloc[1,16]
        index_date=pd.date_range(start=pd.to_datetime(periode.split(' au ')[0][3:], dayfirst=True), end=pd.to_datetime(periode.split(' au ')[1], dayfirst=True))
        df_pl=df.iloc[17:24,2:26].copy()
        df_tv=df.iloc[27:34,2:26].copy()
        df_pl.columns=[f'h{str(i)}_{str(i+1)}' for i in range(24)]
        df_tv.columns=[f'h{str(i)}_{str(i+1)}' for i in range(24)]
        df_pl=df_pl.assign(jour=index_date, type_veh='PL')
        df_tv=df_tv.assign(jour=index_date, type_veh='TV')
        return tmja, pl, periode, df_tv, df_pl
     
    def listCptCreer(self):
        """
        interroger la bdd pour cpnnaitre les cpt localisés à la main avec les noms de fichiers inscrits
        """ 
        with ct.ConnexionBdd('gti_otv_pg11') as c : 
            rqt="select id_comptag, id_cpt from comptage.na_2010_2019_p where id_cpt is not null"
            ids=pd.read_sql(rqt, c.sqlAlchemyConn)
        return ids
        
    def sommer2sens(self, id_comptag, id_cpt):
        """
        calculer les données argegees et horaires à partir d'un id_comptag et de la référence aux fichier
        in : 
            id_comptag : string, issu d'une requete sql sur la table comptag cf listCptCreer()
            id_cpt : string, issu d'une requete sql sur la table comptag, continet le nom des fchiers separes par ';' cf listCptCreer()
        """ 
        Donnees2Sens=list(zip(*[self.extraireDonneesSterela(fichier) for fichier in [f"SEMAINE {i}{os.sep}{a.replace('SEMAINE 1',f'SEMAINE {i}')}" for a in id_cpt.split(';') for i in range(1,4)]]))
        tmja=sum(Donnees2Sens[0])
        pl,pc_pl=sum(Donnees2Sens[1]),round(sum(Donnees2Sens[1])*100/tmja,2)
        df_tv=pd.concat([*Donnees2Sens[3]], axis=0, sort=False).groupby('jour').sum().assign(type_veh='TV', id_comptag=id_comptag).reset_index()
        df_pl=pd.concat([*Donnees2Sens[4]], axis=0, sort=False).groupby('jour').sum().assign(type_veh='PL', id_comptag=id_comptag).reset_index()  
        return tmja, pl,pc_pl, df_tv, df_pl
    
    def creer_dfs(self,ids):
        """
        fusionner l'ensemble des données de trafic dans les df des données agrégées et celes des données horaires
        in : 
            ids : la liste des id_comptag et id_cpt issues de listCptCreer
        """
        list_id_comptag,list_fichiers=ids.id_comptag.tolist(), ids.id_cpt.tolist()
        list_tmja, list_pc_pl, list_df_tv, list_df_pl=[],[],[],[]
        for id_comptag,id_cpt in zip(list_id_comptag,list_fichiers) : 
            tmja, pl,pc_pl, df_tv, df_pl=self.sommer2sens(id_comptag, id_cpt)
            list_tmja.append(tmja)
            list_pc_pl.append(pc_pl)
            list_df_tv.append(df_tv)
            list_df_pl.append(df_pl)
        self.df_attr=pd.DataFrame({'id_comptag':list_id_comptag, 'tmja_2015':list_tmja, 'pc_pl_2015':list_pc_pl,'src_2015':'agglo LaRochelle',
                      'fichier':[self.dossier+a for a in list_fichiers]})
        self.df_attr_horaire=pd.concat([*list_df_tv,*list_df_pl],axis=0, sort=False)
        
    def update_bdd_LaRochelle(self, schema, table):
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr,['id_comptag','tmja_2015','pc_pl_2015','src_2015', 'fichier'])
        dico_attr_modif={f'tmja_2016':'tmja_2015', f'pc_pl_2016':'pc_pl_2015',f'src_2016':'src_2015', 'fichier':'fichier'}
        self.update_bdd(schema, table, valeurs_txt,dico_attr_modif)  
        
class Comptage_Anglet(Comptage):
    """
    basé sur l'utilisation des fichiers .mdb au format MHCorbin du module Donnees_sources
    fonctionne par annee ; utilisation manuelle de 'mon geocodeur' IGN en cours de route
    """
    
    def __init__(self, annee, dossier, dossierResume):
        """
        aatribuut : 
            anneee : string : 4 caracteres
            dossier : rawstring du path du dossier contenant les fihcires
            dossierResume : rawstring du path du dossier allant recevoir le fichier de geoloc
            codepostal : 64600
            commune : Anglet
            gdfGeoloc : gdf avec le nom de fichier, des infos sur la geoloc et les attributs necessaires pour creer des compteurs dans bdd
            dfHoraire : dataframe des donnees horaires, tous points confondus
            dfIndicAgrege : dataframe des donnees agreges, tous points confondus
        """
        self.annee=annee 
        self.dossier=dossier
        self.codepostal='64600'
        self.commune='Anglet'
        self.resumeComptage=os.path.join(dossierResume,f'Comptage{annee}.csv')
        self.resumeComptageGeoloc=os.path.join(dossierResume,f'Comptage{annee}_geoloc.csv')
    
    def exporterCsvAGeocoder(self):
        """
        creer le csv qui va permettre le geocodage manuel dans 'monGeocodeur de l'IGN'
        """
        #lister les fichiers et exporter le fichier csv pour géocodage IGN
        dicoAdresse={'fichier':[], 'adresse':[], 'codePostal':self.codepostal, 'Commune':self.commune}
        #exporter les adresses (géocodées ensuite via mongeocodeur)
        def replaceAdresse(matchobj):
            return matchobj.group(0).replace('.',',')
        for files in os.listdir(self.dossier):
            if os.path.isfile(os.path.join(self.dossier, files)) and files.endswith('.mdb'):
                dicoAdresse['fichier'].append(os.path.join(self.dossier, files))
                dicoAdresse['adresse'].append(re.sub('^[1-9]+\.', replaceAdresse,files[:-4] ))
        dfFichierAdresse=pd.DataFrame(dicoAdresse)
        dfFichierAdresse.to_csv(self.resumeComptage, sep=';')  
        return 
    
    def creerDfGeoloc(self):  
        """
        creer la df qui va contenir tous les attributs relatifs aux compteurs
        """
        #importer le fichier des points de comptage geolocalise et le mettre en forme
        gdfGeoloc=gp.read_file(self.resumeComptageGeoloc, encoding='utf-8')#C:\Users\martin.schoreisz\Documents\temp\OTV\Anglet\Comptage2019_geoloc.csv
        gdfGeoloc.geometry=gp.points_from_xy(x=gdfGeoloc['longitude'], y=gdfGeoloc['latitude'], crs=4326)
        gdfGeoloc=gdfGeoloc.to_crs('epsg:2154')
        gdfGeoloc['id_comptag']=gdfGeoloc.apply(lambda x : 'Anglet-'+re.sub('_{2,50}', '_',re.sub('(,|\.|\'| +)', '_', x['adresse']))+f"-{round(float(x['longitude']), 4)};{round(float(x['latitude']),4)}", axis=1)
        gdfGeoloc['route']=gdfGeoloc.adresse.apply(lambda x : re.sub('^([1-9]{1,3},)+','', x))
        gdfGeoloc['reseau']='VC'
        gdfGeoloc['dep']='64'
        gdfGeoloc['gestionnai']='Anglet'
        gdfGeoloc['concession']=False
        gdfGeoloc['type_poste']='ponctuel'
        gdfGeoloc['src_geo']='adresse'
        gdfGeoloc['obs_geo']=gdfGeoloc.apply(
            lambda x : f"geocodé avec 'mon géocodeur' de l'IGN ; qualite : {x.qualite} ; precision geocodage : {x['precision geocodage']} ; ID adresse : {x['ID adresse']}; adresse geocodee : {x['adresse geocodee']}", axis=1)
        gdfGeoloc['x_l93']=round(gdfGeoloc.geometry.x,3)
        gdfGeoloc['y_l93']=round(gdfGeoloc.geometry.y,3)
        gdfGeoloc['fictif']=False
        gdfGeoloc['src_cpt']='gestionnaire'
        gdfGeoloc['convention']=False
        self.gdfGeoloc=gdfGeoloc
        return
     
    def indicsTousFichiers(self):
        """
        a partir de la gdfGeoloc, pour chaque fihcier on va recuperer les donnees horaire et calculer les agrege
        """
        listDfHoraires=[]
        for row in self.gdfGeoloc.itertuples(index=False) : 
            print(row.fichier)
            try :
                mhc=MHCorbin(row.fichier)
            except DataSensError : 
                print(f'Erreur de sens sur le fichier {self.row.fichier}')
                continue
            listDfHoraires.append(mhc.formaterDonneesHoraires(mhc.formaterDonneesIndiv(mhc.dfAgreg2Sens)).reset_index().assign(
                id_comptag=row.id_comptag,
                fichier=mhc.fichierCourt,
                obs_supl=f"{mhc.dfAgreg2Sens.obs_supl.values[0]} ; {mhc.dfAgreg2Sens.sensTxt.values[0]}",
                sens_cpt=mhc.dfAgreg2Sens.nb_sens.values[0],
                note_manuelle_qualite=mhc.indicQualite,
                obs_qualite=mhc.comQualite))
        #calcul des dF d'indicateurs
        self.dfHoraire=pd.concat(listDfHoraires)
        self.dfIndicAgrege=tmjaDepuisHoraire(self.dfHoraire.assign(annee=self.annee)).merge(self.dfHoraire[['id_comptag','obs_supl','sens_cpt', 
                                                    'note_manuelle_qualite', 'obs_qualite', 'fichier']].drop_duplicates(), on='id_comptag', how='left')
        return
           
    def plusProcheVoisinBddRegroupe(self, tableLinauto, distance=15):
        """
        passer la df dans la bdd pour trouver le plus proche voisin selon le regroupement du schema linauto
        in : 
            tableLinauto : string : nom de la table dans le chema linauto a uinterroger
            distance : integer : distance au dela de laquelle on en echerche plus de voisin
        out : 
            dfPPV : dataframe du plud proche voisin, sur la base de gdfGeoloc
        """  
        self.insert_bdd('public', f'cpt_anglet_{self.annee}',self.gdfGeoloc.drop('fichier', axis=1), if_exists='replace' )  
        rqt=f"""WITH 
                cpteur_base AS (
                SELECT DISTINCT ON (c.id_comptag) c.id_comptag id_comptag_bdd, t.gid, st_distance(c.geom, t.geom)
                 FROM comptage.compteur c left JOIN linauto.{tableLinauto} t ON st_dwithin(c.geom, t.geom, {distance})
                 ORDER BY c.id_comptag, st_distance(c.geom, t.geom))
                SELECT DISTINCT ON (c.field_1) 
                                    c.*, t.gid, st_distance(c.geom, t.geom),
                                    c2.id_comptag_bdd
                 FROM public.cpt_anglet_{self.annee} c left JOIN linauto.{tableLinauto} t ON st_dwithin(c.geom, t.geom, {distance})
                                               LEFT JOIN cpteur_base c2 ON t.gid=c2.gid
                 ORDER BY c.field_1, st_distance(c.geom, t.geom)"""
        with ct.ConnexionBdd(nomConnBddOtv) as c : 
            dfPPV=gp.read_postgis(rqt, c.sqlAlchemyConn, crs='epsg:2154').merge(self.dfIndicAgrege, on='id_comptag', how='left'
                            ).merge(periodeDepuisHoraire(self.dfHoraire.assign(annee=self.annee)), how='left', on='id_comptag')
            dfPPV=O.gp_changer_nom_geom(dfPPV, 'geometrie')
        return dfPPV
    
    def isolerComptagesDoublons(self,dfPPV ):
        """
        obtenir la liste des comptages situes sur un mm troncon homogene de trafic
        in : 
            dfPPV : donnees des compteurs, avec info supp, issue de plusProcheVoisinBddRegroupe
        out : 
            numpy array avec les id_comptages (cree automatiquementdans le cadre des donnees, pas ceux de la bdd) situes sur des troncon homogene equivalent
        """
        return dfPPV.loc[(dfPPV.duplicated(['gid', 'indicateur'], keep=False)) & (~dfPPV.gid.isna())]
    
    def cptSsDblEtSsGeom(self, dfPPV, dicoAssoc):
        """
        récupérer la df des comptages sans doublons, mais avec les geometries
        in : 
            dfPPV : donnees des compteurs, avec info supp, issue de plusProcheVoisinBddRegroupe
            DicoAssoc : dico des id_comptag creer a partir des donnees, enclé: les id_comptag des cpt ref, en value une liste des id_comptag des cpt assoc
        out : 
            dfPPVCptRef : df des comptages sans doublon pour insertion dans bdd
            dfPPVCptAssoc : df des comptages associes a retravailler avant insertion
        """
        dfPPVFiltreAconserver=dfPPV.loc[~dfPPV.id_comptag.isin([v for e in dicoAssoc.values() for v in e]) & (dfPPV.indicateur=='tmja') & (~dfPPV.gid.isna())]
        if not (dfPPVFiltreAconserver.gid.value_counts()==1).all() : 
            raise ComptageMultipleSurTronconHomogeneError(list(dfPPVFiltreAconserver.loc[dfPPVFiltreAconserver.gid.isin(dfPPVFiltreAconserver.gid.value_counts()!=1)].gid.unique()))
        dfPPVCptRef=dfPPV.loc[dfPPV.id_comptag.isin(dfPPVFiltreAconserver.id_comptag.tolist()+dfPPV.loc[dfPPV.gid.isna()].id_comptag.tolist())].copy()
        dfPPVCptAssoc=dfPPV.loc[dfPPV.id_comptag.isin([v for e in dicoAssoc.values() for v in e])].merge(
                pd.DataFrame.from_dict(dicoAssoc, orient='index').reset_index().rename(columns={'index':'cpt_ref', 0:'cpt_assoc'}), left_on='id_comptag', right_on='cpt_assoc')
        #attentionaux doublons dus aux fichiers en double
        dfPPVCptRef.drop_duplicates(['indicateur','valeur', 'periode'], inplace=True)
        dfPPVCptAssoc.drop_duplicates(['indicateur','valeur', 'periode'], inplace=True)
        return dfPPVCptRef, dfPPVCptAssoc
    
    def insererCompteur(self, dfPPVSsDbl, schema='comptage', table='compteur'):
        """
        Preparer les donnes liees au compteur et les inserer (si besoin) dans la bdd
        in : 
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
        """
        #creer les compteurs a inserer dans la bdd
        dfCptAcreer=dfPPVSsDbl.loc[dfPPVSsDbl.id_comptag_bdd.isna()]
        if not dfCptAcreer.empty :
            compteur=dfCptAcreer[attBddCompteur].drop_duplicates().copy()
            #inserer
            self.insert_bdd(schema, table, compteur)
        else : 
            print('pas de nouveau compteur a inserer')
            return
        
    def insererComptage(self, dfPPVSsDbl, schema='comptage', table='comptage'):
        """
        Preparer les donnes liees au comptages et les inserer dans la bdd
        in : 
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
        """
        dfPPVSsDblTemp=dfPPVSsDbl.copy()
        dfPPVSsDblTemp['id_comptag_final']=dfPPVSsDblTemp.apply(lambda x : x.id_comptag_bdd if not pd.isnull(x.id_comptag_bdd) else x.id_comptag, axis=1)
        comptage=dfPPVSsDblTemp[['id_comptag_final', 'annee', 'periode']].assign(src='gestionnaire', type_veh='tv/pl').rename(columns={'id_comptag_final':'id_comptag'}).drop_duplicates()
        #inserer
        self.insert_bdd(schema, table,comptage)
        return
        
    def insererIndics(self, dfPPVSsDbl,schema='comptage', tableAgrege='indic_agrege', tableHoraire='indic_horaire'):
        """
        Preparer les donnes liees au indicateurs(horaires, agrege) et les inserer dans la bdd
        in : 
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
        """
        dfPPVSsDblTemp=dfPPVSsDbl.copy()
        dfPPVSsDblTemp['id_comptag_final']=dfPPVSsDblTemp.apply(lambda x : x.id_comptag_bdd if not pd.isnull(x.id_comptag_bdd) else x.id_comptag, axis=1)
        indicAgrege=dfPPVSsDblTemp.merge(self.recupererIdUniqComptage(dfPPVSsDblTemp.id_comptag_final.tolist(), self.annee)
                             , left_on=['id_comptag_final', 'annee'], right_on=['id_comptag', 'annee'])[['id_comptag_uniq', 'indicateur', 'valeur', 'fichier']]
        dfHoraireSsDbl=self.dfHoraire.merge(dfPPVSsDblTemp[['id_comptag', 'id_comptag_final']], on='id_comptag')
        indicHoraire=dfHoraireSsDbl.assign(annee=self.annee, 
            indicateur=dfHoraireSsDbl.indicateur.str.upper()).merge(self.recupererIdUniqComptage(
            dfHoraireSsDbl.id_comptag_final.tolist(), self.annee).rename(columns={'id_comptag':'id_comptag_final'}), on=['id_comptag_final', 'annee']).drop([
            'sens_cpt', 'note_manuelle_qualite', 'obs_qualite', 'annee', 'id_comptag','id_comptag_final', 'obs_supl'], axis=1)
        self.insert_bdd(schema, tableAgrege,indicAgrege)
        self.insert_bdd(schema, tableHoraire,indicHoraire)
        return
        
    def insererDatas(self, dfPPVSsDbl, schema='comptage', tableComptag='comptage',
                     tableCpteur='compteur', tableAgrege='indic_agrege', tableHoraire='indic_horaire'):
        """
        preparer et inserere les donnees de compteurs, comptages et indicateurs
        in : 
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
        """
        self.insererCompteur(dfPPVSsDbl, schema, tableCpteur)
        self.insererComptage(dfPPVSsDbl, schema, tableComptag)
        self.insererIndics(dfPPVSsDbl, schema, tableAgrege, tableHoraire)
        
        
    def insererQualiteFaible(self, dfPPVCptRef, schema='qualite', table='qualite_manuelle_cptag'):
        """
        inserer les elements de qualite faible des comptages precedemment inseres dans la bdd
        in : 
            dfPPVCptRef : df des comptages de references inseres dans la bdd
        """  
        O.checkAttributsinDf(dfPPVCptRef, ['id_comptag_bdd', 'id_comptag','note_manuelle', 'obs'])
        dfQualiteFaible=dfPPVCptRef.loc[dfPPVCptRef.note_manuelle==1].copy()
        #y associer un id_comptage_final
        dfQualiteFaible['id_comptag_final']=dfQualiteFaible.apply(lambda x : x.id_comptag_bdd if not pd.isnull(x.id_comptag_bdd) else x.id_comptag, axis=1)
        #joindre avec les comptages existant
        dfQualiteFaibleAInsere=dfQualiteFaible.merge(self.recupererIdUniqComptage(dfQualiteFaible.id_comptag_final.tolist(), self.annee), 
                              left_on='id_comptag_final', right_on='id_comptag')[['id_comptag_uniq','note_manuelle','obs'
                                ]].drop_duplicates(['id_comptag_uniq','note_manuelle','obs'])
        self.insert_bdd(schema, table,dfQualiteFaibleAInsere)
        return                        
    
    def creerComptageAssoc(self, dfPPVCptAssoc):    
        """
        creer la table des comptage du schema comptage_assoc
        in  : 
            dfPPVCptAssoc : df des comptages associes issu de cptSsDblEtSsGeom
        out : 
            dfAssoc : df au format de la table en bdd
        """
        dfPPVCptAssocInterne=dfPPVCptAssoc.copy()
        dfPPVCptAssocInterne['id_comptag_final']=dfPPVCptAssocInterne.apply(lambda x : x.id_comptag_bdd if not pd.isnull(x.id_comptag_bdd) else x.cpt_ref, axis=1)
        #trouver son id_comptag_uniq
        dfAssoc=dfPPVCptAssocInterne.merge(self.recupererIdUniqComptage(dfPPVCptAssocInterne.id_comptag_final.tolist(), self.annee), 
                                      left_on='id_comptag_final', right_on='id_comptag')
        #ajouter les colonnes manquantes
        dfAssocInsert=dfAssoc.assign(src='gestionnaire', type_veh='tv/pl', obs=None).drop_duplicates(['id_comptag_uniq', 'periode'])
        dfAssocInsert['rang']=dfAssocInsert.assign(toto=1).groupby('id_comptag_uniq').toto.rank(method='first')
        #inserer les comptages
        self.insert_bdd(schemaComptageAssoc, tableComptage, dfAssocInsert[['id_comptag_uniq', 'periode', 'src', 'type_veh', 'rang', 'obs']].rename(columns={'id_comptag_uniq':'id_cptag_ref'}))
        return dfAssoc
    
    def creerIndicsAssoc(self, dfAssoc, schema='comptage_assoc', tableAgreg='indic_agrege', tableHoraire='indic_horaire'):
        """
        creer les tables en bdd des indcis_agreges et indic_horaires pour les comptages associes
        in : 
            dfAssoc : df issu de creerComptageAssoc
        out : 
            dfIndicAgregeAssoc : df integree a la bdd et sortie pour ifo
            dfHoraireassoc : df integree a la bdd et sortie pour ifo
        """
        dfIndicAgregeAssoc=dfAssoc.merge(self.recupererIdUniqComptageAssoc(dfAssoc.id_comptag_uniq.tolist()).rename(
            columns={'id_comptag_uniq':'id_comptag_uniq_assoc'}), left_on='id_comptag_uniq', right_on='id_cptag_ref')[['id_comptag_uniq_assoc', 'indicateur', 'valeur', 'fichier']]
        dfHoraireassoc=self.dfHoraire.loc[self.dfHoraire.fichier.isin(dfIndicAgregeAssoc.fichier.tolist())].merge(dfIndicAgregeAssoc[['id_comptag_uniq_assoc', 'fichier']], on='fichier').drop(
            ['id_comptag','obs_supl','sens_cpt', 'note_manuelle_qualite', 'obs_qualite'], axis=1).rename(columns={'id_comptag_uniq_assoc':'id_comptag_uniq'})
        dfHoraireassoc['indicateur']=dfHoraireassoc.indicateur.str.upper()    
        self.insert_bdd(schema, tableAgreg,dfIndicAgregeAssoc.rename(columns={'id_comptag_uniq_assoc':'id_comptag_uniq'}))
        self.insert_bdd(schema, tableHoraire,dfHoraireassoc)
        return dfIndicAgregeAssoc, dfHoraireassoc
        
class Comptag_Angouleme(Comptage):
    """
    constrcution en 2021 pour traiter des données 2020 au format MHCorbin (.mdb). Souvent 1 fichier par sens
    """
    def __init__(self, annee, dossierSource):
        """
        attributs : 
            annee : string 4 caractères
            dossierSource rw string du dossier pointant vers le repertoir "en_cours" du dossier partagé OTV dédié à Angouleme
            gdfGeoloc : df de geolocalisatiopn avec donnees de la table compteur
            dfHoraire : dataframe correspondant à la table indic_horaire
            dfIndicAgrege : dataframe correspondant à la table indic_agrege
        """
        self.annee = annee
        self.dossierSource = dossierSource
        
    def extraireDonneesBrutes(self):
        """
        à partir du dossier source, parcourir tous les fichiers et en extraire les données
        out : 
            dFDataBrutes : df horaire avec attributs bdd (sauf id_comptag)
            dfHshdr : concatenation des donnees de la tale hshdr des fichier MHCorbin
        """
        listHshdr=[]
        listDfDataBrutes=[]
        for root, dirs, files in os.walk(self.dossierSource) : 
            for f in files : 
                fichier=os.path.join(root, f)
                if f.endswith('.mdb') : 
                    try : 
                        cpt=MHCorbin(fichier)
                        listHshdr.append(cpt.dicoTables['hshdr'].assign(fichier=fichier))
                        listDfDataBrutes.append(cpt.formaterDonneesHoraires(cpt.formaterDonneesIndiv(cpt.dfAgreg2Sens)).assign(fichier=fichier))
                    except ValueError : 
                        print(f'pb sur fichier {fichier}')
                        continue
        dFDataBrutes=pd.concat(listDfDataBrutes)
        dfHshdr=pd.concat(listHshdr)
        return dFDataBrutes, dfHshdr
    
    def assignerDossierRefEtAdresse(self, dFDataBrutes, dfHshdr):
        """
        ajouter aux données brutes les attributs de descriptions du dossier de réference, pour regroupement des différents sens
        ajouter l'adresse aux données HsHdr
        lève une érreur si un comptage n'as pas de repertoir de référence
        in : 
            dFDataBrutes : df horaire avec attributs bdd (sauf id_comptag), issu de extraireDonneesBrutes()
            dfHshdr : concatenation des donnees de la tale hshdr des fichier MHCorbin, issu de extraireDonneesBrutes()
        """
                
        # isoler les donnes fournies par la commune dans le path
        dFDataBrutes['spliPath']=dFDataBrutes.fichier.str.split('en_cours')
        dfHshdr['spliPath']=dfHshdr.fichier.str.split('en_cours')
        # trouver le nb de sous-dossier
        dFDataBrutes['nb_niv_dir']=dFDataBrutes.spliPath.apply(lambda x : len(PureWindowsPath(x[1]).parents))
        dfHshdr['nb_niv_dir']=dfHshdr.spliPath.apply(lambda x : len(PureWindowsPath(x[1]).parents))
        # trouver le dossier parent représentatif des deux sens A MODIFIER : SI " VERS " dans un dossier parent, alors on prend celui d'avant
        dFDataBrutes.loc[dFDataBrutes.nb_niv_dir==4, 'dir_ref']=dFDataBrutes.loc[dFDataBrutes.nb_niv_dir==4].spliPath.apply(lambda x : PureWindowsPath(x[1]).parents[1].name)
        dFDataBrutes.loc[dFDataBrutes.nb_niv_dir.isin((2,3)), 'dir_ref']=dFDataBrutes.loc[dFDataBrutes.nb_niv_dir.isin((2,3))
                                                                                         ].spliPath.apply(lambda x : PureWindowsPath(x[1]).parents[0].name if not 'vers ' in PureWindowsPath(x[1]).parts[-2]
                                                                                                         else PureWindowsPath(x[1]).parents[1].name)
        dfHshdr.loc[dfHshdr.nb_niv_dir==4, 'dir_ref']=dfHshdr.loc[dfHshdr.nb_niv_dir==4].spliPath.apply(lambda x : PureWindowsPath(x[1]).parents[1].name)
        dfHshdr.loc[dfHshdr.nb_niv_dir.isin((2,3)), 'dir_ref']=dfHshdr.loc[dfHshdr.nb_niv_dir.isin((2,3))
                                                                                         ].spliPath.apply(lambda x : PureWindowsPath(x[1]).parents[0].name if not 'vers ' in PureWindowsPath(x[1]).parts[-2]
                                                                                                         else PureWindowsPath(x[1]).parents[1].name)
        # verif que tout le monde a un dossier parent
        if dFDataBrutes.dir_ref.isna().any() or dfHshdr.dir_ref.isna().any():
            raise ValueError('il y a un comptage sans repertoire de référence') 
        # creation de l'adresse si possible
        dfHshdr['adresse']=dfHshdr.apply(lambda x : re.sub('(numero|n°)', '', re.search('(numero *[0-9]+|n° *[0-9]+)', re.sub('é|è|ê', 'e', x.Location.lower()))[0]) + f' {x.Street}' if re.search('(numero|n°)', 
                        re.sub('é|è|ê', 'e', x.Location.lower())) else np.NaN, axis=1)
        
    def adresseUniques(self,dfHshdr ):
        """
        à partir du fichier dfHshdr, creer une df avec un adresse unique par repetroire de référence
        in : 
            dfHshdr : concatenation des donnees de la tale hshdr des fichier MHCorbin, avec Adresse, issu de assignerDossierRefEtAdresse()
        """
        def uniformiserAdresse(groupe) : 
            """Fonction interne pour repercuter l'adresse d'un fihcier vers tous ceux associé au même dossier"""
            if len(set(groupe))==1 and not any([pd.isnull(e) for e in groupe]) : 
                return list(groupe)[0]
            elif len(set(groupe))>1 and not all([pd.isnull(e) for e in groupe]) :
                return [e for e in groupe if not pd.isnull(e)][0]
            else : 
                return None
        # uniformiser les adressses d'un même dossier
        dfAdresse=dfHshdr[['fichier','dir_ref', 'adresse']].groupby('dir_ref').agg({'adresse': lambda x : uniformiserAdresse(x)}).reset_index()
        return dfAdresse
    
    def assignerAdressesManuelles(self, dfAdresse, dicoAdresseMano):
        """
        ajouter des adresses manuelles aux dir_ref ne présentant pas une adresse complete en numero, type de rue et nom de rue
        in : 
            dfAdresse : df des adresses, issus de adresseUniques()
            dicoAdresseMano : dico cree manuellement avec en clé le dir_ref de dfAdresse et en value une adresse en numero et 
                              type de et nom de rue. peux contenir des coordonnées WGS84 en '46.789654, 0.145863' par exemple (au moins 4 chiffres apres le '.'
        out : 
            dfAdresse2 : df adresse sans valeur Na dans le champs adresse (car complete par le dico)
        """
        # vrerif
        if not len(dicoAdresseMano)==len(dfAdresse.loc[dfAdresse['adresse'].isna()].dir_ref.unique()) : 
            raise ValueError('le nombre d\'adresse inconnue et d\'adresse fournie par dico doit etre le mm')
        # assignation des nouvelles adresses
        dfAdresse2=dfAdresse.merge(
            pd.DataFrame.from_dict(dicoAdresseMano, orient='index', columns=['adresse']).reset_index().rename(columns={'index':'dir_ref'}), on='dir_ref', how='left')
        #verif finale
        dfAdresse2['adresse']=dfAdresse2.apply(lambda x: x.adresse_x if not pd.isnull(x.adresse_x) else x.adresse_y, axis=1)
        dfAdresse2.drop(['adresse_x', 'adresse_y'], axis=1, inplace=True)
        if not dfAdresse2.adresse.notna().all() :
            raise ValueError 
        return dfAdresse2
    
    def exporterCsvAGeocoder(self, dfAdresse2):
        """
        à partir du fichier des adresse, on exporte toute celles qu ne sont pas relatives à des coordonnées WGS84
        in : 
            dfAdresse2 : df adresse sans valeur Na dans le champs adresse (car complete par le dico) issu de assignerAdressesManuelles()
        """
        dfAdresse2.loc[dfAdresse2.adresse.apply(lambda x : False if re.search('[0-9]{2}\.[0-9]{4,9}', x) else True)].assign(
            codepostal='16000', commune='Angoulême').to_csv(os.path.join(self.dossierSource,f'Comptage{self.annee}.csv'))
            
    def creerDfGeoloc(self, dfAdresse2):
        """
        creer la df geolocalise avec tous les atrributs de la table compteur, en associant la geoloc par IGN et les adresse en coordonnees
        wgs84
        in : 
            dfAdresse2 : df adresse sans valeur Na dans le champs adresse (car complete par le dico) issu de assignerAdressesManuelles()
        """
        gdfGeoloc=gp.read_file(os.path.join(self.dossierSource,f'Comptage{self.annee}_geoloc.csv'), encoding='utf-8')#C:\Users\martin.schoreisz\Documents\temp\OTV\Anglet\Comptage2019_geoloc.csv
        gdfGeoloc.geometry=gp.points_from_xy(x=gdfGeoloc['longitude'], y=gdfGeoloc['latitude'], crs=4326)
        gdfPoints=gp.GeoDataFrame(dfAdresse2.loc[dfAdresse2.adresse.apply(lambda x : True if re.search('[0-9]{2}\.[0-9]{4,9}', x) else False)],
            geometry=gp.points_from_xy(
                dfAdresse2.loc[dfAdresse2.adresse.apply(lambda x : True if re.search('[0-9]{2}\.[0-9]{4,9}', x) else False)].adresse.apply(
                lambda x : float(x.split(', ')[1])),
                dfAdresse2.loc[dfAdresse2.adresse.apply(lambda x : True if re.search('[0-9]{2}\.[0-9]{4,9}', x) else False)].adresse.apply(
                lambda x : float(x.split(', ')[0])), crs='epsg:4326'))
        gdfPoints['longitude']=gdfPoints.adresse.apply(lambda x : x.split(', ')[1])
        gdfPoints['latitude']=gdfPoints.adresse.apply(lambda x : x.split(', ')[0])
        gdfGeoloc=pd.concat([gdfPoints,gdfGeoloc])
        gdfGeoloc=gdfGeoloc.to_crs('epsg:2154')
        gdfGeoloc['id_comptag'] = gdfGeoloc.apply(lambda x : 'Angouleme-'+re.sub('_{2,50}', '_',re.sub('(,|\.|\'| +)', '_', x['adresse'].strip()))+f"-{round(float(x['longitude']), 4)};{round(float(x['latitude']),4)}" 
                                                  if not pd.isnull(x.codepostal) else
                                                  'Angouleme-'+f"-{round(float(x['longitude']), 4)};{round(float(x['latitude']),4)}", axis=1)
        gdfGeoloc['route'] = gdfGeoloc.apply(lambda x : re.sub('^([1-9]{1,3},)+','', x.adresse) if not pd.isnull(x.codepostal) else np.NaN, axis=1)
        gdfGeoloc['reseau'] = 'VC'
        gdfGeoloc['dep']  ='16'
        gdfGeoloc['gestionnai'] = 'Angouleme'
        gdfGeoloc['concession'] = False
        gdfGeoloc['type_poste'] = 'ponctuel'
        gdfGeoloc['src_geo'] = gdfGeoloc.apply(lambda x : 'adresse' if not pd.isnull(x.codepostal) else 'manuelle', axis=1)
        gdfGeoloc['obs_geo'] = gdfGeoloc.apply(
            lambda x : f"geocodé avec 'mon géocodeur' de l'IGN ; qualite : {x.qualite} ; precision geocodage : {x['precision geocodage']} ; ID adresse : {x['ID adresse']}; adresse geocodee : {x['adresse geocodee']}"
                        if not pd.isnull(x.codepostal) else "pas d'adressee à proximite", axis=1)
        gdfGeoloc['x_l93'] = round(gdfGeoloc.geometry.x,3)
        gdfGeoloc['y_l93'] = round(gdfGeoloc.geometry.y,3)
        gdfGeoloc['fictif'] = False
        gdfGeoloc['src_cpt'] = 'gestionnaire'
        gdfGeoloc['convention'] = False
        self.gdfGeoloc=gdfGeoloc
    
    def creerIndicsTousFichiers(self, dFDataBrutes):
        """
        ajouter l'id_comptag à la dFDataBrutes pour créer le dfHoraire et le dfIndicAgrege
        """
        #r apatriement de l'id_comptag
        dfDataBrutesIdcomptag = dFDataBrutes.reset_index().merge(self.gdfGeoloc[['dir_ref', 'id_comptag']], how='left')
        # verif
        if dfDataBrutesIdcomptag.id_comptag.isna().any() : 
            raise ValueError('il manque un id_comptag sur une valeur')
        O.checkAttributsinDf(dfDataBrutesIdcomptag, attributsHoraire+['jour', 'id_comptag', 'fichier', 'indicateur'])
        dfDataBrutesIdcomptag['indicateur'] = dfDataBrutesIdcomptag.indicateur.str.upper()
        self.dfHoraire=dfDataBrutesIdcomptag[attributsHoraire+['jour', 'id_comptag', 'fichier', 'indicateur']]
        self.dfIndicAgrege=tmjaDepuisHoraire(self.dfHoraire.assign(annee=self.annee))
        
    def plusProcheVoisinBddRegroupe(self, tableLinauto, distance=15):
        """
        passer la df dans la bdd pour trouver le plus proche voisin selon le regroupement du schema linauto
        in : 
            tableLinauto : string : nom de la table dans le chema linauto a uinterroger
            distance : integer : distance au dela de laquelle on en echerche plus de voisin
        out : 
            dfPPV : dataframe du plud proche voisin, sur la base de gdfGeoloc
        """  
        self.insert_bdd(nomConnBddOtv, 'public', f'cpt_angouleme_{self.annee}',self.gdfGeoloc, if_exists='replace' )  
        rqt=f"""WITH 
                cpteur_base AS (
                SELECT DISTINCT ON (c.id_comptag) c.id_comptag id_comptag_bdd, t.gid, st_distance(c.geom, t.geom)
                 FROM comptage.compteur c left JOIN linauto.{tableLinauto} t ON st_dwithin(c.geom, t.geom, {distance})
                 ORDER BY c.id_comptag, st_distance(c.geom, t.geom))
                SELECT DISTINCT ON (c.id_comptag) 
                                    c.*, t.gid, st_distance(c.geom, t.geom),
                                    c2.id_comptag_bdd
                 FROM public.cpt_angouleme_{self.annee} c left JOIN linauto.{tableLinauto} t ON st_dwithin(c.geom, t.geom, {distance})
                                               LEFT JOIN cpteur_base c2 ON t.gid=c2.gid
                 ORDER BY c.id_comptag, st_distance(c.geom, t.geom)"""
        with ct.ConnexionBdd(nomConnBddOtv) as c : 
            dfPPV=gp.read_postgis(rqt, c.sqlAlchemyConn, crs='epsg:2154').merge(self.dfIndicAgrege, on='id_comptag', how='left'
                            ).merge(periodeDepuisHoraire(self.dfHoraire.assign(annee=self.annee)), how='left', on='id_comptag')
            dfPPV=O.gp_changer_nom_geom(dfPPV, 'geometrie')
        return dfPPV
    
    def ajouterFichierConcatener(self):
        """
        remplacer la reference au fichier de base par la ref aux deux fichier pour chaque sens 
        et ajouter le nombre de sens qui en découle
        """
        dfFichiersConcat=self.dfHoraire.groupby('id_comptag').fichier.apply( lambda x : ', '.join([os.path.basename(e) for e in set(x)]))
        self.dfHoraire=self.dfHoraire.drop('fichier', axis=1, errors='ignore').merge(dfFichiersConcat.reset_index(), on='id_comptag')
        self.dfIndicAgrege=self.dfIndicAgrege.drop('fichier', axis=1, errors='ignore').merge(dfFichiersConcat.reset_index(), on='id_comptag')
        self.dfIndicAgrege['sens_cpt']=self.dfIndicAgrege.fichier.apply(lambda x: 'double sens' if len(x.split(',')) > 1 else 'sens unique')

    def isolerComptagesDoublons(self,dfPPV ):
        """
        obtenir la liste des comptages situes sur un mm troncon homogene de trafic
        in : 
            dfPPV : donnees des compteurs, avec info supp, issue de plusProcheVoisinBddRegroupe
        out : 
            numpy array avec les id_comptages (cree automatiquementdans le cadre des donnees, pas ceux de la bdd) situes sur des troncon homogene equivalent
        """
        return dfPPV.loc[(dfPPV.duplicated(['gid', 'indicateur'], keep=False)) & (~dfPPV.gid.isna())]   
    
    def cptSsDblEtSsGeom(self, dfPPV, dicoAssoc, gererErreur=False):
        """
        récupérer la df des comptages sans doublons, mais avec les geometries
        in : 
            dfPPV : donnees des compteurs, avec info supp, issue de plusProcheVoisinBddRegroupe
            gererErreur : booleen pour prendre en compte ou non l'erreur qui bloque si pluseiurs pt sur une section homogene de trafic. pour la premeire passe laisse sur False
                          si erreur sur les sections gomgene passer à true
            DicoAssoc : dico des id_comptag creer a partir des donnees, enclé: les id_comptag des cpt ref, en value une liste des id_comptag des cpt assoc
        out : 
            dfPPVCptRef : df des comptages sans doublon pour insertion dans bdd
            dfPPVCptAssoc : df des comptages associes a retravailler avant insertion
        """
        dfPPVFiltreAconserver=dfPPV.loc[~dfPPV.id_comptag.isin([v for e in dicoAssoc.values() for v in e]) & (dfPPV.indicateur=='tmja') & (~dfPPV.gid.isna())]
        if not gererErreur : 
            if not (dfPPVFiltreAconserver.gid.value_counts()==1).all() : 
                raise ComptageMultipleSurTronconHomogeneError(list(dfPPVFiltreAconserver.loc[dfPPVFiltreAconserver.gid.isin(dfPPVFiltreAconserver.gid.value_counts()!=1)].gid.unique()))
        dfPPVCptRef=dfPPV.loc[dfPPV.id_comptag.isin(dfPPVFiltreAconserver.id_comptag.tolist()+dfPPV.loc[dfPPV.gid.isna()].id_comptag.tolist())].copy()
        dfPPVCptAssoc=dfPPV.loc[dfPPV.id_comptag.isin([v for e in dicoAssoc.values() for v in e])].merge(
                pd.DataFrame.from_dict(dicoAssoc, orient='index').reset_index().rename(columns={'index':'cpt_ref', 0:'cpt_assoc'}), left_on='id_comptag', right_on='cpt_assoc')
        #attentionaux doublons dus aux fichiers en double
        dfPPVCptRef.drop_duplicates(['indicateur','valeur', 'periode'], inplace=True)
        dfPPVCptAssoc.drop_duplicates(['indicateur','valeur', 'periode'], inplace=True)
        return dfPPVCptRef, dfPPVCptAssoc 
    
    def transfererAssocSupp(self, dfPPVCptRef, dfPPVCptAssoc, listCptAssoc):
        """
        fonction de transfert de donées entre la table de référence e la table des comptages Assoc, pour des transfert a posteriori de la 
        fonction cptSsDblEtSsGeom
        in : 
            listCptAssoc : list d'id_comptag a transférer
            dfPPVCptAssoc : df Des comptages associé issue de cptSsDblEtSsGeom()
            dfPPVCptRef : df des comptage sde références, issue de cptSsDblEtSsGeom()
        out : 
            dfPPVCptAssoc : df d'entrée avec les cpt a asocier en plus
            dfPPVCptRef : df d'entrée sans les cpt a ssocier
        """
        dfPPVCptAssoc=pd.concat([dfPPVCptAssoc, dfPPVCptRef.loc[dfPPVCptRef.id_comptag.isin(listCptAssoc)]])
        dfPPVCptRef=dfPPVCptRef.drop(dfPPVCptRef.loc[dfPPVCptRef.id_comptag.isin(listCptAssoc)].index)
        return dfPPVCptRef, dfPPVCptAssoc
        
    
    def insererCompteur(self, dfPPVSsDbl, CptAForcer=None, schema='comptage', table='compteur'):
        """
        Preparer les donnes liees au compteur et les inserer (si besoin) dans la bdd
        in : 
            CptAForcer : list d'id_comptag que l'on veut inclure dans la Bdd, de force, mm si un autre id_comptag est proche
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
        """
        #creer les compteurs a inserer dans la bdd
        dfCptAcreer=dfPPVSsDbl.loc[(dfPPVSsDbl.id_comptag_bdd.isna()) | (dfPPVSsDbl.id_comptag.isin(CptAForcer))]
        if not dfCptAcreer.empty :
            compteur=dfCptAcreer[[e for e in dfCptAcreer.columns if e in attBddCompteur]].drop_duplicates().copy()
            #inserer
            self.insert_bdd(schema, table, compteur)
        else : 
            print('pas de nouveau compteur a inserer')
            return
        
    def insererComptage(self, dfPPVSsDbl, CptAForcer=None, schema='comptage', table='comptage'):
        """
        Preparer les donnes liees au comptages et les inserer dans la bdd
        in : 
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
        """
        dfPPVSsDblTemp=dfPPVSsDbl.copy()
        dfPPVSsDblTemp['id_comptag_final']=dfPPVSsDblTemp.apply(lambda x : x.id_comptag_bdd if not pd.isnull(x.id_comptag_bdd) 
                                                                and x.id_comptag not in CptAForcer  else x.id_comptag, axis=1)
        comptage=dfPPVSsDblTemp[['id_comptag_final', 'annee', 'periode']].assign(src='gestionnaire', type_veh='tv/pl').rename(columns={'id_comptag_final':'id_comptag'}).drop_duplicates()
        #inserer
        self.insert_bdd(schema, table,comptage)
        return
        
    def insererIndics(self, dfPPVSsDbl,CptAForcer=None, schema='comptage', tableAgrege='indic_agrege', tableHoraire='indic_horaire'):
        """
        Preparer les donnes liees au indicateurs(horaires, agrege) et les inserer dans la bdd
        in : 
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
        """
        dfPPVSsDblTemp=dfPPVSsDbl.copy()
        dfPPVSsDblTemp['id_comptag_final']=dfPPVSsDblTemp.apply(lambda x : x.id_comptag_bdd if not pd.isnull(x.id_comptag_bdd) and x.id_comptag not in CptAForcer
                                                                else x.id_comptag, axis=1)
        indicAgrege=dfPPVSsDblTemp.merge(self.recupererIdUniqComptage(dfPPVSsDblTemp.id_comptag_final.tolist(), self.annee)
                             , left_on=['id_comptag_final', 'annee'], right_on=['id_comptag', 'annee'])[['id_comptag_uniq', 'indicateur', 'valeur', 'fichier']]
        dfHoraireSsDbl=self.dfHoraire.merge(dfPPVSsDblTemp[['id_comptag', 'id_comptag_final']], on='id_comptag')
        indicHoraire=dfHoraireSsDbl.assign(annee=self.annee, 
            indicateur=dfHoraireSsDbl.indicateur.str.upper()).merge(self.recupererIdUniqComptage(
            dfHoraireSsDbl.id_comptag_final.tolist(), self.annee).rename(columns={'id_comptag':'id_comptag_final'}), on=['id_comptag_final', 'annee']).drop([
            'sens_cpt', 'note_manuelle_qualite', 'obs_qualite', 'annee', 'id_comptag','id_comptag_final', 'obs_supl'], axis=1, errors='ignore')
        self.insert_bdd(schema, tableAgrege,indicAgrege)
        self.insert_bdd(schema, tableHoraire,indicHoraire)
        return
    
    def insererDatas(self, dfPPVSsDbl, CptAForcer=None, schema='comptage', tableComptag='comptage',
                     tableCpteur='compteur', tableAgrege='indic_agrege', tableHoraire='indic_horaire'):
        """
        preparer et inserere les donnees de compteurs, comptages et indicateurs
        in : 
            dfPPVSsDbl : donnees issues de cptSsDblEtSsGeom()
            CptAForcer : list d'id_comptag que l'on veut inclure dans la Bdd, de force, mm si un autre id_comptag est proche. creation des compteurs
        """
        self.insererCompteur(dfPPVSsDbl, CptAForcer, schema, tableCpteur)
        self.insererComptage(dfPPVSsDbl, CptAForcer, schema, tableComptag)
        self.insererIndics(dfPPVSsDbl, CptAForcer, schema, tableAgrege, tableHoraire)
        
    def creerComptageAssoc(self, dfPPVCptAssoc):    
        """
        creer la table des comptage du schema comptage_assoc
        in  : 
            dfPPVCptAssoc : df des comptages associes issu de cptSsDblEtSsGeom
        out : 
            dfAssoc : df au format de la table en bdd
        """
        dfPPVCptAssocInterne=dfPPVCptAssoc.copy()
        dfPPVCptAssocInterne['id_comptag_final']=dfPPVCptAssocInterne.apply(lambda x : x.id_comptag_bdd if not pd.isnull(x.id_comptag_bdd) else x.cpt_ref, axis=1)
        #trouver son id_comptag_uniq
        dfAssoc=dfPPVCptAssocInterne.merge(self.recupererIdUniqComptage(dfPPVCptAssocInterne.id_comptag_final.tolist(), self.annee), 
                                      left_on='id_comptag_final', right_on='id_comptag')
        #ajouter les colonnes manquantes
        dfAssocInsert=dfAssoc.assign(src='gestionnaire', type_veh='tv/pl', obs=None).drop_duplicates(['id_comptag_uniq', 'periode'])
        dfAssocInsert['rang']=dfAssocInsert.assign(toto=1).groupby('id_comptag_uniq').toto.rank(method='first')
        #inserer les comptages
        #self.insert_bdd(schemaComptageAssoc, tableComptage, dfAssocInsert[['id_comptag_uniq', 'periode', 'src', 'type_veh', 'rang', 'obs']].rename(columns={'id_comptag_uniq':'id_cptag_ref'}))
        return dfAssoc
    
    def creerIndicsAssoc(self, dfAssoc):
        """
        ATTENTION : 
        creer les tables en bdd des indcis_agreges et indic_horaires pour les comptages associes
        in : 
            dfAssoc : df issu de creerComptageAssoc
        out : 
            dfIndicAgregeAssoc : df integree a la bdd et sortie pour ifo
            dfHoraireassoc : df integree a la bdd et sortie pour ifo
        """
        dfIndicAgregeAssoc=dfAssoc.merge(self.recupererIdUniqComptageAssoc(dfAssoc.id_comptag_uniq.tolist()).rename(
            columns={'id_comptag_uniq':'id_comptag_uniq_assoc'}), left_on='id_comptag_uniq', right_on='id_cptag_ref')[['id_comptag_uniq_assoc', 'indicateur', 'valeur', 'fichier']]
        dfHoraireassoc=self.dfHoraire.loc[self.dfHoraire.fichier.isin(dfIndicAgregeAssoc.fichier.tolist())].merge(dfIndicAgregeAssoc[['id_comptag_uniq_assoc', 'fichier']], on='fichier').drop(
            ['id_comptag','obs_supl','sens_cpt', 'note_manuelle_qualite', 'obs_qualite'], axis=1, errors='ignore').rename(columns={'id_comptag_uniq_assoc':'id_comptag_uniq'})
        dfHoraireassoc['indicateur']=dfHoraireassoc.indicateur.str.upper()    
        self.insert_bdd(schemaComptageAssoc, tableIndicAgrege,dfIndicAgregeAssoc.rename(columns={'id_comptag_uniq_assoc':'id_comptag_uniq'}))
        self.insert_bdd(schemaComptageAssoc, tableIndicHoraire,dfHoraireassoc)
        return dfIndicAgregeAssoc, dfHoraireassoc
 
class Comptage_Dira(Comptage):
    """
    pour la DIRA on a besoinde connaitre le nom du fichierde syntheses tmja par section, 
    et le nom du dossier qui contient les fichiers excel de l'année complete, plus le nom du fihcier de Comptage 
    issu de la carto (cf dossier historique)
    on ajoute aussi le nom de table et de schema comptagede la Bdd pour pouvoir de suite se connecter et obtenir les infos 
    des points existants
    """ 
    def __init__(self,fichierTmjaParSection, dossierAnneeComplete,fichierComptageCarto, annee,table):
        self.fichierTmjaParSection = fichierTmjaParSection
        self.fichierComptageCarto = fichierComptageCarto
        self.dossierAnneeComplete = dossierAnneeComplete
        self.annee = annee
        self.dfCorrespExistant = self.correspIdsExistant(table)
        
    def ouvrirMiseEnFormeFichierTmja(self):
        """
        a partir du fichier, on limite les données, on ajoute le type de donnees, on supprime les NaN, 
        on prepare un id de jointure sans carac spéciaux, on change les noms de mois
        """
        dfBase = pd.read_excel(self.fichierTmjaParSection, engine='odf', sheet_name=f'TMJM_{self.annee}', na_values=['No data'])
        dfLimitee = dfBase.iloc[:,0:14]
        mois, colonnes = [m for m in dfLimitee.iloc[0,1:].tolist() if m in [a for v in dico_mois.values() for a in v]
                          ], ['nom']+dfLimitee.iloc[0,1:].tolist()
        dfLimitee.columns = colonnes
        dfFiltree = dfLimitee.loc[dfLimitee.apply(lambda x : not all([pd.isnull(x[m]) for m in mois]), axis=1)].iloc[2:,:].copy()
        # mise en forme données mensuelles
        dfFiltree.loc[~dfFiltree.nom.isna(), 'donnees_type'] = 'tmja'
        dfFiltree.loc[dfFiltree.nom.isna(), 'donnees_type'] = 'pc_pl'
        dfFiltree.nom = dfFiltree.nom.fillna(method='pad')
        dfFiltree.nom = dfFiltree.nom.apply(lambda x: re.sub('(é|è|ê)','e',re.sub('( |_|-)','',x.lower())).replace('ç','c'))
        dfFiltree.rename(columns={c: k for k, v in dico_mois.items() for c in dfFiltree.columns if c not in 
                                  ('nom','MJA','donnees_type') if c in v}, inplace=True)
        dfFiltree['nom_simple'] = dfFiltree.nom.apply(lambda x: x.split('mb')[0]+')' if 'mb' in x else x)
        return dfFiltree.drop_duplicates()
    
    def decomposeObsSupl(self, obs_supl, id_comptag, id_cpt) : 
        """
        decomposer le obs_supl des comptages existant de la base de donnees en 2 lignes renoyant 
        au mm id_comptag pour la mise a jour des donnees mesnuelle
        """
        if obs_supl:
            if 'voie : ' in obs_supl: 
                #print(id_comptag, id_cpt)
                voie = obs_supl.split(';')[0].split('voie : ')[1]
                nb_voie = len(voie.split(',')) 
                site = obs_supl.split(';')[3].split('nom_site : ')[1]
                if id_cpt: 
                    df = pd.DataFrame({'id_comptag': [id_comptag]*nb_voie, 'voie': voie.split(','),
                                       'site': [site]*nb_voie, 'id_cpt': id_cpt.split(',')})
                else : 
                    df = pd.DataFrame({'id_comptag': [id_comptag]*nb_voie, 'voie': voie.split(','),
                                       'site': [site] * nb_voie, 'id_cpt': ['NC'] * nb_voie})
                df['id_dira'] = df.apply(lambda x: re.sub('(é|è|ê)', 'e', re.sub('( |_|-)', '', x['site'].lower()))
                                         + '(' + re.sub('(é|è|ê)', 'e', re.sub('( |_)', '', x['voie'].lower())) + ')',
                                         axis=1).reset_index(drop=True)
                return df
        return pd.DataFrame([])
    
    def correspIdsExistant(self, table_cpt):
        """
        obtenir une df de correspndance entre l'id_comptag et les valeusr stockées dans obs_supl de la Bdd
        """
        self.existant = compteur_existant_bdd(table_cpt, schema='comptage', dep=False, type_poste=False, gest='DIRA')
        return pd.concat([self.decomposeObsSupl(b, c, d) for b, c, d in zip(
            self.existant.obs_supl.tolist(), self.existant.id_comptag.tolist(), self.existant.id_cpt.tolist())], axis=0, sort=False)
    
    def jointureExistantFichierTmja(self, nomAttrFichierDira='nom_simple'):
        """
        creer une df jointe entre les points de comptages existants dans la bdd et le fichier excel de comptage
        in : 
            nomAttrFichierDira : string : nom de l'attribut issu du fihcier source DIRA pour faire la jointure
        """
        #jointure avec donnees mensuelle
        dfFiltree = self.ouvrirMiseEnFormeFichierTmja()
        dfMens = dfFiltree.merge(self.dfCorrespExistant, left_on=nomAttrFichierDira, right_on='id_dira')
        return dfMens.drop_duplicates([k for k in dico_mois.keys()]  + ['nom'])  # car pb de doublons pas clair
    
    def verifValiditeMensuelle(self, dfMens) : 
        """
        pour tout les id_comptag il faut vérifier que si 2 sens sont présent, alors on a bien les données 2 
        sens pour calculer la valeur mensuelle
        """
        verifValidite = dfMens.merge(dfMens.groupby('id_comptag').id_dira.count().reset_index().rename(
            columns={'id_dira':'nb_lgn'}), on='id_comptag')
        for idComptag in verifValidite.id_comptag.tolist():
            test = verifValidite.loc[verifValidite['id_comptag'] == idComptag]
            if (test.nb_lgn == 4).all(): 
                #pour le tmja
                testTmja = test.loc[test['donnees_type'] == 'tmja']
                colInvalidesTmja = testTmja.columns.to_numpy()[testTmja.isnull().any().to_numpy()]
                verifValidite.loc[(verifValidite['donnees_type'] == 'tmja') & (verifValidite['id_comptag'] == idComptag),
                                  colInvalidesTmja] = -99
                #pour le pc_pl
                testTmja = test.loc[test['donnees_type'] == 'pc_pl']
                colInvalidesTmja = testTmja.columns.to_numpy()[testTmja.isnull().any().to_numpy()]
                verifValidite.loc[(verifValidite['donnees_type'] == 'pc_pl') & (verifValidite['id_comptag'] == idComptag
                                                                                ), colInvalidesTmja] = -99
            elif  test.nb_lgn.isin((1,2)).all(): 
                verifValidite.loc[verifValidite['id_comptag'] == idComptag] = verifValidite.loc[
                    verifValidite['id_comptag'] == idComptag].fillna(-99)
            else : 
                print(f'cas non prevu {test.nb_lgn}, {idComptag}')
        return verifValidite
    
    def MiseEnFormeMensuelleAnnuelle(self, verifValidite):
        #regroupement par id_comptag
        dfMensGrpTmja = verifValidite.loc[verifValidite['donnees_type'] == 'tmja'][['id_comptag','MJA'] + [
            k for k in dico_mois.keys()]].groupby('id_comptag').sum().assign(donnees_type='tmja')
        dfMensGrpTmja.MJA = dfMensGrpTmja.MJA.astype('int64')
        dfMensGrpPcpl = pd.concat([verifValidite.loc[verifValidite['donnees_type'] == 'pc_pl'][
            [k for k in dico_mois.keys()]+['MJA']].astype('float64')*100, verifValidite.loc[
                verifValidite['donnees_type'] =='pc_pl'][['id_comptag']]], axis=1, sort=False).groupby('id_comptag').mean().assign(
                    donnees_type='pc_pl')
        dfMensGrpPcpl = dfMensGrpPcpl.loc[dfMensGrpPcpl['MJA']<=99]#cas bizarre avec 100% PL
        dfMensGrp = pd.concat([dfMensGrpPcpl, dfMensGrpTmja], axis=0, sort=False).sort_values('id_comptag').reset_index()
        cond = dfMensGrp[[k for k in dico_mois.keys()]] > 0
        valMoins99 = dfMensGrp[[k for k in dico_mois.keys()]].where(cond, -99)
        dfMensGrpHomogene = pd.concat([valMoins99, dfMensGrp[['id_comptag','MJA','donnees_type']]], axis=1, sort=False
                                      ).set_index('id_comptag')
        return dfMensGrpHomogene
    
    def MiseEnFormeAnnuelle(self, dfMensGrp, idComptagAexclure=None):
        #mise à jour des TMJA null dans la Bdd
        idComptagAexclure = idComptagAexclure if isinstance(idComptagAexclure,list) else [idComptagAexclure,]
        dfMensGrpFiltre = dfMensGrp.loc[~dfMensGrp.index.isin(idComptagAexclure)].copy()
        self.df_attr = dfMensGrpFiltre.loc[dfMensGrpFiltre['donnees_type'] == 'tmja'][['MJA']].rename(columns={'MJA':'tmja'}).merge(
                        dfMensGrpFiltre.loc[dfMensGrpFiltre['donnees_type'] == 'pc_pl'][['MJA']].rename(
                            columns={'MJA':'pc_pl'}), how='left', right_index=True, left_index=True)
        self.df_attr['src'] = 'donnees_mensuelle tableur'
        self.df_attr['fichier'] = os.path.basename(self.fichierTmjaParSection)
        self.df_attr.reset_index(inplace=True)
        
    def MiseEnFormeMensuelle(self,dfMensGrp, idComptagAexclure=None):
        """
        Mise en forme des données avant intégration dans a base.
        in :
            idComptagAexclure string ou list de string : id_comptag à ne pas prendre en compte
        """
        self.df_attr_mens = dfMensGrp.fillna(-99).drop('MJA', axis=1)
        idComptagAexclure = idComptagAexclure if isinstance(idComptagAexclure,list) else [idComptagAexclure,]
        self.df_attr_mens = self.df_attr_mens.loc[~self.df_attr_mens.index.isin(idComptagAexclure)].copy()
        self.df_attr_mens.loc[self.df_attr_mens['donnees_type'] == 'pc_pl','janv':'dece'] = self.df_attr_mens.loc[
            self.df_attr_mens['donnees_type'] == 'pc_pl','janv':'dece'].applymap(lambda x : round(x,2))
        self.df_attr_mens.reset_index(inplace=True)
        self.df_attr_mens['annee'] = self.annee
        self.df_attr_mens['fichier'] = os.path.basename(self.fichierTmjaParSection)
        
        
    def enteteFeuilHoraire(self, fichier, feuille):
        site = ' '.join(fichier[feuille].iloc[2,0].split(' ')[5:])[1:-1]
        voie = fichier[feuille].iloc[3,0].split('Voie : ')[1]
        idDira = re.sub('ç','c',re.sub('(é|è|ê)', 'e', re.sub('( |_|-)', '', site.lower()))) + '(' + re.sub(
            'ç', 'c', re.sub('(é|è|ê)', 'e', re.sub('( |_)', '', voie.lower())))+')'
        return site, voie, idDira
    
    def miseEnFormeFeuille(self, fichier, feuille, nbHeure0Max=8, FlagHorsOTV=False):
        """
        transformer une feuille horaire en df
        in : 
            nbHeure0Max : nombre d'heure consecutive avec 0 véhicules
            FlagHorsOTV : drapeau pour outre-passer les verifs liées aux références OTV
        """
        voie, idDira = self.enteteFeuilHoraire(fichier, feuille)[1:3]
        flagVoie = False
        if not FlagHorsOTV:
            if not idDira in self.dfCorrespExistant.id_dira.tolist(): #il n'y a pas de correspondance avec un point de comptage
                if not voie in self.dfCorrespExistant.id_cpt.str.replace('_', ' ').tolist() :
                    raise self.BoucleNonConnueError(idDira)
                else : 
                    flagVoie = True
        colonnes = ['jour','type_veh'] + ['h'+c[:-1].replace('-','_') for c in fichier[feuille].iloc[4,:].values if c[-1] == 'h']
        df_horaire = fichier[feuille].iloc[5:fichier[feuille].loc[fichier[feuille].iloc[:,0] == 'Moyenne Jours'].index.values[0]-1, :26]
        df_horaire.columns = colonnes
        df_horaire.jour.fillna(method='pad', inplace=True)
        if not FlagHorsOTV:
            if flagVoie:
                df_horaire['id_dira'] = voie
                df_horaire = df_horaire.merge(self.dfCorrespExistant.assign(id_cpt=self.dfCorrespExistant.id_cpt.str.replace('_', ' ')),
                                             left_on='id_dira', right_on='id_cpt')
            else: 
                df_horaire['id_dira'] = idDira
                df_horaire = df_horaire.merge(self.dfCorrespExistant, on='id_dira')
        else : 
            df_horaire['id_dira'] = idDira
            df_horaire['voie'] = voie
        if any([(len(df_horaire.loc[(df_horaire.isna().any(axis=1)) & (df_horaire['type_veh']==t)]) > len(
            df_horaire.loc[df_horaire['type_veh'] == t]) / 2) or len(df_horaire.loc[df_horaire['type_veh'] == t].loc[
                df_horaire.loc[df_horaire['type_veh'] == t][df_horaire.loc[df_horaire['type_veh'] == t] == 0].count(axis=1) > nbHeure0Max]
            ) > len(df_horaire.loc[df_horaire['type_veh'] == t])/2 for t in ('PL','TV')]):
            raise self.FeuilleInvalideError(feuille, idDira)
        return df_horaire, idDira
    
    def verif2SensDispo(self, df):
        """
        vérifier que pour les id_comptages relatifs au section courante, chauqe date à bien des valeusr VL et PL pour chauqe sens
        """
        # on ne regarder que les id_comptag decrit par deux feuilles différentes
        dfJourValeur = df.loc[df.voie.apply(lambda x: re.sub('ç', 'c', re.sub('(é|è|ê)', 'e', re.sub('( |_)', '', x.lower()))))
                           .isin(denominationSens)].groupby(['jour','id_comptag']).agg(
                               {'voie': lambda x : x.count(), 'type_veh': lambda x: tuple(x)})
        # liste des jours avec moins de 4 valeurs (i.e pas 2 sens en VL et PL) et un type veh en (VL,PL)
        JoursPb = dfJourValeur.loc[(dfJourValeur['voie'] == 2) & (dfJourValeur.type_veh.isin([('VL', 'PL'), ('PL', 'VL')]))].index.tolist()
        dfFinale = df.loc[df.apply(lambda x: (x['jour'], x['id_comptag']) not in JoursPb, axis=1)].copy()
        return dfFinale
    
    def getIdEquiv(self, idDira):
        """
        pour un idDira, savoir si il y a un autre id a asocier pour faire un id_comptag
        """
        try:
            idDira2 = self.dfCorrespExistant.loc[self.dfCorrespExistant['id_comptag'].isin(
            self.dfCorrespExistant.loc[self.dfCorrespExistant['id_dira'] == idDira].id_comptag.tolist()) &
                                  (self.dfCorrespExistant['id_dira'] != idDira)].id_dira.values[0]
        except IndexError: 
            idDira2 = None
        return idDira2
    
    def miseEnFormeFichier(self, nomFichier, nbJoursValideMin=7, dicoModifVerifValiditeHoraire=None, FlagHorsOTV=False, nbHeure0Max=8):
        """
        transofrmer un fichier complet en df
        in : 
            nomFichier : nom du fichier, sans le chemin (deja stocke dan self.dossierAnneeComplete
            dicoModifVerifValiditeHoraire : un dictionnaire avec en clé le nom de fichier (sans le chemin) et en value le nombre d'heure
                                            continue avec une valeur 0 tolérable. Premte d'ajuster plus finement les vérif de validité
                                            pour les fichier de comptage à faible trafic. cf Donnees_horaires.verifValiditeFichier()
            FlagHorsOTV : drapeau pour outre-passer les verifs liées aux références OTV
        """
        fichier = pd.read_excel(os.path.join(self.dossierAnneeComplete,nomFichier), sheet_name=None)#A63_Ech24_Trimestre2_2019.xls
        dicoFeuille = {}
        listFail = [] # pour la gestion du cas où une des 2 feuilles de la section courantes est invalide, il faut pouvoir identifier l'autre et la virer
        for feuille in [k for k in fichier.keys() if k[:2]!="xx"]: 
            print(feuille)
            if not FlagHorsOTV:
                try :
                    print(feuille)
                    if not dicoModifVerifValiditeHoraire or nomFichier not in dicoModifVerifValiditeHoraire.keys():
                        df_horaire, idDira = self.miseEnFormeFeuille(fichier, feuille, nbHeure0Max=nbHeure0Max)
                    else:
                        df_horaire, idDira = self.miseEnFormeFeuille(fichier, feuille, dicoModifVerifValiditeHoraire[nomFichier],
                                                                     nbHeure0Max=nbHeure0Max)
                    if idDira in listFail: 
                        # print(f'feuille a jeter : {nomFichier}.{idDira}')
                        continue
                    # print(f'feuille en cours : {nomFichier}.{idDira}')
                    dicoFeuille[idDira] = df_horaire
                except self.BoucleNonConnueError:
                    continue
                except self.FeuilleInvalideError as e: 
                    # print(f'feuille a jeter : {nomFichier}.{e.idDira}')
                    # trouver l'autre feuille si traitee avant
                    idDira2 = self.getIdEquiv(e.idDira)
                    if idDira2:
                        if idDira2 in dicoFeuille.keys(): 
                            dicoFeuille[idDira2] = pd.DataFrame([])
                        else:
                            listFail.append(idDira2)
                    continue
            else:
                if not dicoModifVerifValiditeHoraire or nomFichier not in dicoModifVerifValiditeHoraire.keys():
                    df_horaire, idDira = self.miseEnFormeFeuille(fichier, feuille, FlagHorsOTV=FlagHorsOTV, nbHeure0Max=nbHeure0Max)
                else:
                    df_horaire, idDira = self.miseEnFormeFeuille(fichier, feuille, dicoModifVerifValiditeHoraire[nomFichier],
                                                                 FlagHorsOTV=FlagHorsOTV, nbHeure0Max=nbHeure0Max)
                dicoFeuille[idDira] = df_horaire.assign(id_comptag=idDira)
        # print([f.empty for f in dicoFeuille.values()])
        if not all([f.empty for f in dicoFeuille.values()]):
            dfHoraireFichier = pd.concat(dicoFeuille.values(), axis=0, sort=False)
        else: 
            raise self.AucuneBoucleConnueError(nomFichier) 
        if not dicoModifVerifValiditeHoraire or nomFichier not in dicoModifVerifValiditeHoraire.keys():
            dfHoraireFichierFiltre = verifValiditeFichier(dfHoraireFichier, NbHeures0Max=nbHeure0Max)[0]  # tri des feuilles sur le nb de valeusr NaN ou 0
        else:
            dfHoraireFichierFiltre = verifValiditeFichier(dfHoraireFichier, dicoModifVerifValiditeHoraire[nomFichier], NbHeures0Max=nbHeure0Max)[0]
        if not FlagHorsOTV:
            # test si fichier comprenant des id_comptag à 2 sens 
            dfFiltre = dfHoraireFichierFiltre.loc[dfHoraireFichierFiltre.id_comptag.isin(self.dfCorrespExistant.set_index('id_comptag').loc[
                self.dfCorrespExistant.id_comptag.value_counts() == 2].index.unique())]
            if not dfFiltre.empty: 
                # tri des donnes pour que les sections courantes ai bien une valeur VL et PL dans les deux sesn
                dfHoraireFichierFiltre = self.verif2SensDispo(dfHoraireFichierFiltre) 
                # verif que les deux sens sont concordant : 
                try:
                    comparer2Sens(dfHoraireFichierFiltre, attributSens='voie', attributIndicateur='type_veh')
                except SensAssymetriqueError as e:
                    dfHoraireFichierFiltre = dfHoraireFichierFiltre.loc[dfHoraireFichierFiltre.apply(
                        lambda x : (x['jour'],x['id_comptag']) not in zip(
                            e.dfCompInvalid.jour.tolist(), e.dfCompInvalid.id_comptag.tolist()), axis=1)].copy()
                except KeyError:
                    warnings.warn("attention, un des attribut nécéssaire à la vérification des sens est manquant")
            
        dfHoraireFichierFiltre = verifNbJoursValidDispo(dfHoraireFichierFiltre, nbJoursValideMin)[0]#tri sur id_comptag avec moins de 15 jours de donnees
        return dfHoraireFichierFiltre.assign(fichier=nomFichier)
    
    
    def concatTousFichierHoraire(self, dicoModifVerifValiditeHoraire=None):
        """
        rassemebler l'intégraliteé des données de fcihiers horaires dans une df
        in : 
            dicoModifVerifValiditeHoraire : cf miseEnFormeFichier()
        """
        listDf = []
        for i,fichier in enumerate(os.listdir(self.dossierAnneeComplete)) : 
            if fichier.endswith('.xls') : 
                try :
                    listDf.append(concatIndicateurFichierHoraire(self.miseEnFormeFichier(fichier, dicoModifVerifValiditeHoraire=dicoModifVerifValiditeHoraire)))
                except self.AucuneBoucleConnueError : 
                    print(i, fichier, '\n    aucune boucle dans ce fichier')
                except Exception as e:
                    print(fichier, f'ERREUR {e}')
            continue
        dfTousFichier = pd.concat(listDf, axis=0, sort=False)
        dblSupprime = dfTousFichier.loc[dfTousFichier.duplicated(['id_comptag', 'jour','type_veh'], keep=False)]
        dfTousFichier.drop_duplicates(['id_comptag', 'jour','type_veh'], inplace=True)
        idCptNonAffectes = self.dfCorrespExistant.loc[~self.dfCorrespExistant['id_comptag'].isin(
            dfTousFichier.id_comptag.tolist())].id_comptag.tolist()
        idCptNonAffectes.sort()
        return dfTousFichier, idCptNonAffectes, dblSupprime
    
    def cptCartoOuvrir(self):
        """
        les donnes de cmptage de la carto de la DIRA doiecvnt etre transferees dans un tableur sur la base du 
        modele 2020 dira_tmja_2020.ods
        on ouvre
        """
        return pd.read_excel(self.fichierComptageCarto, engine='odf')
    
    def cptCartoVerif(self, tableurCarto):
        """
        a partir de cptCartoOuvrir on verifie que les champs sont bien renseignes
        in : 
            tableurCarto dataframe issue de cptCartoOuvrir
        """       
        if not (tableurCarto.loc[tableurCarto[f'carto_{self.annee}']=='t'].reset_index(drop='True').equals(tableurCarto.loc[~tableurCarto.src.isna()].reset_index(drop='True')) and 
            tableurCarto.loc[tableurCarto.diffusable=='t'].reset_index(drop='True').equals(tableurCarto.loc[~tableurCarto.src.isna()].reset_index(drop='True')) ) : 
            raise ValueError(f"les valeurs dans les champs carto_{self.annee}, src, diffusable ne sont pas concordante")
        return tableurCarto
    
    def cptCartoForme(self, tableurCarto):
        """
        a partir des donnees ouvertes et verifiee, mise en forme des donnees
        in : 
            tableurCarto donnees issues de cptCartoVerif
        """
        #limitation du tableau aux donnees de la carte et calcul des indicateurs tout sens confondus
        donneesCarto=tableurCarto.loc[tableurCarto.carto_2020=='t']
        #verifier que toutes la valeurs sans TMJA on tbien la mention ''Donnees Non Disponibles''
        if (donneesCarto.loc[donneesCarto.tmja.isna()].obs.apply(lambda x : 'Donnees Non Disponibles' not in x if not pd.isnull(x) else True).any()
            or
            donneesCarto.loc[~donneesCarto.tmja.isna()].obs.apply(lambda x : 'Donnees Non Disponibles' in x if not pd.isnull(x) else False).any()): 
            raise ValueError(f"les lignes sans valeur de TMJA doivent contenir la mention 'Donnees Non Disponibles' (et inversement) dans le champs obs ")
        
        donneesAgregees=tableurCarto.loc[tableurCarto.carto_2020=='t'].groupby('id_comptag').agg({
            'tmja':'sum', 'pc_pl':'mean', 'src':lambda x : tuple(x)[0], 'obs':lambda x : tuple(x)[0] if not any(['Donnees Non Disponibles' in y if not pd.isnull(y) else False for y in tuple(x)]) else 'Donnees Non Disponibles' }).reset_index()
        #corriger toutes les donnees en DND
        donneesAgregees.loc[donneesAgregees.obs.apply(lambda x : 'Donnees Non Disponibles' in x if not pd.isnull(x) else False), 'tmja']=np.nan
        donneesAgregees.loc[donneesAgregees.obs.apply(lambda x : 'Donnees Non Disponibles' in x if not pd.isnull(x) else False), 'pc_pl']=np.nan
        #ajout du fichier, de l'année, etc..
        donneesAgregees['fichier']=os.path.basename(self.fichierComptageCarto)
        donneesAgregees['annee']=self.annee
        return donneesAgregees
    
    def cptCartoInsertUpdate(self):
        """
        classer les comptages carto selon qu'ils doivent mettre à jour les données issus du tableau annuel, ou creer un nouveau comptage
        out : 
            donneesAgregeesInsert : donnes issus des cpt carto agrege tout sens pour lesquelles il faut creer un nouveau comptage
            donneesAgregeesUpdate : donnes issus des cpt carto agrege tout sens pour lesquelles il faut MaJ un comptage existant
        """
        donneesAgregees=self.cptCartoForme(self.cptCartoVerif(self.cptCartoOuvrir()))
        donneesAgregeesUpdate=donneesAgregees.merge(self.recupererIdUniqComptage(donneesAgregees.id_comptag.tolist(), self.annee), on=('id_comptag'))
        donneesAgregeesInsert=donneesAgregees.loc[~donneesAgregees.id_comptag.isin(self.recupererIdUniqComptage(donneesAgregees.id_comptag.tolist(), self.annee).id_comptag.tolist())]
        return donneesAgregeesUpdate, donneesAgregeesInsert


    def update_bdd_Dira(self,schema, table, nullOnly=True):
        """
        metter à jour la table des comptage
        in : 
            nullOnly : boolean : True si la requet de Mise a jour ne concerne que les voies dont le tmja est null, false pour mettre 
                      toute les lignes à jour
        """
        valeurs_txt=self.creer_valeur_txt_update(self.df_attr,['id_comptag','tmja','pc_pl','src', 'fichier'])
        dico_attr_modif={f'tmja_{self.annee}':'tmja', f'pc_pl_{self.annee}':'pc_pl',f'src_{self.annee}':'src', 'fichier':'fichier'}
        if nullOnly : 
            self.update_bdd(schema, table, valeurs_txt,dico_attr_modif,'c.tmja_2019 is null')
        else : 
            self.update_bdd(schema, table, valeurs_txt,dico_attr_modif)
    
    class BoucleNonConnueError(Exception):
        """
        Exception levee si la reference d'une boucle n'est pas dans le fichier des id_existant de la Bdd
        """     
        def __init__(self, idDira):
            Exception.__init__(self, f'la boucle {idDira} n\'est pas dans le champs obs_supl d\'un des points de la bdd')
            
    class FeuilleInvalideError(Exception):
        """
        Exception levee si la feuille contient trop de valeur 0 ou NaN
        """     
        def __init__(self, feuille,idDira):
            self.idDira = idDira
            Exception.__init__(self, f'la feuille {feuille} contient trop de valuers NaN ou 0')
            
    class AucuneBoucleConnueError(Exception):
        """
        Exception levee si l'ensemble des boucles d'un fichier n'est pas dans le fichier des id_existant de la Bdd
        """     
        def __init__(self, nomFichier):
            Exception.__init__(self, f'le fichier {nomFichier} ne comporte aucune boucle dans le champs obs_supl d\'un des points de la bdd, ou toute les feuilles sont corrompues') 
    
class Comptage_Dirco(Comptage):
    """
    ne traite our le moemnt que de la partie MJM et horaire
    besoin du TMJA pour faire un premiere jointure entre les points de comptaget lees refereces des fichiers excele horaire
    """ 
    def __init__(self,fichierTmja, fichierTmjm, dossierHoraire, annee):
        self.fichierTmja = fichierTmja
        self.fichierTmjm = fichierTmjm
        self.dossierHoraire = dossierHoraire
        self.annee = annee
        self.dfSourceTmjm = pd.read_excel(self.fichierTmjm, skiprows=7, sheet_name=None)

    def miseEnFormeMJA(self):
        """
        ouvrir le fichier contanant les TMJA, le mettre en forme pour obtenir toute les infos pour verifier si les comptages
        existent et sinon les creer
        """
        #import des données
        dfTmjaBrute = pd.read_excel(self.fichierTmja, skiprows=3, na_values=['ND', '#VALEUR !'])
        #Mise en forme
        O.checkAttributsinDf(dfTmjaBrute, ['trafic cumulé', ' PL', 'Dépt.', 'Route', 'PR', 'Abc', 'Nom de la station' ] )
        dfTraficBrut = dfTmjaBrute.loc[(~dfTmjaBrute['trafic cumulé'].isna()) & (dfTmjaBrute['trafic cumulé']!='TV')
                                       ].copy().rename(columns={'trafic cumulé': 'tmja',' PL': 'pc_pl' })
        dfTraficBrut['id_comptag'] = dfTraficBrut.apply(
            lambda x: f"{str(int(x['Dépt.']))}-{x['Route'].replace('RN', 'N')}-{str(int(x['PR']))}+{str(int(x['Abc']))}", axis=1)
        dfTraficBrut['obs_supl'] = dfTraficBrut.apply(lambda x: f"site : {x['Nom de la station']}", axis=1)
        dfTraficBrut['fichier'] = os.path.basename(self.fichierTmja)
        dfTraficBrut['pc_pl'] = dfTraficBrut['pc_pl']*100
        dfTrafic = dfTraficBrut[['id_comptag', 'tmja', 'pc_pl', 'obs_supl', 'fichier']].copy()
        return dfTrafic

    def miseEnFormeMJM(self, feuille):
        """
        ouvrir et prepare le fichier des TMJM
        in : 
            feuille : df d'une feuille de self.dfSourceTmjm
        """
        dfTmjmDirco=feuille.iloc[1:].copy()
        dfTmjmDirco['Route']=dfTmjmDirco.Route.apply(lambda x : x if not isinstance(x,int) else f"RN{str(x)}")
        dfTmjmDirco['pr']=dfTmjmDirco.apply(lambda x : x['Route'].replace(' ','') if '+' in x['Route'] else np.NaN, axis=1).fillna(method='bfill')
        dfTmjmDirco['voie']=dfTmjmDirco.apply(lambda x : x['Route'].replace('RN' ,'N') if not '+' in x['Route'] else np.NaN, axis=1).fillna(method='ffill')
        dfTmjmDirco['dept']=dfTmjmDirco.apply(lambda x : x['Intitulé du PM'] if not '+' in x['Route'] 
                                              else np.NaN, axis=1).fillna(method='ffill')
        dfTmjmDirco['id_comptag']=dfTmjmDirco.apply(lambda x : f"{str(int(x['dept']))}-{x['voie']}-{str(x['pr'])}", axis=1)
        dfTmjmDirco['donnees_type']=dfTmjmDirco['TMJA  '].apply(lambda x : 'pc_pl' if x<1 else 'tmja')
        dfTmjmDirco.drop(['dept','Intitulé du PM','Route','Unnamed: 1','pr','voie','TMJA  '], axis=1, inplace=True)
        dfTmjmDirco.rename(columns={c:k for k, v in dico_mois.items()  for c in dfTmjmDirco.columns 
                                    if c not in ('id_comptag','donnees_type') if c in v}, inplace=True)
        return dfTmjmDirco
     
    def indicateurGlobalFeuilleMJM(self,dfTmjmDirco):   
        """
        regourper les resultats de miseEnFormeMJM() et mettre ne forme les Pc_pl
        """
        Agreg = dfTmjmDirco.merge(dfTmjmDirco.groupby('id_comptag').donnees_type.count(), on='id_comptag'
                ).rename(columns={'donnees_type_x': 'donnees_type', 'donnees_type_y' : 'nbLigne' })
        dfTv2SensCompte = Agreg.loc[(Agreg['donnees_type'] == 'tmja') & (Agreg.nbLigne == 4)][
            [e for e in Agreg.columns if e not in ('donnees_type', 'nbLigne')]].groupby(
                ['id_comptag']).sum().assign(donnees_type='tmja').replace(0, np.nan)
        dfTv1SensCompte = Agreg.loc[(Agreg['donnees_type'] == 'tmja') & (Agreg.nbLigne == 2)][
            [e for e in Agreg.columns if e not in ('donnees_type', 'nbLigne')]].groupby(
                ['id_comptag']).agg(lambda x: 2*x).assign(
            donnees_type = 'tmja', obs='1 seul sens *2')
        dfPl = Agreg.loc[Agreg['donnees_type'] == 'pc_pl'].fillna(-99)[
            [e for e in Agreg.columns if e not in ('donnees_type', 'nbLigne')]].groupby(
                ['id_comptag']).mean().applymap(lambda x: x * 100 if x > 0 else np.NaN).assign(donnees_type='pc_pl')
        return pd.concat([dfTv2SensCompte,dfTv1SensCompte, dfPl], axis=0, sort=False).sort_values(
            ['id_comptag','donnees_type'])

    
    def indicateurGlobalFichierMJM(self, cptARetirer):
        """
        concatener et mettre en forme les donnée de chaque feuille issues de indicateurGlobalFeuilleMJM
        in : 
            cptARetirer : string ou list de string : les id_comptag à ne pas prendre en compte
        """
        dfTouteFeuilles = pd.concat([self.indicateurGlobalFeuilleMJM(self.miseEnFormeMJM(self.dfSourceTmjm[k])) 
                          for k in self.dfSourceTmjm.keys()], axis=0, sort=False).assign(annee=self.annee).reset_index()
        dfTouteFeuilles = corresp_nom_id_comptag(dfTouteFeuilles)
        cptARetirer = cptARetirer if isinstance(cptARetirer,list) else [cptARetirer,]
        dfTouteFeuilles = dfTouteFeuilles.loc[~dfTouteFeuilles.id_comptag.isin(cptARetirer)]
        # mettre le TMJA en type integer
        dfTouteFeuilles.loc[dfTouteFeuilles.donnees_type == 'tmja', [k for k in dico_mois.keys()]
                   ] = dfTouteFeuilles.loc[dfTouteFeuilles.donnees_type == 'tmja', [k for k in dico_mois.keys()]
                                          ].applymap(lambda x: int(x) if not pd.isnull(x) else np.NaN)
        return dfTouteFeuilles
    
    def miseEnFormeFichierTmjaPourHoraire(self):
        """
        utiliser les données du fihier de TMJA pour faire un lien entre les données horaires et les points de
        cpt de la Bdd
        """
        def nomIdFichier(route,debStation,finStation,nomStation):
            if route == 'A20':
                return f"SRDT {route}_{debStation}_{finStation}"
            elif route == 'RN21':
                if nomStation in ('CROIX_BLANCHE','LA_COQUILLE'):
                    return "SRDT N021_" + nomStation.replace('_',' ').upper()
                elif nomStation in ('LAYRAC', 'TRELISSAC'):
                    return nomStation
                elif nomStation.upper() in ('ASTAFFORT', 'CANCON'):
                    return nomStation.upper()
                elif nomStation == 'NOTRE DAME':
                    return 'ND_SANILHAC'
                else:
                    return "SRDT N021_" + re.sub('( |_|\')','',nomStation).upper()
            elif route in ('RN1021','RN1113'):
                return nomStation.replace(' ','_').upper()
            elif route == 'RN141':
                return 'SRDT N141_'+nomStation.upper()
            elif route == 'RN520':
                return 'SRDT N520_PR'
            elif route == 'RN145':
                if nomStation == 'ST MAURICE':
                    return 'SAINT_MAURICE'
                elif nomStation == 'LA_RIBIERE':
                    return 'SRDT N145_LA RIBIERE'
                if nomStation == 'SAINT VAURY':
                    return 'SRDT N145_SAINT-VAURY'
                return 'SRDT N145' + '_' + nomStation.upper()
            elif route in ('RN147','RN149','RN249'):
                dico_corresp = {'AYRON': 'Station MBV86.O','EPANOURS': 'Station MBW87.J',
                                'FLEURE': 'Station MBV86.J', 'GENIEC': 'Station MBV86.H',
                                'LA CROIX CHAMBERT': 'Station MBX79.B', 'LE VINCOU': 'Station MBW87.K',
                                'LOUBLANDE': 'Station MBX79.A', 'MAISONNEUVE': 'Station MBW87.A',
                                'MOULINET': 'Station MBV86.A', 'MOULISMES': 'Station MBW87.I',
                                'MILETERIE CHU': 'MBV86.F', 'MARGOUILLET': 'MBV86.G', 'BUXEROLLES': 'MBV86.W', 
                                'LA FOLIE': 'MBV86.X', 'MIGNEAUXANCES':'MBV86.S',
                                'CHARDONCHAMP': 'MBV86.Y', 'MIGNALOUX': 'MBV86.V' }
                for k in dico_corresp.keys():
                    if nomStation.upper() == k.upper() :
                        return k.upper()

        dfFichierTmja = pd.read_excel(self.fichierTmja, skiprows=3, na_values=['ND', '#VALEUR !'])
        #dfFichierTmja.columns=dfFichierTmja.iloc[2].values
        dfFichierTmja = dfFichierTmja.loc[~dfFichierTmja['Nom de la station'].isna()].iloc[1:].copy()
        dfFichierTmja.columns = ['route', 'dept', 'pr', 'absc', 'nomStation', 'debStation', 'finStation',
                                 'xl93', 'yl93', 'voies', 'tmjaSens1', 'tmjaSens2', 'lcSens1', 'lcSens2',
                                 'tmja', 'pc_pl']
        dfFichierTmja['pr'] = dfFichierTmja.pr.apply(lambda x: int(x) if not pd.isnull(x) else -999)
        dfFichierTmja.absc = dfFichierTmja.absc.apply(lambda x: int(x) if not pd.isnull(x) else x)
        dfFichierTmja['idFichier'] = dfFichierTmja.apply(
            lambda x: nomIdFichier(x['route'], x['debStation'], x['finStation'], x['nomStation']), axis=1 )
        dfFichierTmja['id_comptag'] = dfFichierTmja.apply(
            lambda x: f"{int(x['dept'])}-{x['route'].replace('RN','N')}-{x['pr']}+{x['absc']}", axis=1)
        dfFichierTmja = corresp_nom_id_comptag(dfFichierTmja)
        self.existant = compteur_existant_bdd('compteur', gest='DIRCO')
        return dfFichierTmja
    
    def ouvrirFichierHoraire(self, fichier) : 
        """
        ouvrir un fichier et extraire les donnees brutes
        in : 
            fihcier : raw string de path du fichier
        """
        dfBrut = pd.read_excel(fichier, sheet_name=None)
        try:
            feuilleBrut = [a for a in dfBrut.keys() if 'xxcpt' in a][0]
        except IndexError: 
            raise ValueError(f'il n\'y a pas de feuille avec \'xxcpt\' dans le fichier {fichier}')
        return dfBrut[feuilleBrut]
        
    def caracTypeFichier(self, dfFichier):
        """
        identifier si la df fourni continet 1 seul station (format 2019) ou plusieurs stations (format 2020)
        in :
            dfFichier : dataframe de donnees brutes issue de ouvrirFichierHoraire
        """
        O.checkAttributsinDf(dfFichier, 'Station')
        if len(dfFichier.Station.unique())==1 : 
            return 'station unique'
        elif len(dfFichier.Station.unique())>1 :
            return 'stations multiples'
        else : 
            raise ValueError('pb sur le nombre de station dans df < 1')
    
    def jointureHoraireIdcpt(self, fichier, dfFichier, dfFichierTmja):
        """
        joindre les données contenuesdans les fichiers horaires avec un id_comptag de notre bdd
        attention, la structure des données horaires a changé, 2019 : un fichier par cpt, 2020 : plusieurs cpt sur un fichier
        in : 
            dfFichierTmja : issue de miseEnFormeFichierTmjaPourHoraire, df issu du fichier de tmja, avec ajourt de idFichier et idComptag
            fichier : raw string de path complet du fichier
            dfFichier, : dataframe du fichier de comptag horaire issue de caracTypeFichier
        """
        if 'A20' in fichier.upper():
            dfFichier['Station'] = dfFichier.Station.apply(lambda x: '_'.join(x.split('_')[:2]) + '_' + 
                                                           re.split('(_| |-)',x.split('_')[2])[0])    
        elif 'N21' in fichier.upper():
            dfFichier['Station'] = dfFichier.Station.apply(lambda x: x.upper())
        elif 'N141' in fichier.upper():
            dfFichier['Station'] = dfFichier.Station.apply(lambda x: x.split('_')[0] + '_' + 
                                                           x.split('_')[1].split('(')[0].strip().upper())
        elif 'N145' in fichier.upper():
            dfFichier['Station'] = dfFichier.Station.apply(lambda x: x.split('_')[0] + '_' +
                                                           x.split('_')[1].split('(')[0].replace('È','E').strip().upper()) 
        dfFichier['fichier'] = os.path.basename(fichier)
        return dfFichier.merge(dfFichierTmja.loc[~dfFichierTmja.idFichier.isna()][['idFichier','id_comptag']], left_on='Station', right_on='idFichier', how='left')
           
    def miseEnFormeHoraire(self, dfHoraireAgregee):
        """
        a partir des donnees horaires de base agregee tousFichierHoraires(), nettoyer, 
        filtrer les données hors section courante
        """
        def findTypeVeh(nature):
            if nature == 'Débit': 
                return 'TV'
            else: 
                return 'Autre' 
        dfDonnees = dfHoraireAgregee.loc[(~dfHoraireAgregee.mesure.isna())].copy()
        #convertir les temps au format bdd
        dfDonnees['heure'] = dfDonnees.apply(lambda x: f"h{str(x['sequence'].hour)}_{str(x['sequence'].hour+1)}", axis=1)
        dfSc = dfDonnees.loc[dfDonnees['Code Canal'].isin((0,1))].copy()
        dfSc['voie'] = dfSc['Code Canal'].apply(lambda x : 'sens 1' if x==0 else 'sens 2') # a ameliorer si besoin prise en compte Bretelle et autre
        #pour info : dfDonnees['Libellé groupage'].unique()#permetde voir que des libélés non désiré existent : 'Bretelle sortie',Bretelle Entrée'...
        dfSc['type_veh'] = dfSc['nature de mesure'].apply(lambda x : findTypeVeh(x))
        return dfSc
    
    def correspondanceStationIdcpt(self,dfSc):
        """
        obtenir le nom de station équivalent pour chaque id_comptag
        """
        #pour verif que 1 seule station par id_comptag : (dfSc.groupby('id_comptag')['Code Station'].unique().apply(lambda x : len(x))>1).unique()
        return dfSc.groupby('id_comptag')['Code Station'].unique().apply(lambda x : x[0]).reset_index().rename(columns={'Code Station':'codestation'})
    
    def horaireParSens(self,dfSc):
        """
        à partir des données de section courante, pivoter la table
        puis test sur la concordance des deux sens et check de la validite pour cahque sens
        """
        dfScParSens = dfSc[['id_comptag','jour','heure','mesure','type_veh','voie', 'fichier']
                           ].pivot(index=['id_comptag','jour','type_veh','voie', 'fichier'], 
                                   columns='heure',values='mesure').reset_index()
        dicoCptError = {}
        dicoCptok = {}
        dfScParSensValide, dfJourIdcptARetirer = verifValiditeFichier(dfScParSens)
        if not dfJourIdcptARetirer.empty:
            for idCpt in dfJourIdcptARetirer.id_comptag.unique():
                print(f'certains jour supprimes pour {idCpt}')
                dicoCptError[idCpt]={'typeError':'jours supprimes', 
                                     'df':dfJourIdcptARetirer.loc[dfJourIdcptARetirer.id_comptag == idCpt].copy()}
        for idCpt in dfScParSensValide.id_comptag.unique():
            try:
                dfComp = comparer2Sens(dfScParSensValide.loc[dfScParSensValide['id_comptag'] == idCpt],
                                     attributIndicateur='type_veh')[1]
                dicoCptok[idCpt] = dfComp
            except SensAssymetriqueError as e:
                print(f'erreur sens assymetriques sur {idCpt}')
                dicoCptError[idCpt] = {'typeError':'sens assymetriques', 'df': e.dfComp}
        dfScParSensOk = dfScParSensValide.loc[~dfScParSensValide['id_comptag'].isin(dicoCptError.keys())]
        return dfScParSensOk, dicoCptok, dicoCptError
        
    def horaire2SensFormatBdd(self,dfScParSensOk):
        """
        à partir des données de horaireParSens : somme des deux sens puis analyse de la validite des données
        """
        dfSc2Sens = concatIndicateurFichierHoraire(dfScParSensOk)
        #dfHoraire=dfSc2Sens.pivot(index=['id_comptag','jour','type_veh'], columns='heure',values='mesure').reset_index()
        dfHoraireValidNbJourOk,idCptInvalid,dfCptInvalid=verifNbJoursValidDispo(dfSc2Sens,10)
        return dfHoraireValidNbJourOk,idCptInvalid,dfCptInvalid
    
    def tousFichierHoraires(self, dfFichierTmja):
        """
        assembler tous les fichiers horaires en une seule df, en fonction du type de fichier (mono ou multi station)
        in : 
            dfFichierTmja : issue de miseEnFormeFichierTmjaPourHoraire, 
            df issu du fichier de tmja, avec ajourt de idFichier et idComptag
        """
        dicoCptOkTot, dicoCptErrorTot = {},{}
        listDfFinale = []
        for fichier in [os.path.join(root, f) for (root, dirs, files) in os.walk(self.dossierHoraire) for f in files]:
            print(fichier)
            dfFichier = self.ouvrirFichierHoraire(fichier)
            typeFichier = self.caracTypeFichier(dfFichier)
            if typeFichier == 'station unique':
                dfFichierHoraire = self.jointureHoraireIdcpt(fichier, dfFichier, dfFichierTmja)
                dfFichierHoraire = self.miseEnFormeHoraire(dfFichierHoraire)
                dfScParSensValide, dicoCptok, dicoCptError = self.horaireParSens(dfFichierHoraire)
                for k, v in dicoCptError.items():
                    dicoCptErrorTot[k] = v
                for k, v in dicoCptok.items():
                    dicoCptOkTot[k] = v
                dfFinale, idCptInvalid, dfCptInvalid = self.horaire2SensFormatBdd(dfScParSensValide)
                dfFinale['fichier'] = fichier
                if idCptInvalid:
                    for e in idCptInvalid:
                        print('fichier inavlid duree donees : {e}')
                        dicoCptError[e] = dfCptInvalid.loc[dfCptInvalid.id_comptag == e]
                listDfFinale.append(dfFinale.assign(fichier=os.path.basename(fichier)))
            else:
                for station in dfFichier.Station.unique():
                    print(f'station multi ; station : {station}, fichier : {fichier}')
                    dfFichierStationUniq = dfFichier.loc[dfFichier.Station == station].copy()
                    dfFichierStationUniq['Station'] = dfFichierStationUniq.Station.str.upper()
                    dfFichierHoraire = self.jointureHoraireIdcpt(fichier, dfFichierStationUniq, dfFichierTmja)
                    dfFichierHoraire = self.miseEnFormeHoraire(dfFichierHoraire)
                    try:
                        dfScParSensValide, dicoCptok, dicoCptError = self.horaireParSens(dfFichierHoraire.drop_duplicates(
                        ['id_comptag', 'jour', 'type_veh', 'voie', 'heure', 'mesure']))
                    except Exception as e:
                        dicoCptErrorTot[k] = e
                    for k, v in dicoCptError.items():
                        dicoCptErrorTot[k] = v
                    for k, v in dicoCptok.items():
                        dicoCptOkTot[k] = v
                    if dfScParSensValide.empty:
                        continue
                    dfFinale, idCptInvalid, dfCptInvalid = self.horaire2SensFormatBdd(dfScParSensValide)
                    if idCptInvalid:
                        for e in idCptInvalid:
                            print('fichier inavlid duree donees : {e}')
                            dicoCptError[e] = dfCptInvalid.loc[dfCptInvalid.id_comptag == e]
                    listDfFinale.append(dfFinale.assign(fichier=os.path.basename(fichier)))
        self.df_attr_horaire = pd.concat(listDfFinale, axis=0)
        return dicoCptOkTot,dicoCptErrorTot
     
     
class PasAssezMesureError(Exception):
    """
    Exception levee si le fichier comport emoins de 7 jours
    attribut : 
        nbjours : integer : nombrede jour de mesure
    """     
    def __init__(self, nbjours):
        Exception.__init__(self,f'le fichier comporte moins de 7 jours de mesures. Nb_jours: : {nbjours} ')   
  
        
class ComptageMultipleSurTronconHomogeneError(Exception):
    """
    Exception levee si sur un tronçon homogene 
    attribut : 
        refTroncHomo : list integer : identifiant unique du tronçn homogène
    """
    def __init__(self, refTroncHomo):
        Exception.__init__(self,f"le troncon homogene {','.join(refTroncHomo)} supporte plus que 1 comptage de reference")
        
        
class FormatError(Exception):
    """
    Exception levee si un format de donnnees source n'est pas prevu dans les codes 
    attribut : 
        formatDonnees : string : format de foichier ou donnees
    """
    def __init__(self, formatDonnees):
        Exception.__init__(self,f"le format {formatDonnees} n'est pas prevu dans le code")
        
        
